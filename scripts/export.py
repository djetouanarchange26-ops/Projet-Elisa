"""
EXPORT DES RÉSULTATS D'ANALYSE — PDF / Excel
==============================================
Feuille de route, point 4.2 : permettre à l'analyste de joindre le résultat
d'une analyse à un mémo de comité de crédit, sans repasser par un
copier-coller manuel depuis l'app.

CHOIX: fpdf2 (pur Python, pas de dépendance système type wkhtmltopdf/
weasyprint) pour le PDF, openpyxl (déjà utilisé pour lire corpus_cao_ifc.xlsx)
pour l'Excel — pas de nouvelle dépendance lourde.
FRAGILE: fpdf2 core fonts (Helvetica) sont limitées à Latin-1 — les résumés
LLM peuvent contenir des caractères hors Latin-1 (tirets typographiques,
guillemets courbes...). `_safe()` les remplace plutôt que de faire planter
l'export.
"""

import html
import io
from datetime import datetime

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Ponctuation Unicode courante (LLM, texte collé) sans équivalent Latin-1 —
# remplacée par un équivalent ASCII plutôt que par "?" (fallback encode()).
_UNICODE_REPLACEMENTS = {
    "—": "-", "–": "-",           # em/en dash
    "‘": "'", "’": "'",           # guillemets courbes simples
    "“": '"', "”": '"',           # guillemets courbes doubles
    "…": "...",                        # ellipse
}


def _safe(text):
    """Sanitize pour fpdf2 (Latin-1 core fonts) — remplace la ponctuation
    Unicode courante par un équivalent ASCII, puis remplace le reste
    (rare) plutôt que de planter l'export."""
    text = str(text)
    for char, repl in _UNICODE_REPLACEMENTS.items():
        text = text.replace(char, repl)
    return text.encode("latin-1", "replace").decode("latin-1")


def _unescape(text):
    """`display["signals"]`/`display["similar_cases"]` viennent de
    app._map_result_to_display, qui html.escape() certains champs pour le
    rendu Streamlit (unsafe_allow_html) — pas pertinent pour un export
    PDF/Excel, donc on inverse ici plutôt que de dupliquer la logique
    d'agrégation pour repartir des données brutes de `result`."""
    return html.unescape(text)


# ============================================================================
# PDF
# ============================================================================

def build_pdf_report(document_label, result, display):
    """Construit un rapport PDF à partir du résultat de analyze() (`result`)
    et de sa version mappée pour l'UI (`display`, voir app._map_result_to_display).

    Retourne des bytes, prêts pour st.download_button.
    """
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "ESG Risk Intelligence - Analysis Report", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 6, _safe(f"Document: {document_label}"), new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    # --- Risk Assessment Summary ---
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Risk Assessment Summary", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, _safe(f"Risk Grade: {display['risk_grade']} ({display['risk_label']}) — Score: {display['risk_score']}/100"), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(1)
    recommendation = result.get("recommendation") or "N/A"
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "Recommendation:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(0, 6, _safe(recommendation), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # --- Flag Scores ---
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Flag Scores", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    for label, score in display["flag_scores"].items():
        pdf.cell(0, 6, _safe(f"{label}: {score}/100"), new_x="LMARGIN", new_y="NEXT")
        for e in display.get("evidence_by_flag", {}).get(label, []):
            pdf.set_font("Helvetica", "I", 8)
            pdf.set_text_color(120, 120, 120)
            pdf.multi_cell(0, 4, _safe(f"  - {e['name']} ({e['score']}%): {_unescape(e['excerpt'])}"), new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Helvetica", "", 10)
    pdf.ln(3)

    # --- Detected Signals --- (result["detected_signals"], pas display["signals"]
    # qui est html.escape() par app._map_result_to_display pour le rendu Streamlit)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Detected Signals", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    if result["detected_signals"]:
        for s in result["detected_signals"][:8]:
            pdf.set_font("Helvetica", "B", 9)
            pdf.multi_cell(0, 5, _safe(f"Flag {s['source_flag']} — {s['signal'].upper()} ({s['occurrences']}x)"), new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(90, 90, 90)
            pdf.multi_cell(0, 5, _safe(s["evidence_excerpt"]), new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(0, 0, 0)
            pdf.ln(1)
    else:
        pdf.cell(0, 6, "No ESG signal detected in this document.", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # --- Historical Similar Cases ---
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Historical Similar Cases", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    if display["similar_cases"]:
        for case in display["similar_cases"]:
            pdf.set_font("Helvetica", "B", 9)
            pdf.multi_cell(0, 5, _safe(f"{case['name']} — Similarity: {case['similarity']:.0%} · {case['flag_type']}"), new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(90, 90, 90)
            pdf.multi_cell(0, 5, _safe(case["outcome"]), new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(130, 130, 130)
            pdf.multi_cell(0, 5, _safe(f'"{_unescape(case["excerpt"])}"'), new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(0, 0, 0)
            pdf.ln(1)
    else:
        pdf.cell(0, 6, "No similar historical case found.", new_x="LMARGIN", new_y="NEXT")

    # CHOIX: footer "en flux" (pas pdf.set_y(-15) pour l'ancrer en bas de
    # page) — set_y avec une valeur négative juste avant un cell() déclenche
    # un saut de page intempestif dans fpdf2 dès que le contenu au-dessus
    # est court, laissant une page 2 quasi vide avec juste le footer dessus.
    pdf.ln(6)
    pdf.set_font("Helvetica", "I", 7)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 10, "CA-CIB ESG Risk Intelligence - NLP Prototype MVP - Public IFC/CAO data")

    return bytes(pdf.output())


# ============================================================================
# EXCEL
# ============================================================================

_HEADER_FILL = PatternFill(start_color="006F4E", end_color="006F4E", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_header_row(ws, row, headers):
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL


def _autofit(ws, widths):
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def build_excel_report(document_label, result, display):
    """Construit un classeur Excel (3 feuilles : Summary, Detected Signals,
    Similar Cases) à partir du résultat de analyze(). Retourne des bytes."""
    wb = Workbook()

    # --- Feuille Summary ---
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ("Document", document_label),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Risk Grade", display["risk_grade"]),
        ("Risk Score (/100)", display["risk_score"]),
        ("Risk Label", display["risk_label"]),
        ("Recommendation", result.get("recommendation") or "N/A"),
        ("", ""),
    ]
    for label, value in rows:
        ws.append([label, value])
    ws.append(["Flag", "Score (/100)"])
    for cell in ws[ws.max_row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    for label, score in display["flag_scores"].items():
        ws.append([label, score])
    # CHANTIER SIMPLIFICATION PIPELINE (2026-08-08) : max_row=6, pas 7 --
    # une ligne de moins depuis le retrait de "Probability of ESG event".
    for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=1):
        row[0].font = Font(bold=True)
    _autofit(ws, [38, 60])

    # --- Feuille Detected Signals --- (result["detected_signals"], données
    # brutes non échappées — display["signals"] est html.escape() pour Streamlit)
    ws2 = wb.create_sheet("Detected Signals")
    _write_header_row(ws2, 1, ["Flag", "Signal", "Occurrences", "Confidence", "Evidence excerpt"])
    for s in result["detected_signals"]:
        ws2.append([
            f"Flag {s['source_flag']}", s["signal"], s["occurrences"],
            round(s["confidence"], 2), s["evidence_excerpt"],
        ])
    _autofit(ws2, [10, 22, 12, 12, 80])

    # --- Feuille Similar Cases ---
    ws3 = wb.create_sheet("Similar Cases")
    _write_header_row(ws3, 1, ["Project", "Similarity", "Flag Type", "Outcome", "Summary"])
    for case in display["similar_cases"]:
        ws3.append([case["name"], f"{case['similarity']:.0%}", case["flag_type"], case["outcome"], _unescape(case["excerpt"])])
    _autofit(ws3, [28, 12, 16, 30, 80])

    # --- Feuille Evidence by Flag --- (traçabilité : quels voisins FAISS
    # ont fait monter chaque flag_score, voir app._map_result_to_display)
    ws4 = wb.create_sheet("Evidence by Flag")
    _write_header_row(ws4, 1, ["Flag", "Project", "Score", "Excerpt"])
    for flag_label, evidence in display.get("evidence_by_flag", {}).items():
        for e in evidence:
            ws4.append([flag_label, e["name"], f"{e['score']}%", _unescape(e["excerpt"])])
    _autofit(ws4, [28, 28, 10, 80])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ============================================================================
# GRILLE ESG V4 — PDF / EXCEL (directive CC-V4-10)
# ============================================================================
# Export séparé de build_pdf_report/build_excel_report ci-dessus (pipeline
# ancien, 3 flags) — même conventions (fpdf2/_safe, openpyxl/_write_header_row/
# _autofit), pas de nouvelle dépendance. `result_v4` est le dict retourné par
# grid_analyze.analyze_grid() (cf. grid_result.build_grid_result pour le
# contrat exact).
#
# FRAGILE: plusieurs champs de result_v4 sont des clés TOUJOURS PRÉSENTES
# mais dont la valeur peut être None (mitigation_label, evidence_r["page"],
# evidence_a["page"]...) — un simple `dict.get(clé, défaut)` ne retombe sur
# `défaut` que si la clé est ABSENTE, pas si sa valeur est None. Traité en
# `dict.get(clé) or défaut` partout ici (même correction que grid_display.py,
# CC-V4-08/09). `reading_mode_label` est un champ de NIVEAU RÉSULTAT
# (result_v4["reading_mode_label"]), pas de result_v4["scoring"].

_COLOR_RGB_V4 = {
    "VERT": (46, 204, 113),
    "JAUNE": (241, 196, 15),
    "ORANGE": (230, 126, 34),
    "ROUGE": (231, 76, 60),
}

# SEUIL: largeurs calibrées pour tenir sur une page A4 portrait (~190mm
# utiles avec les marges par défaut de fpdf2) — 7 colonnes, somme = 183mm.
_GRID_V4_TABLE_COLS = [
    ("Sous-theme", 62), ("Code", 18), ("Statut", 20),
    ("Mitigation", 35), ("Penalite", 18), ("Gain", 15), ("Net", 15),
]


def _truncate(text, max_len):
    """Tronque `text` à `max_len` caractères avec une ellipse — évite qu'une
    cellule de tableau fpdf2 (largeur fixe, pas de wrap automatique dans
    cell()) déborde visuellement sur la colonne suivante."""
    text = str(text)
    return text if len(text) <= max_len else text[: max_len - 1].rstrip() + "…"


class _GridV4PDF(FPDF):
    """Sous-classe locale à la Grille V4 (pied de page "Confidentiel", cf.
    directive refonte layout 2026-08-20) — n'affecte PAS build_pdf_report
    (pipeline legacy, hors périmètre de cette directive), qui continue à
    utiliser FPDF nu."""

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(120, 120, 120)
        self.cell(
            0, 10,
            _safe(f"Confidentiel - Genere le {datetime.now().strftime('%Y-%m-%d %H:%M')}"),
            align="C",
        )
        self.set_text_color(0, 0, 0)


def _write_segments(pdf, segments, size=9, line_height=5, ln_after=6):
    """Ecrit une liste de segments (texte, gras) en continu via pdf.write()
    (contrairement a cell(), write() ne force pas de retour a la ligne et
    wrap automatiquement sur la largeur de page) — factorise le pattern
    phrase-narrative-avec-gras-partiel partage par _write_context_sentence
    et _write_question_sentence (fpdf2 n'interprete pas le markdown, d'ou
    la bascule manuelle de police par segment)."""
    for text, bold in segments:
        pdf.set_font("Helvetica", "B" if bold else "", size)
        pdf.write(line_height, _safe(text))
    if ln_after:
        pdf.ln(ln_after)
    pdf.set_font("Helvetica", "", size)


def _write_context_sentence(pdf, context):
    """Phrase narrative du contexte dossier (BLOC D), avec les valeurs
    discriminantes en gras — remplace l'ancienne ligne brute
    `A | Sensible | 400 | Mandated Lead Arranger` (cf. retour Elisa
    2026-08-20, "en faire un vrai rapport"). Miroir de
    grid_display._format_context_sentence, mais fpdf2 n'interprete pas le
    markdown : la phrase est decoupee en segments (texte, gras) ecrits en
    continu via pdf.write() (contrairement a cell(), write() ne force pas
    de retour a la ligne et wrap automatiquement sur la largeur de page)."""
    ep = context.get("ep_classification")
    sensitivity = context.get("sensitivity")
    amount = context.get("financing_amount")
    role = context.get("cacib_role")

    if not any([ep, sensitivity, amount, role]):
        return

    segments = [("Ce dossier est ", False)]
    if ep:
        segments += [("classe ", False), (f"Categorie {ep}", True), (" (Equator Principles)", False)]
    else:
        segments.append(("presente le contexte suivant", False))

    clauses = []
    if sensitivity:
        clauses.append([("un statut de portefeuille ", False), (sensitivity, True)])
    if amount:
        clauses.append([("un financement de ", False), (str(amount), True)])
    if role:
        clauses.append([("la banque intervenant en tant que ", False), (role, True)])
    if clauses:
        segments.append((", avec ", False))
        for i, clause in enumerate(clauses):
            if i > 0:
                segments.append((", ", False))
            segments += clause
    segments.append((".", False))

    _write_segments(pdf, segments)


def _question_status_sentence(q):
    """Phrase narrative statut/mitigation/impact d'une question, avec les
    valeurs decisionnelles en gras — remplace les lignes brutes
    "Statut : OUI" / "Mitigation : NON - forme insuffisante" /
    "Penalite : -15 | Gain : +0 | Net : -15" (cf. retour Elisa 2026-08-20,
    meme demarche que _write_context_sentence). N'est appelee que pour
    q["status"] in ("OUI", "INCONNU") — seuls statuts presents dans la
    section "Detail par question" (cf. build_grid_v4_pdf, filtre
    detail_questions) ; NA et NON n'y apparaissent jamais, pas geres ici.

    Pilotee par les donnees (mitigation_label + signe du gain), pas par
    une correspondance figee sur les 4 valeurs de
    grid_questions.MITIGATION_STATUTS — un libelle qui change de wording
    cote grid_questions.py n'oblige pas a toucher ce template.

    Retourne une liste de segments (texte, gras) pour _write_segments().
    """
    if q["status"] == "INCONNU":
        return [
            ("Les elements disponibles dans le document ne permettent pas de conclure sur ce point.", False),
            (" Aucune penalite appliquee.", False),
        ]

    # status == "OUI" (seul autre cas possible ici, cf. docstring)
    penalty = q["penalty"]
    gain = q.get("gain", 0)
    net = penalty + gain
    mitigation_label = q.get("mitigation_label")

    segments = [("Un risque a ete identifie ", False), ("(OUI)", True), (" sur ce point.", False)]

    if mitigation_label is None:
        segments.append((" Aucune mitigation n'a ete evaluee.", False))
    elif gain > 0:
        segments += [
            (" Une mitigation a ete mise en place et jugee ", False),
            (mitigation_label, True),
            (".", False),
        ]
    else:
        segments += [
            (" Une mitigation a ete recherchee mais jugee ", False),
            (mitigation_label, True),
            (" — aucun gain d'attenuation.", False),
        ]

    segments += [
        (" Penalite ", False), (f"{penalty} pts", True),
        (", gain ", False), (f"{gain:+d} pts", True),
        (", impact net ", False), (f"{net:+d} pts", True),
        (".", False),
    ]
    return segments


def _write_question_sentence(pdf, q):
    _write_segments(pdf, _question_status_sentence(q), ln_after=2)


def _write_project_metadata_sentence(pdf, metadata):
    """Phrase narrative sponsor/pays/secteur/client/type de projet (audit
    UI 2026-08-20, cf. grid_metadata.py) — memes segments (texte, gras)
    que _write_context_sentence. Retourne False (rien ecrit) si aucun
    champ n'a ete trouve — extraction LLM fail-open, jamais une valeur
    inventee ici."""
    sponsor = metadata.get("sponsor")
    country = metadata.get("country")
    sector = metadata.get("sector")
    client = metadata.get("client")
    project_type = metadata.get("project_type")

    if not any([sponsor, country, sector, client, project_type]):
        return False

    lead = f"un projet de type {project_type}" if project_type else "un projet"
    segments = [("Ce dossier concerne ", False), (lead, True)]

    clauses = []
    if sponsor:
        clauses.append([("porte par ", False), (sponsor, True)])
    if country:
        clauses.append([("pays ", False), (country, True)])
    if sector:
        clauses.append([("secteur ", False), (sector, True)])
    if client:
        clauses.append([("client ", False), (client, True)])
    if clauses:
        segments.append((", ", False))
        for i, clause in enumerate(clauses):
            if i > 0:
                segments.append((", ", False))
            segments += clause
    segments.append((".", False))

    _write_segments(pdf, segments)
    return True


# ============================================================================
# ANNEXE METHODOLOGIQUE (statique, identique pour tous les dossiers)
# ============================================================================
# CHOIX: contenu statique, PAS conditionnel au resultat du dossier analyse --
# l'objectif est l'auditabilite (un auditeur ou un comite de credit doit
# pouvoir comprendre le bareme, la regle du silence et les criteres de
# mitigation en lisant l'export seul, sans acces a l'application ni au
# code). Deux dossiers evalues avec le meme bareme doivent produire la
# meme annexe -- une variation deviendrait elle-meme une source de doute.
# Definie une seule fois ici (liste de tuples (label, explication),
# explication=None pour un titre de section), reutilisee telle quelle par
# build_grid_v4_pdf et build_grid_v4_excel : une seule source a maintenir
# si le bareme evolue (cf. grid_scoring.py/grid_questions.py, valeurs
# recopiees ici en texte -- pas d'import de constantes de calcul, cette
# annexe est de la documentation, pas une source de verite).
_METHODOLOGY_CONTENT = [
    ("1. Bareme de scoring", None),
    ("Categorie A (Bloqueurs)",
     "-25 pts par risque confirme (R=OUI), +5 pts si mitigation validee "
     "(statut \"OUI - prouvee\" uniquement, ratio 20%)."),
    ("Categorie B (Structurants)",
     "-15 pts par risque confirme, +3 pts si mitigation validee (ratio 20%)."),
    ("Calcul du score",
     "Score = max(0, 100 - somme des penalites + somme des attenuations), "
     "attenuation totale plafonnee a +20 pts."),
    ("Seuils de verdict",
     "VERT >= 75 / JAUNE 50-74 / ORANGE 25-49 / ROUGE < 25. Un score de 0 "
     "est signale \"ROUGE - Eliminatoire\" (score plancher, distinct du "
     "reste de la zone ROUGE)."),

    ("2. Regle du silence - interpretation de l'absence d'information", None),
    ("Question \"Evenement\" (ex. greve, action en justice, deplacement de population)",
     "Aucune mention -> NON (pas de mention = pas d'evenement survenu)."),
    ("Question \"Etat\" (ex. depassements de seuils de pollution)",
     "Aucune mention -> INCONNU (ne pas mesurer/documenter n'est pas une "
     "preuve d'absence du probleme)."),
    ("Question \"Absence-confirme\" (B.3.1 et B.3.2 uniquement)",
     "Aucune mention -> OUI penalisant. Ces deux questions portent sur "
     "l'absence d'un dispositif (donnees de reference socio-economiques / "
     "reporting ESG periodique) : ne rien trouver confirme directement "
     "cette absence."),

    ("3. Criteres de validation de la mitigation", None),
    ("Filtre temporel",
     "Seul l'accompli compte (mesure deja realisee) - une intention, un "
     "plan ou un engagement futur/conditionnel echoue a ce filtre."),
    ("Filtre de preuve",
     "Seules 3 formes sont acceptees : accord formel signe, investissement "
     "materiel receptionne, verification par un tiers independant. Un "
     "plan, une procedure, une politique, un recrutement ou une formation "
     "ne sont PAS des preuves, meme enonces au passe."),
    ("Gain de points",
     "Seul le statut \"OUI - prouvee\" (passe les deux filtres, aucune "
     "defaillance constatee) genere un gain. Le statut \"OUI - "
     "defaillante\" (mesure existante mais document etablissant qu'elle "
     "n'a pas produit son effet, ou qu'elle a ete interrompue) ne genere "
     "aucun gain."),
    ("Perimetre de la preuve",
     "Le fait ou la preuve doit concerner le projet lui-meme (la SPV) : "
     "un manquement ou une mesure attribues au preteur (IFC/SFI) ou a un "
     "tiers non imputable au projet ne comptent pas."),

    ("4. Limites de l'analyse", None),
    ("Un seul document a la fois",
     "L'evaluation ne croise pas plusieurs rapports entre eux."),
    ("Score non calcule par l'IA",
     "L'IA extrait les passages, le calcul du score est deterministe "
     "(arithmetique, aucun appel au modele de langage dans le calcul "
     "lui-meme)."),
    ("Sur-detection volontaire (fail-open)",
     "Un faux positif se verifie en 30 secondes a la lecture du document, "
     "un faux negatif peut couter beaucoup plus cher."),
]


def _write_methodology_annex(pdf):
    """Ecrit l'annexe methodologique en fin de PDF (derniere page) -- cf.
    _METHODOLOGY_CONTENT pour le detail du contenu et pourquoi il est
    statique (identique pour tous les dossiers)."""
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Annexe - Methodologie d'evaluation", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    for label, explanation in _METHODOLOGY_CONTENT:
        if explanation is None:
            pdf.ln(2)
            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(0, 8, _safe(label), new_x="LMARGIN", new_y="NEXT")
            continue
        _write_segments(pdf, [(f"{label} : ", True), (explanation, False)], size=9, line_height=5, ln_after=6)


def build_grid_v4_pdf(result_v4, project_name="", filename="esg_grid_v4.pdf"):
    """Génère un rapport PDF de la Grille V4 (12 questions), maquette Elisa
    2026-08-20.

    CHOIX: fpdf2 comme l'export existant — pas de nouvelle dépendance.
    `filename` n'est pas utilisé ici (le nom du fichier téléchargé est du
    ressort de st.download_button côté appelant, cf. app.py) — conservé
    dans la signature pour matcher le contrat attendu par CC-V4-10.

    Structure (maquette) :
    - Page 1 : en-tete "ESG Risk Intelligence", contexte du dossier, score
      (fond colore), synthese (si result_v4.get("synthesis"), sinon
      section absente), signaux identifies (OUI/INCONNU, verbatim tronque
      a 200 caracteres) — synthese placee juste apres le score, MEME ordre
      que grid_display._render_executive_summary (cf. _write_project_
      metadata_sentence pour le detail cote UI)
    - Page 2 : tableau des 12 questions + ligne de total
    - Pages suivantes : detail par question, UNIQUEMENT OUI et INCONNU
      (les NON ne sont plus detailles, cf. "Ce qu'il ne faut PAS faire")
    - Derniere page : annexe "Methodologie d'evaluation", texte STATIQUE
      identique pour tous les dossiers (cf. _METHODOLOGY_CONTENT et
      _write_methodology_annex)
    - Pied de page "Confidentiel - Genere le [date]" sur toutes les pages
      (cf. _GridV4PDF.footer)

    Retourne des bytes, prêts pour st.download_button.
    """
    scoring = result_v4["scoring"]
    score = scoring["score"]
    color = scoring["color"]
    mode_label = result_v4.get("reading_mode_label") or "-"
    context = result_v4.get("context") or {}
    project_metadata = result_v4.get("project_metadata") or {}
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    pdf = _GridV4PDF()
    pdf.set_auto_page_break(auto=True, margin=20)

    # --- Page 1 : en-tete + contexte + score + signaux + synthese ---
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 10, "ESG Risk Intelligence", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "B", 12)
    subtitle = f"Evaluation ESG - {project_name}" if project_name else "Evaluation ESG"
    pdf.cell(0, 8, _safe(subtitle), new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 6, f"Genere le : {now_str}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    # --- Metadonnees projet (audit UI 2026-08-20) : AVANT le contexte
    # bancaire (BLOC D) - "quel est ce projet" precede "comment la banque
    # le classe". Extraction LLM fail-open (grid_metadata.py), section
    # absente si rien trouve.
    if _write_project_metadata_sentence(pdf, project_metadata):
        pdf.ln(3)

    # --- Contexte du dossier (BLOC D, CC-V4-11) : AVANT le score, cf.
    # directive — saisie manuelle analyste, jamais extraite du document.
    if context:
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 6, "Contexte du dossier", new_x="LMARGIN", new_y="NEXT")
        _write_context_sentence(pdf, context)
        pdf.cell(0, 5, _safe(f"Mode : {mode_label}"), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

    # --- Score : fond colore (cf. directive, fpdf2 fill_color) ---
    rgb = _COLOR_RGB_V4.get(color, (100, 100, 100))
    pdf.set_fill_color(*rgb)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 20)
    score_label = f"Score : {score} / 100 - {color}"
    if scoring.get("saturation"):
        score_label += " (Eliminatoire)"
    pdf.cell(0, 14, _safe(score_label), new_x="LMARGIN", new_y="NEXT", fill=True, align="C")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)

    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Questions actives : {scoring['questions_active']} / 12", new_x="LMARGIN", new_y="NEXT")
    if scoring["questions_na"] > 0:
        pdf.cell(0, 6, f"Questions N/A : {scoring['questions_na']}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(
        0, 6, f"Plafond d'attenuation : {'Oui' if scoring['cap_applied'] else 'Non'}",
        new_x="LMARGIN", new_y="NEXT",
    )
    pdf.ln(4)

    # --- Synthese (optionnelle, absente du contrat result_v4 aujourd'hui) :
    # juste apres le score, AVANT les signaux — meme ordre que
    # grid_display._render_executive_summary (retour Elisa 2026-08-21 :
    # la synthese n'etait pas au meme endroit dans le PDF que dans l'UI).
    synthesis = result_v4.get("synthesis")
    if synthesis:
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, "Synthese", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.multi_cell(0, 6, _safe(synthesis), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    # --- Signaux identifies (KPI, OUI + INCONNU) ---
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Signaux identifies", new_x="LMARGIN", new_y="NEXT")

    signal_questions = [q for q in result_v4["questions"] if q["status"] in ("OUI", "INCONNU")]
    if not signal_questions:
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 6, "Aucun risque identifie.", new_x="LMARGIN", new_y="NEXT")
    else:
        for q in signal_questions:
            net = q["penalty"] + q.get("gain", 0)
            mit_label = q.get("mitigation_label")

            pdf.set_font("Helvetica", "B", 10)
            pdf.multi_cell(
                0, 6, _safe(f"- {q['sous_theme']} ({q['code']}) : {net:+d} pts"),
                new_x="LMARGIN", new_y="NEXT",
            )
            if mit_label:
                pdf.set_font("Helvetica", "", 9)
                pdf.cell(0, 5, _safe(f"  Mitigation : {mit_label}"), new_x="LMARGIN", new_y="NEXT")

            ev_r = q.get("evidence_r")
            verbatim = ev_r.get("passage") if ev_r else None
            if verbatim:
                # Tronque a 200 caracteres (maquette) — le verbatim complet
                # reste disponible dans le detail par question, page 3+.
                display = verbatim[:200] + "..." if len(verbatim) > 200 else verbatim
                pdf.set_font("Helvetica", "I", 9)
                pdf.set_text_color(90, 90, 90)
                pdf.multi_cell(0, 5, _safe(f'  "{display}"'), new_x="LMARGIN", new_y="NEXT")
                pdf.set_text_color(0, 0, 0)
            elif q["status"] == "INCONNU" or q.get("confidence_note"):
                # Directive "gestion INCONNU" (2026-08-20) : rien invente
                # quand aucun element n'a ete extrait. Etendu (2026-08-21)
                # au cas OUI sans verbatim (_SILENCE_CONFIRMS_ABSENCE,
                # B.3.1/B.3.2) : sans ca, une penalite s'affichait ici sans
                # aucune justification visible avant la page 3.
                no_element_note = q.get("confidence_note") or "Aucun element n'a ete trouve."
                pdf.set_font("Helvetica", "I", 9)
                pdf.set_text_color(90, 90, 90)
                pdf.cell(0, 5, _safe(f"  {no_element_note}"), new_x="LMARGIN", new_y="NEXT")
                pdf.set_text_color(0, 0, 0)
            pdf.ln(1)

    # --- Page 2 : tableau des 12 questions ---
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Grille d'evaluation - 12 questions", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(0, 111, 78)
    pdf.set_text_color(255, 255, 255)
    for header, width in _GRID_V4_TABLE_COLS:
        pdf.cell(width, 8, header, border=1, fill=True, align="C")
    pdf.ln()
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8)

    for q in result_v4["questions"]:
        gain = q.get("gain", 0)
        row = [
            _truncate(q["sous_theme"], 38),
            q["code"],
            q["status"],
            _truncate(q.get("mitigation_label") or "-", 22),
            str(q["penalty"]),
            f"{gain:+d}" if gain else "0",
            str(q["penalty"] + gain),
        ]
        for (_, width), value in zip(_GRID_V4_TABLE_COLS, row):
            pdf.cell(width, 7, _safe(value), border=1)
        pdf.ln()

    net_total = scoring["total_penalty"] + scoring["total_gain_capped"]
    pdf.set_font("Helvetica", "B", 8)
    total_row = [
        "", "", "", "TOTAL",
        str(scoring["total_penalty"]),
        f"{scoring['total_gain_capped']:+d}",
        str(net_total),
    ]
    for (_, width), value in zip(_GRID_V4_TABLE_COLS, total_row):
        pdf.cell(width, 7, _safe(value), border=1)
    pdf.ln()
    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(
        0, 6,
        f"Score = max(0, 100 + {scoring['total_penalty']} + {scoring['total_gain_capped']}) = {score}",
        new_x="LMARGIN", new_y="NEXT",
    )

    # --- Pages suivantes : détail par question — UNIQUEMENT OUI et INCONNU
    # (cf. directive "Ne PAS inclure les questions NON dans les pages de
    # détail du PDF"), contrairement à l'ancien filtre _question_has_detail
    # qui incluait aussi un NON attesté.
    detail_questions = [q for q in result_v4["questions"] if q["status"] in ("OUI", "INCONNU")]
    if detail_questions:
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "Detail par question", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        for q in detail_questions:
            pdf.set_font("Helvetica", "B", 11)
            pdf.multi_cell(0, 7, _safe(f"{q['sous_theme']} ({q['code']})"), new_x="LMARGIN", new_y="NEXT")
            _write_question_sentence(pdf, q)
            pdf.ln(1)

            ev_r = q.get("evidence_r")
            if ev_r and ev_r.get("passage"):
                pdf.set_font("Helvetica", "B", 9)
                pdf.cell(0, 5, f"Preuve de risque (page {ev_r.get('page') or '?'}) :", new_x="LMARGIN", new_y="NEXT")
                pdf.set_font("Helvetica", "I", 9)
                pdf.set_text_color(90, 90, 90)
                pdf.multi_cell(0, 5, _safe(f'"{ev_r["passage"]}"'), new_x="LMARGIN", new_y="NEXT")
                pdf.set_text_color(0, 0, 0)
                pdf.ln(1)

            ev_a = q.get("evidence_a")
            if ev_a and ev_a.get("verbatim_mesure"):
                pdf.set_font("Helvetica", "B", 9)
                pdf.cell(
                    0, 5, f"Preuve de mitigation (page {ev_a.get('page') or '?'}) :",
                    new_x="LMARGIN", new_y="NEXT",
                )
                pdf.set_font("Helvetica", "I", 9)
                pdf.set_text_color(60, 130, 70)
                pdf.multi_cell(0, 5, _safe(f'"{ev_a["verbatim_mesure"]}"'), new_x="LMARGIN", new_y="NEXT")
                pdf.set_text_color(0, 0, 0)
                pdf.ln(1)
            if ev_a and ev_a.get("verbatim_defaillance"):
                pdf.set_font("Helvetica", "B", 9)
                pdf.cell(0, 5, "Defaillance :", new_x="LMARGIN", new_y="NEXT")
                pdf.set_font("Helvetica", "I", 9)
                pdf.set_text_color(180, 130, 30)
                pdf.multi_cell(0, 5, _safe(f'"{ev_a["verbatim_defaillance"]}"'), new_x="LMARGIN", new_y="NEXT")
                pdf.set_text_color(0, 0, 0)
                pdf.ln(1)

            if q.get("confidence_note"):
                # Directive "gestion INCONNU" (2026-08-20) : une absence
                # totale de passage (silence_applied) n'est pas un "doute"
                # du LLM — pas de label, texte canonique affiché tel quel.
                if not q.get("silence_applied"):
                    pdf.set_font("Helvetica", "B", 9)
                    pdf.cell(0, 5, "Doute :", new_x="LMARGIN", new_y="NEXT")
                pdf.set_font("Helvetica", "I", 9)
                pdf.set_text_color(120, 120, 120)
                pdf.multi_cell(0, 5, _safe(q["confidence_note"]), new_x="LMARGIN", new_y="NEXT")
                pdf.set_text_color(0, 0, 0)
                pdf.ln(1)

            qualifying = q.get("qualifying")
            if qualifying:
                qual_items = [(k, v) for k, v in qualifying.items() if v and v is not True]
                if qual_items:
                    pdf.set_font("Helvetica", "B", 9)
                    pdf.cell(0, 5, "Champs qualifiants :", new_x="LMARGIN", new_y="NEXT")
                    pdf.set_font("Helvetica", "", 9)
                    for k, v in qual_items:
                        pdf.multi_cell(0, 5, _safe(f"  {k} : {v}"), new_x="LMARGIN", new_y="NEXT")

            pdf.ln(2)
            pdf.set_draw_color(220, 220, 220)
            pdf.line(pdf.get_x(), pdf.get_y(), pdf.get_x() + 190, pdf.get_y())
            pdf.set_draw_color(0, 0, 0)
            pdf.ln(3)

    # --- Derniere page : annexe methodologique (statique) ---
    _write_methodology_annex(pdf)

    return bytes(pdf.output())


# --- Excel : remplissages conditionnels sur la colonne Statut de la
# feuille "Grille" (maquette Elisa 2026-08-20) : OUI -> rouge clair,
# NON -> vert clair, INCONNU -> jaune clair. ---
_PENALTY_FILL_V4 = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
_GAIN_FILL_V4 = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
_INCONNU_FILL_V4 = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
_STATUS_FILL_V4 = {"OUI": _PENALTY_FILL_V4, "NON": _GAIN_FILL_V4, "INCONNU": _INCONNU_FILL_V4}
_WRAP_TOP_V4 = Alignment(wrap_text=True, vertical="top")


def _build_methodology_sheet(wb):
    """Ajoute la feuille "Methodologie" (derniere feuille du classeur) --
    cf. _METHODOLOGY_CONTENT pour le detail du contenu et pourquoi il est
    statique (identique pour tous les dossiers)."""
    ws = wb.create_sheet("Methodologie")
    _write_header_row(ws, 1, ["Regle", "Explication"])
    for label, explanation in _METHODOLOGY_CONTENT:
        if explanation is None:
            ws.append([label, ""])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
            continue
        ws.append([label, explanation])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=2).alignment = _WRAP_TOP_V4
    _autofit(ws, [40, 90])


def build_grid_v4_excel(result_v4, project_name="", filename="esg_grid_v4.xlsx"):
    """Génère un classeur Excel de la Grille V4 (12 questions), maquette
    Elisa 2026-08-20.

    CHOIX: openpyxl comme l'export existant — pas de nouvelle dépendance.
    `filename` non utilisé (cf. build_grid_v4_pdf ci-dessus).

    4 feuilles (maquette — remplace l'ancien découpage à 4 feuilles
    Synthese/Grille/Evidence/Qualifiants) :
    - Synthese : score, contexte, mode, métriques agrégées + tableau des
      signaux identifiés (OUI/INCONNU), verbatim TRONQUÉ à 200 caractères
    - Grille   : tableau des 12 questions, fond conditionnel par Statut
      (cf. _STATUS_FILL_V4)
    - Detail   : une ligne par question OUI/INCONNU, verbatims COMPLETS,
      jamais tronqués (cf. directive : "l'analyste veut pouvoir
      copier-coller"), colonnes larges + wrap_text
    - Methodologie : annexe statique (barème, règle du silence, critères
      de mitigation, limites) — identique pour tous les dossiers, cf.
      _METHODOLOGY_CONTENT et _build_methodology_sheet

    Retourne des bytes, prêts pour st.download_button.
    """
    scoring = result_v4["scoring"]
    context = result_v4.get("context") or {}
    project_metadata = result_v4.get("project_metadata") or {}
    wb = Workbook()

    # --- Feuille Synthese ---
    ws1 = wb.active
    ws1.title = "Synthese"

    ws1.append(["ESG Risk Intelligence - Evaluation ESG"])
    ws1.cell(row=ws1.max_row, column=1).font = Font(bold=True, size=14)
    ws1.append([project_name or "-", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws1.append([])

    # --- Metadonnees projet (audit UI 2026-08-20) : AVANT le contexte
    # bancaire - extraction LLM fail-open (grid_metadata.py), section
    # absente si rien trouve (jamais une valeur inventee).
    if any(project_metadata.get(k) for k in ("sponsor", "country", "sector", "client", "project_type")):
        ws1.append(["PROJET"])
        ws1.cell(row=ws1.max_row, column=1).font = Font(bold=True)
        _write_header_row(ws1, ws1.max_row + 1, ["Sponsor", "Pays", "Secteur", "Client", "Type de projet"])
        ws1.append([
            project_metadata.get("sponsor") or "-", project_metadata.get("country") or "-",
            project_metadata.get("sector") or "-", project_metadata.get("client") or "-",
            project_metadata.get("project_type") or "-",
        ])
        ws1.append([])

    ws1.append(["CONTEXTE DU DOSSIER"])
    ws1.cell(row=ws1.max_row, column=1).font = Font(bold=True)
    _write_header_row(ws1, ws1.max_row + 1, ["EP", "Sensibilite", "Montant", "Role"])
    ws1.append([
        context.get("ep_classification") or "-", context.get("sensitivity") or "-",
        context.get("financing_amount") or "-", context.get("cacib_role") or "-",
    ])
    ws1.append([])

    ws1.append([f"SCORE : {scoring['score']} / 100 - {scoring['color']}"])
    ws1.cell(row=ws1.max_row, column=1).font = Font(bold=True, size=13)
    ws1.append([f"Mode de lecture : {result_v4.get('reading_mode_label') or '-'}"])
    ws1.append([f"Questions actives : {scoring['questions_active']} / 12"])
    ws1.append([f"Plafond d'attenuation applique : {'Oui' if scoring['cap_applied'] else 'Non'}"])
    ws1.append([])

    ws1.append(["SIGNAUX IDENTIFIES"])
    ws1.cell(row=ws1.max_row, column=1).font = Font(bold=True)
    _write_header_row(ws1, ws1.max_row + 1, ["Code", "Sous-theme", "Penalite", "Verbatim (tronque)"])
    signal_questions = [q for q in result_v4["questions"] if q["status"] in ("OUI", "INCONNU")]
    if not signal_questions:
        ws1.append(["-", "Aucun risque identifie.", "-", "-"])
    else:
        for q in signal_questions:
            ev_r = q.get("evidence_r")
            verbatim = ev_r.get("passage") if ev_r else None
            if verbatim:
                display = (verbatim[:200] + "...") if len(verbatim) > 200 else verbatim
            else:
                # Cas OUI/INCONNU sans verbatim (silence R5, dont
                # _SILENCE_CONFIRMS_ABSENCE B.3.1/B.3.2) : confidence_note
                # plutot qu'un simple "-" qui ne justifie pas la penalite.
                display = q.get("confidence_note") or "-"
            ws1.append([q["code"], q["sous_theme"], q["penalty"], display])
    _autofit(ws1, [30, 30, 14, 70])

    # --- Feuille Grille ---
    ws2 = wb.create_sheet("Grille")
    _write_header_row(ws2, 1, ["Code", "Sous-theme", "Statut", "Mitigation", "Penalite", "Gain", "Net"])
    for q in result_v4["questions"]:
        gain = q.get("gain", 0)
        ws2.append([
            q["code"], q["sous_theme"], q["status"], q.get("mitigation_label") or "-",
            q["penalty"], gain, q["penalty"] + gain,
        ])
        fill = _STATUS_FILL_V4.get(q["status"])
        if fill:
            ws2.cell(row=ws2.max_row, column=3).fill = fill
    _autofit(ws2, [8, 34, 10, 24, 10, 8, 8])

    # --- Feuille Detail (OUI/INCONNU uniquement, verbatims complets) ---
    ws3 = wb.create_sheet("Detail")
    _write_header_row(ws3, 1, [
        "Code", "Statut", "Verbatim risque", "Page", "Verbatim mitigation",
        "Defaillance", "Confiance",
    ])
    for q in result_v4["questions"]:
        if q["status"] not in ("OUI", "INCONNU"):
            continue
        ev_r = q.get("evidence_r") or {}
        ev_a = q.get("evidence_a") or {}
        ws3.append([
            q["code"], q["status"],
            ev_r.get("passage") or "-", ev_r.get("page") or "-",
            ev_a.get("verbatim_mesure") or "-", ev_a.get("verbatim_defaillance") or "-",
            q.get("confidence_note") or "-",
        ])
        row_idx = ws3.max_row
        # Verbatims complets (pas tronqués, cf. directive) — wrap_text pour
        # rester lisible malgré la longueur.
        for col in (3, 5, 6, 7):
            ws3.cell(row=row_idx, column=col).alignment = _WRAP_TOP_V4
    _autofit(ws3, [10, 10, 60, 8, 60, 60, 40])

    # --- Feuille Methodologie (statique, cf. _METHODOLOGY_CONTENT) ---
    _build_methodology_sheet(wb)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
