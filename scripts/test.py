"""
SUITE DE TESTS MVP
==================
3 niveaux :
  1. Tests unitaires      — chaque fichier/composant isolé
  2. Tests d'intégration  — pipeline bout en bout
  3. Tests de cohérence   — les résultats ont-ils du sens métier ?

Usage :
    cd scripts/
    python tests.py              # lance tout
    python tests.py --unit       # unitaires seulement
    python tests.py --integ      # intégration seulement
    python tests.py --business   # cohérence métier seulement
"""

import sys
import time
import numpy as np
import pandas as pd
from pathlib import Path

# --- Chemins ---
BASE = Path(__file__).resolve().parent.parent
CHUNKS_PATH      = BASE / "data/processed/chunks.csv"
EMBEDDING_PATH   = BASE / "models/embeddings.npy"
METADATA_PATH    = BASE / "models/chunks_metadata.pkl"
FAISS_INDEX_PATH = BASE / "models/faiss_index.bin"
ANNOTATIONS_PATH = BASE / "data/raw/corpus_cao_ifc.xlsx"

# Compteurs globaux
_passed = 0
_failed = 0
_warnings = 0


def _test(name, condition, msg_fail="", warning_only=False):
    global _passed, _failed, _warnings
    if condition:
        print(f"  ✅ {name}")
        _passed += 1
    elif warning_only:
        print(f"  ⚠️  {name} — {msg_fail}")
        _warnings += 1
    else:
        print(f"  ❌ {name} — {msg_fail}")
        _failed += 1


# ============================================================================
# 1. TESTS UNITAIRES
# ============================================================================

def test_unit():
    print("\n--- 1. Tests unitaires ---\n")

    # 1.1 Fichiers nécessaires
    _test("chunks.csv existe",      CHUNKS_PATH.exists(),      f"Manquant : {CHUNKS_PATH}")
    _test("embeddings.npy existe",  EMBEDDING_PATH.exists(),   f"Manquant : {EMBEDDING_PATH}")
    _test("metadata.pkl existe",    METADATA_PATH.exists(),    f"Manquant : {METADATA_PATH}")
    _test("faiss_index.bin existe", FAISS_INDEX_PATH.exists(), f"Manquant : {FAISS_INDEX_PATH}")
    _test("annotations existe",     ANNOTATIONS_PATH.exists(), f"Manquant : {ANNOTATIONS_PATH}")

    # 1.2 Intégrité chunks.csv
    if CHUNKS_PATH.exists():
        chunks = pd.read_csv(CHUNKS_PATH)
        required_cols = ["project_name", "text", "chunk_id"]
        _test("chunks.csv colonnes requises",
              all(c in chunks.columns for c in required_cols),
              f"Colonnes : {list(chunks.columns)}")
        _test("chunks.csv > 100 lignes",
              len(chunks) > 100,
              f"Seulement {len(chunks)} lignes",
              warning_only=True)
        _test("Pas de texte vide",
              chunks["text"].notna().all() and (chunks["text"].str.len() > 10).all(),
              "Chunks avec texte vide ou trop court")

    # 1.3 Cohérence embeddings / chunks
    if EMBEDDING_PATH.exists() and CHUNKS_PATH.exists():
        embs = np.load(EMBEDDING_PATH)
        n_chunks = len(pd.read_csv(CHUNKS_PATH))
        _test("Embeddings shape cohérente",
              embs.shape[0] == n_chunks,
              f"embeddings={embs.shape[0]} vs chunks={n_chunks}")
        _test("Dimension = 768",
              embs.shape[1] == 768,
              f"Dimension = {embs.shape[1]}")
        norms = np.linalg.norm(embs[:10], axis=1)
        _test("Normalisés L2",
              np.allclose(norms, 1.0, atol=0.01),
              f"Norme moyenne = {norms.mean():.3f}")

    # 1.4 Grade ESG par règle (CHANTIER SIMPLIFICATION PIPELINE, 2026-08-08 —
    # remplace le Cox, retiré : coefficients flag2/flag3 non significatifs
    # sur 46 projets, décalage train/serve non résolu, cf. checklist.md).
    from model import compute_grade, DEFAULT_RISK_THRESHOLDS

    _test("DEFAULT_RISK_THRESHOLDS strictement croissant",
          all(DEFAULT_RISK_THRESHOLDS[i][0] < DEFAULT_RISK_THRESHOLDS[i + 1][0]
              for i in range(len(DEFAULT_RISK_THRESHOLDS) - 1)),
          f"Seuils : {DEFAULT_RISK_THRESHOLDS}")

    _low = compute_grade({"flag1_community": 5, "flag2_pollution": 5, "flag3_compliance": 5})
    _test("Score bas -> grade D (Vigilance)",
          _low["risk_grade"] == "D" and _low["risk_label"] == "Vigilance",
          f"Obtenu : {_low}")

    _high = compute_grade({"flag1_community": 90, "flag2_pollution": 5, "flag3_compliance": 5})
    _test("Score haut sur un seul flag -> grade A (Escalade), max() pas moyenne",
          _high["risk_grade"] == "A" and _high["risk_label"] == "Escalade",
          f"Obtenu : {_high}")

    # 1.5 llm_backend — abstraction backend LLM (2026-08-07, checklist.md).
    # Pas d'appel réseau réel ici (pas de dépendance à une clé API/à Ollama
    # actif) : on monkey-patche llm_backend._dispatch pour vérifier le
    # contrat (fail-open, fallback, plafond de longueur) indépendamment du
    # backend réellement configuré.
    import config
    import llm_backend

    _orig_dispatch = llm_backend._dispatch

    def _always_fails(*a, **kw):
        raise RuntimeError("backend simulé injoignable")

    llm_backend._dispatch = _always_fails
    try:
        result = llm_backend.call_llm("test", config_key="confirm_risk", timeout=1)
        _test("llm_backend.call_llm fail-open (backend injoignable -> None, jamais d'exception)",
              result is None,
              f"Résultat inattendu : {result!r}")
    finally:
        llm_backend._dispatch = _orig_dispatch

    _orig_backend, _orig_fallback = config.LLM_BACKEND, config.LLM_FALLBACK
    config.LLM_BACKEND, config.LLM_FALLBACK = "together", "ollama"
    calls = []

    def _fail_then_succeed(backend, prompt, model, options, timeout, response_format=None):
        calls.append(backend)
        if backend == "together":
            raise RuntimeError("primaire simulé injoignable")
        return "reponse de secours"

    llm_backend._dispatch = _fail_then_succeed
    try:
        result = llm_backend.call_llm("test", config_key="confirm_risk", timeout=1)
        _test("llm_backend.call_llm bascule sur LLM_FALLBACK si le backend principal échoue",
              result == "reponse de secours" and calls == ["together", "ollama"],
              f"Résultat={result!r}, backends appelés={calls}")
    finally:
        llm_backend._dispatch = _orig_dispatch
        config.LLM_BACKEND, config.LLM_FALLBACK = _orig_backend, _orig_fallback

    opts_capped = llm_backend._resolve_options("confirm_risk", 0.0)
    _test("_resolve_options applique le plafond num_predict pour un config_key connu",
          opts_capped["max_tokens"] == config.OLLAMA_CONFIGS["confirm_risk"]["num_predict"],
          f"max_tokens={opts_capped['max_tokens']}")

    opts_uncapped = llm_backend._resolve_options(None, 0.0)
    _test("_resolve_options NE plafonne PAS quand config_key=None (invariant Pass 2 deep_analysis)",
          opts_uncapped["max_tokens"] is None,
          f"max_tokens={opts_uncapped['max_tokens']}")

    # 1.6 Grille ESG V4 (grid_questions.py, directive CC-V4-01) — structure
    # seule, pas de scoring/LLM ici. Pipeline pas encore activé
    # (config.GRID_V4_ENABLED = False), ce test ne dépend d'aucun réseau.
    test_grid_questions_v4()

    # 1.7 Moteur de scoring V3 (grid_scoring.py, directive CC-02) et
    # 1.8 format de résultat V3 (grid_result.py, directive CC-03) — même
    # raison : purement déterministe, pas de dépendance réseau/LLM.
    test_grid_scoring()
    test_grid_result()

    # 1.7b Scoring V4 (grid_scoring.py, directive CC-V4-02) — 4 zones de
    # couleur, plafond partagé (dormant depuis CC-V4-11), B.3.1 schéma
    # standard, mode de lecture par type de document. Purement déterministe.
    test_grid_scoring_v4()

    # 1.8b Format de résultat V4 (grid_result.py, directive CC-V4-03) —
    # document_type/reading_mode, garde-fou mitigation par a_condition,
    # champ qualifying. Purement déterministe.
    test_grid_result_v4()

    # 1.9 Prompts V4 complets (grid_prompts.py, directive CC-V4-05) —
    # construction/parsing de strings uniquement, aucun appel LLM.
    test_grid_prompts_v4()

    # 1.10 Mitigation à 4 statuts (grid_scoring.py/grid_questions.py,
    # directive CC-07, passe d'annotation CBG) — purement déterministe.
    test_mitigation_4_statuts()

    # 1.11-1.13 Silence événement/état, verrou B.2.1→B.2.2, flag attesté
    # (directive CC-08, passe CBG) — purement déterministe.
    test_silence_evenement_etat()
    test_verrou_b21_b22()
    test_flag_atteste()

    # 1.14 Exclusions parseur ESAP / section plaintes IFC (grid_sections.py,
    # directive CC-09, confirmée par CC-V4-04) — purement déterministe, pas de LLM.
    test_esap_exclusion()
    test_ifc_complaints_exclusion()
    test_no_sections_detected()

    # 1.14d Marquage best-effort des couches temporelles (grid_sections.py,
    # addition V4, directive CC-V4-04, bonus non bloquant) — déterministe.
    test_temporal_layer_marking()

    # 1.15 Orchestrateur Grille V4 (grid_analyze.py, directive CC-V4-06) —
    # backend LLM monkey-patché (même idiome que 1.5), pas d'appel réseau réel.
    test_grid_analyze_v4()

    # 1.23 Synthèse finale post-scoring (directive "évolutions pipeline
    # ESG/risk", 2026-08-20) — prompt (déterministe) puis intégration
    # (backend LLM monkey-patché).
    test_grid_synthesis_prompt()
    test_grid_analyze_synthesis()

    # 1.16 Calibration V4 sur les 4 dossiers annotés (directive CC-V4-07) —
    # purement déterministe (grid_scoring.py seul), pas de LLM.
    test_calibration_cbg()
    test_calibration_mundra()
    test_calibration_aysha()
    test_calibration_indorama()
    test_calibration_ordering()
    test_calibration_shared_cap_mundra()
    test_calibration_b31_inverted()

    # 1.17 Export PDF/Excel de la Grille V4 (directive CC-V4-10) — purement
    # déterministe (fpdf2/openpyxl), pas de LLM.
    test_grid_v4_export()

    # 1.17a Phrase narrative statut/mitigation dans le PDF (retour Elisa
    # 2026-08-20) — purement déterministe, pas de LLM.
    test_pdf_question_sentence()

    # 1.17b Groupes de l'Executive Risk Summary (audit UI 2026-08-20) —
    # purement déterministe, pas de LLM, pas de Streamlit.
    test_grid_display_executive_summary()

    # 1.17c Conservation des analyses sur disque (analysis_store.py) —
    # purement déterministe, isolé dans un dossier temporaire.
    test_analysis_store()

    # 1.18 Détection automatique du type de document (R11, grid_doctype.py)
    # — backend LLM monkey-patché, pas d'appel réseau réel.
    test_grid_doctype()

    # 1.18b Métadonnées projet (grid_metadata.py, audit UI 2026-08-20) —
    # backend LLM monkey-patché, pas d'appel réseau réel.
    test_grid_metadata()

    # 1.19 R10 — catégories étendues du filtre de sujet (AMBIGU, INDIRECT)
    # — string-only sur les prompts + parsing, pas de LLM.
    test_grid_prompts_r10_extended()

    # 1.20 Articulation B.2.3/B.4.1 — CODE MORT DORMANT depuis CC-V4-11
    # (codes retirés) — string-only sur les prompts, pas de LLM.
    test_grid_prompts_b23_b41_articulation()

    # 1.21 R7 — exige que le verbatim de défaillance porte sur la MÊME
    # mesure que le verbatim de mesure (statut 4) — string-only, pas de LLM.
    test_grid_prompts_r7_evidence_linkage()

    # 1.22 Dispatch pipeline unique (pipeline_dispatch.py, refactor pipeline
    # unique) — garde-fou anti double-exécution, monkey-patché.
    test_pipeline_dispatch_single_execution()


def test_grid_prompts_r10_extended():
    """CC-V4-12 : R10 (filtre de sujet) vit désormais dans la Passe 1
    (EXTRACTION, champ `subject`), plus dans un appel unique — 5
    catégories SPV/LENDER/SUBSTITUTION/AMBIGUOUS/INDIRECT (extension
    volontaire par rapport au schéma à 3 valeurs de la directive CC-V4-12,
    cf. docstring grid_prompts.py). String-only, aucun appel LLM."""
    print("\n--- 1.19 R10 — filtre de sujet étendu (grid_prompts.py, CC-V4-12) ---\n")

    import grid_prompts

    prompt_a21 = grid_prompts.get_extraction_prompt("A.2.1", ["chunk"])
    _test("R10 : catégorie SUBSTITUTION documentée dans le prompt d'extraction", "SUBSTITUTION" in prompt_a21)
    _test("R10 : catégorie AMBIGUOUS documentée dans le prompt d'extraction", "AMBIGUOUS" in prompt_a21)
    _test("R10 : catégorie INDIRECT documentée dans le prompt d'extraction", "INDIRECT" in prompt_a21)
    _test("R10 : consigne explicite de ne pas attribuer par défaut à la SPV",
          "JAMAIS" in prompt_a21 and "SPV" in prompt_a21 and "par défaut" in prompt_a21)

    # Parsing (Passe 1) : les 5 catégories sont acceptées telles quelles,
    # sans réécriture — sauf repli AMBIGUOUS si absente/invalide (jamais
    # SPV par défaut, cf. parse_extraction_response).
    import json
    resp_ambiguous = json.dumps({"code": "A.2.1", "found": True, "verbatim": "x", "page": 5, "subject": "AMBIGUOUS", "brief": "test"})
    parsed_ambiguous = grid_prompts.parse_extraction_response(resp_ambiguous)
    _test("Parsing extraction : subject='AMBIGUOUS' accepté", parsed_ambiguous["subject"] == "AMBIGUOUS",
          f"Obtenu : {parsed_ambiguous['subject']!r}")

    resp_indirect = json.dumps({"code": "A.2.1", "found": True, "verbatim": "x", "page": 5, "subject": "INDIRECT", "brief": "test"})
    parsed_indirect = grid_prompts.parse_extraction_response(resp_indirect)
    _test("Parsing extraction : subject='INDIRECT' accepté", parsed_indirect["subject"] == "INDIRECT",
          f"Obtenu : {parsed_indirect['subject']!r}")

    resp_bad_subject = json.dumps({"code": "A.2.1", "found": True, "verbatim": "x", "page": 5, "subject": "AUTRE", "brief": "test"})
    parsed_bad = grid_prompts.parse_extraction_response(resp_bad_subject)
    _test("Parsing extraction : subject invalide -> repli AMBIGUOUS (JAMAIS SPV, cf. R10)",
          parsed_bad["subject"] == "AMBIGUOUS", f"Obtenu : {parsed_bad['subject']!r}")


def test_grid_prompts_b23_b41_articulation():
    """CC-V4-11 : B.2.3/B.4.1 n'existent plus dans grid_questions.py
    (absents de la Maquette Vierge). CC-V4-12 : grid_prompts.py réécrit en
    entier (architecture 2 passes) — R2bis (_ARTICULATION_B23_B41) n'a
    même plus été reportée dans la réécriture (elle ne servait qu'à
    articuler B.2.3/B.4.1, tous deux inexistants ; texte original
    récupérable dans l'historique git, commit 3fbf8b5, si un besoin
    similaire réapparaissait). Ce test vérifie seulement que les codes
    retirés renvoient None des deux passes. String-only, aucun appel LLM."""
    print("\n--- 1.20 Codes retirés B.2.3/B.4.1 (grid_prompts.py, CC-V4-11/12) ---\n")

    import grid_prompts

    for code in ("B.2.3", "B.4.1"):
        _test(f"{code} retiré : get_extraction_prompt('{code}', ...) -> None",
              grid_prompts.get_extraction_prompt(code, ["chunk"]) is None)
        _test(f"{code} retiré : get_qualification_prompt('{code}', ...) -> None",
              grid_prompts.get_qualification_prompt(code, "verbatim", "SPV") is None)


def test_grid_prompts_r7_evidence_linkage():
    """Tests dédiés au renforcement R7 (grid_prompts.py) : le verbatim de
    défaillance (statut 4, OUI_DEFAILLANTE) doit porter explicitement sur
    la MÊME mesure que le verbatim de mesure — une simple conjonction
    concessive (however/although) isolée, ou une défaillance d'une AUTRE
    mesure, ne suffit plus : rester au statut 3 en cas de doute sur le
    lien. CC-V4-12 : cette règle vit désormais dans la Passe 2
    (QUALIFICATION), pas dans l'appel unique. String-only, aucun appel LLM
    (le garde-fou structurel — double verbatim non vide — existe déjà côté
    grid_scoring.py et n'est pas modifié ici)."""
    print("\n--- 1.21 R7 — lien mesure/défaillance (grid_prompts.py, CC-V4-12) ---\n")

    import grid_prompts

    prompt_b11 = grid_prompts.get_qualification_prompt("B.1.1", "verbatim de test", "SPV")
    _test("R7 : exige que la défaillance porte sur la MÊME mesure",
          "MÊME" in prompt_b11 and "mesure" in prompt_b11.lower())
    _test("R7 : cas 'lien non explicite -> rester au statut 3' documenté",
          "OUI_PROUVEE" in prompt_b11 and "OUI_DEFAILLANTE" in prompt_b11)


def test_pipeline_dispatch_single_execution():
    """GARDE-FOU : une analyse ne doit déclencher EXACTEMENT qu'un seul
    pipeline (V4 xor legacy), jamais les deux, jamais aucun silencieusement
    (refactor pipeline unique). Compteurs d'appels sur les deux points
    d'entrée (analyze.analyze / grid_analyze.analyze_grid_auto), monkey-
    patchés — pas d'appel réseau réel."""
    print("\n--- 1.22 Dispatch pipeline unique (pipeline_dispatch.py) ---\n")

    import config
    import pipeline_dispatch
    import analyze as analyze_module
    import grid_analyze

    calls = {"legacy": 0, "v4": 0}

    def _fake_analyze(text, risk_thresholds=None, k=15, document_label="x"):
        calls["legacy"] += 1
        return {"flag_scores": {}, "prediction": {}, "similar_passages": [],
                "detected_signals": [], "signal_spans": [], "recommendation": None,
                "deep_analysis": {"enabled": False}, "processing_time_s": 0}

    def _fake_analyze_grid_auto(chunks, full_text, na_modules=None, document_type_override=None, context=None):
        calls["v4"] += 1
        return {"grid_version": "V4", "document_type": 1, "questions": [],
                "scoring": {"score": 100}, "document_type_detection": {"source": "manuel"},
                "context": context}

    _orig_analyze = analyze_module.analyze
    _orig_analyze_grid_auto = grid_analyze.analyze_grid_auto
    _orig_pipeline = config.ACTIVE_PIPELINE
    analyze_module.analyze = _fake_analyze
    grid_analyze.analyze_grid_auto = _fake_analyze_grid_auto

    try:
        # --- ACTIVE_PIPELINE="v4" : exactement 1 appel V4, 0 legacy ---
        calls["legacy"] = calls["v4"] = 0
        config.ACTIVE_PIPELINE = "v4"
        dispatch_result = pipeline_dispatch.run_active_pipeline(
            "texte de test", chunks_for_grid=[{"text": "x", "page": None}],
            document_type_override=1,
        )
        _test("ACTIVE_PIPELINE='v4' -> exactement 1 appel V4", calls["v4"] == 1, f"Obtenu : {calls}")
        _test("ACTIVE_PIPELINE='v4' -> 0 appel legacy", calls["legacy"] == 0, f"Obtenu : {calls}")
        _test("ACTIVE_PIPELINE='v4' -> dispatch_result['pipeline']='v4'",
              dispatch_result["pipeline"] == "v4")
        _test("ACTIVE_PIPELINE='v4' -> result (legacy) reste None",
              dispatch_result["result"] is None)
        _test("ACTIVE_PIPELINE='v4' -> result_v4 non None",
              dispatch_result["result_v4"] is not None)

        # --- ACTIVE_PIPELINE="legacy" : exactement 1 appel legacy, 0 V4 ---
        calls["legacy"] = calls["v4"] = 0
        config.ACTIVE_PIPELINE = "legacy"
        dispatch_result = pipeline_dispatch.run_active_pipeline("texte de test")
        _test("ACTIVE_PIPELINE='legacy' -> exactement 1 appel legacy", calls["legacy"] == 1, f"Obtenu : {calls}")
        _test("ACTIVE_PIPELINE='legacy' -> 0 appel V4", calls["v4"] == 0, f"Obtenu : {calls}")
        _test("ACTIVE_PIPELINE='legacy' -> dispatch_result['pipeline']='legacy'",
              dispatch_result["pipeline"] == "legacy")
        _test("ACTIVE_PIPELINE='legacy' -> result_v4 reste None",
              dispatch_result["result_v4"] is None)

        # --- Mauvaise config -> erreur explicite, jamais un repli silencieux ---
        calls["legacy"] = calls["v4"] = 0
        config.ACTIVE_PIPELINE = "n_importe_quoi"
        try:
            pipeline_dispatch.run_active_pipeline("texte de test")
            _test("ACTIVE_PIPELINE invalide -> ValueError levée", False, "Aucune exception levée")
        except ValueError:
            _test("ACTIVE_PIPELINE invalide -> ValueError levée", True)
        _test("ACTIVE_PIPELINE invalide -> AUCUN pipeline appelé (pas de fallback silencieux)",
              calls == {"legacy": 0, "v4": 0}, f"Obtenu : {calls}")

    finally:
        analyze_module.analyze = _orig_analyze
        grid_analyze.analyze_grid_auto = _orig_analyze_grid_auto
        config.ACTIVE_PIPELINE = _orig_pipeline


def test_grid_doctype():
    """Tests de la détection automatique du type de document (R11,
    grid_doctype.py). Backend LLM monkey-patché (même idiome que
    test_grid_analyze_v4), pas d'appel réseau réel."""
    print("\n--- 1.18 Détection automatique du type de document (grid_doctype.py) ---\n")

    import llm_backend
    import grid_doctype

    _orig_dispatch = llm_backend._dispatch

    def _make_fake_dispatch(doc_type, confidence="haute"):
        def _fake(backend, prompt, model, options, timeout, response_format=None):
            return f"TYPE: {doc_type}\nCONFIANCE: {confidence}\nEVIDENCE: extrait de justification\n"
        return _fake

    try:
        # --- 4 types, LLM simulé répond correctement pour chacun ---
        for expected_type in (1, 2, 3, 4):
            llm_backend._dispatch = _make_fake_dispatch(expected_type)
            result = grid_doctype.detect_document_type("Texte de test quelconque.")
            _test(f"Type {expected_type} : détecté correctement",
                  result["document_type"] == expected_type, f"Obtenu : {result}")
            _test(f"Type {expected_type} : source='llm'", result["source"] == "llm")
            _test(f"Type {expected_type} : confidence='haute'", result["confidence"] == "haute")
            _test(f"Type {expected_type} : evidence non vide", bool(result["evidence"]))

        # --- Fail-open : LLM injoignable -> repli heuristique lexical,
        # jamais un Type 1 muet indiscernable d'une vraie détection ---
        def _broken_dispatch(backend, prompt, model, options, timeout, response_format=None):
            raise RuntimeError("backend injoignable (simulation)")
        llm_backend._dispatch = _broken_dispatch

        result_fail = grid_doctype.detect_document_type(
            "Compliance Advisor Ombudsman (CAO) monitoring report — Third Monitoring Period."
        )
        _test("Fail-open : LLM injoignable -> pas d'exception, dict retourné",
              isinstance(result_fail, dict))
        _test("Fail-open : source='fallback_heuristique_lexicale' (pas un Type 1 muet)",
              result_fail["source"] == "fallback_heuristique_lexicale", f"Obtenu : {result_fail}")
        _test("Fail-open : indices CAO/monitoring period -> Type 3 par heuristique",
              result_fail["document_type"] == 3, f"Obtenu : {result_fail}")
        _test("Fail-open : confidence='faible' (signale une détection non fiable)",
              result_fail["confidence"] == "faible")

        # --- Fail-open sans aucun indice lexical -> Type 1, mais explicitement marqué ---
        result_no_hint = grid_doctype.detect_document_type("")
        _test("Fail-open sans indice : document_type=1 (repli prudent)",
              result_no_hint["document_type"] == 1)
        _test("Fail-open sans indice : source='fallback_heuristique_aucun_indice'",
              result_no_hint["source"] == "fallback_heuristique_aucun_indice")

        # --- Réponse LLM inexploitable (pas de ligne TYPE) -> repli heuristique ---
        def _garbage_dispatch(backend, prompt, model, options, timeout, response_format=None):
            return "réponse illisible sans le bon format"
        llm_backend._dispatch = _garbage_dispatch
        result_garbage = grid_doctype.detect_document_type("Annual Monitoring Report prepared by the client.")
        _test("Réponse LLM inexploitable -> repli heuristique (pas de crash)",
              result_garbage["source"].startswith("fallback"))
        _test("Réponse LLM inexploitable + indice AMR -> Type 2 par heuristique",
              result_garbage["document_type"] == 2, f"Obtenu : {result_garbage}")

    finally:
        llm_backend._dispatch = _orig_dispatch


def test_grid_metadata():
    """Tests des métadonnées projet (sponsor/pays/secteur/client/type de
    projet — grid_metadata.py, audit UI 2026-08-20). Backend LLM
    monkey-patché (même idiome que test_grid_doctype), pas d'appel réseau
    réel. Contrairement à grid_doctype.py, PAS de repli heuristique ici
    (aucun équivalent lexical raisonnable pour deviner un sponsor) —
    fail-open = tous les champs à None, source='indisponible'."""
    print("\n--- 1.18b Métadonnées projet (grid_metadata.py) ---\n")

    import json as json_module
    import llm_backend
    import grid_metadata

    _orig_dispatch = llm_backend._dispatch

    try:
        # --- Extraction complète, tous les champs trouvés ---
        def _full_dispatch(backend, prompt, model, options, timeout, response_format=None):
            return json_module.dumps({
                "sponsor": "ABC Energy", "country": "Kenya", "sector": "Énergie renouvelable",
                "client": "XYZ SPV", "project_type": "Parc éolien",
            })
        llm_backend._dispatch = _full_dispatch
        result = grid_metadata.detect_project_metadata("Texte de test quelconque.")
        _test("Extraction complète : sponsor correct", result["sponsor"] == "ABC Energy", f"Obtenu : {result}")
        _test("Extraction complète : country correct", result["country"] == "Kenya")
        _test("Extraction complète : sector correct", result["sector"] == "Énergie renouvelable")
        _test("Extraction complète : client correct", result["client"] == "XYZ SPV")
        _test("Extraction complète : project_type correct", result["project_type"] == "Parc éolien")
        _test("Extraction complète : source='llm'", result["source"] == "llm")

        # --- Extraction partielle : certains champs absents du document ---
        def _partial_dispatch(backend, prompt, model, options, timeout, response_format=None):
            return json_module.dumps({
                "sponsor": "ABC Energy", "country": None, "sector": None,
                "client": None, "project_type": None,
            })
        llm_backend._dispatch = _partial_dispatch
        result_partial = grid_metadata.detect_project_metadata("Texte de test.")
        _test("Extraction partielle : sponsor trouvé", result_partial["sponsor"] == "ABC Energy")
        _test("Extraction partielle : country=None (jamais inventé)", result_partial["country"] is None)
        _test("Extraction partielle : source='llm' même partiel", result_partial["source"] == "llm")

        # --- JSON avec backticks markdown (tolérance de parsing) ---
        def _fenced_dispatch(backend, prompt, model, options, timeout, response_format=None):
            return "```json\n" + json_module.dumps({"sponsor": "Fenced Corp"}) + "\n```"
        llm_backend._dispatch = _fenced_dispatch
        result_fenced = grid_metadata.detect_project_metadata("Texte de test.")
        _test("JSON avec backticks : parsing tolérant", result_fenced["sponsor"] == "Fenced Corp",
              f"Obtenu : {result_fenced}")

        # --- Fail-open : LLM injoignable -> tous les champs à None, PAS
        # d'exception, PAS de repli heuristique (contrairement à R11) ---
        def _broken_dispatch(backend, prompt, model, options, timeout, response_format=None):
            raise RuntimeError("backend injoignable (simulation)")
        llm_backend._dispatch = _broken_dispatch
        result_fail = grid_metadata.detect_project_metadata("Texte de test.")
        _test("Fail-open : pas d'exception, dict retourné", isinstance(result_fail, dict))
        _test("Fail-open : source='indisponible'", result_fail["source"] == "indisponible",
              f"Obtenu : {result_fail}")
        _test("Fail-open : tous les champs à None (aucune valeur inventée)",
              all(result_fail[f] is None for f in ("sponsor", "country", "sector", "client", "project_type")),
              f"Obtenu : {result_fail}")

        # --- Réponse LLM inexploitable (JSON invalide) -> fail-open ---
        def _garbage_dispatch(backend, prompt, model, options, timeout, response_format=None):
            return "réponse illisible sans JSON valide"
        llm_backend._dispatch = _garbage_dispatch
        result_garbage = grid_metadata.detect_project_metadata("Texte de test.")
        _test("Réponse inexploitable -> fail-open, pas de crash",
              result_garbage["source"] == "indisponible", f"Obtenu : {result_garbage}")

    finally:
        llm_backend._dispatch = _orig_dispatch


def test_grid_prompts_v4():
    """Tests des prompts, architecture 2 passes + JSON (grid_prompts.py,
    directive CC-V4-12) : Passe 1 EXTRACTION, Passe 2 QUALIFICATION,
    règles réorganisées (R2/R8 en extraction ; R9/R11 en qualification ;
    R10 en extraction), few-shot des 4 dossiers annotés, parsing JSON
    tolérant (backticks, clés manquantes)."""
    print("\n--- 1.9 Prompts Grille V4 — 2 passes (grid_prompts.py, CC-V4-12) ---\n")

    import json as json_module

    import grid_prompts
    import grid_questions

    # 1. get_extraction_prompt("A.1.1", [...]) retourne un string non vide
    prompt_a11 = grid_prompts.get_extraction_prompt("A.1.1", ["chunk1", "chunk2"])
    _test("Test 1 : get_extraction_prompt('A.1.1', ...) -> string non vide",
          isinstance(prompt_a11, str) and len(prompt_a11) > 0)

    # 2. Le prompt d'extraction contient la formulation R exacte d'A.1.1
    question_r_a11 = grid_questions.get_question("A.1.1")["question_r"]
    _test("Test 2 : formulation R exacte d'A.1.1 présente dans le prompt d'extraction",
          question_r_a11 in prompt_a11, f"Attendu dans le prompt : {question_r_a11!r}")

    # 3. Instruction de langue (F2, CLAUDE.md — invariant non régressé par CC-V4-12)
    _test("Test 3 : instruction de langue ('français') présente en extraction",
          "français" in prompt_a11)
    prompt_qual_lang = grid_prompts.get_qualification_prompt("A.1.1", "verbatim test", "SPV")
    _test("Test 3b : instruction de langue ('français') présente en qualification",
          "français" in prompt_qual_lang)
    _test("Test 3c : consigne 'verbatim reste dans sa langue d'origine' (pas de traduction forcée)",
          "verbatim" in prompt_a11.lower() and "traduction" in prompt_a11.lower())

    # 4. B.3.1 (CC-V4-11) : polarité STANDARD — question_a présente dans le
    # prompt de qualification comme les 11 autres, plus de template dédié.
    prompt_b31_qual = grid_prompts.get_qualification_prompt("B.3.1", "verbatim test", "SPV")
    _test("Test 4 : prompt de qualification B.3.1 mentionne la question de mitigation",
          grid_questions.get_question("B.3.1")["question_a"] in prompt_b31_qual)

    # 4b. Aucune question de la Maquette Vierge n'est plus en polarité
    # inversée (CC-V4-11, inchangé par CC-V4-12).
    for q in grid_questions.QUESTIONS:
        _test(f"Test 4b : {q['code']} n'a pas inverted_polarity=True (CC-V4-11)",
              not q.get("inverted_polarity"))

    # --- R2 — matérialisation conditionnelle au reading_mode, en Passe 1 (CC-V4-12) ---
    prompt_instruction = grid_prompts.get_extraction_prompt("A.1.1", ["chunk"], document_type=1)
    prompt_suivi = grid_prompts.get_extraction_prompt("A.1.1", ["chunk"], document_type=3)
    _test("R2 : mode INSTRUCTION mentionne un 'legacy issue' (constat, pas événement daté)",
          "legacy issue" in prompt_instruction)
    _test("R2 : mode SUIVI ne mentionne PAS 'legacy issue' (règle instruction absente)",
          "legacy issue" not in prompt_suivi)

    # 5. get_extraction_prompt/get_qualification_prompt("Z.9.9", ...) -> None (code inconnu)
    _test("Test 5 : get_extraction_prompt code inconnu -> None",
          grid_prompts.get_extraction_prompt("Z.9.9", ["c"]) is None)
    _test("Test 5b : get_qualification_prompt code inconnu -> None",
          grid_prompts.get_qualification_prompt("Z.9.9", "v", "SPV") is None)

    # --- R11 : formes de preuve selon le type de document — Passe 2 (CC-V4-12) ---
    prompt_qual_type1 = grid_prompts.get_qualification_prompt("B.1.1", "verbatim", "SPV", document_type=1)
    _test("R11 : document Type 1 -> 4ème forme de preuve mentionnée",
          "QUATRIÈME" in prompt_qual_type1 or "plan détaillé" in prompt_qual_type1)

    prompt_qual_type3 = grid_prompts.get_qualification_prompt("B.1.1", "verbatim", "SPV", document_type=3)
    _test("R11 : document Type 3 -> seulement 3 formes ('plan seul reste au statut 2')",
          "plan seul reste au statut 2" in prompt_qual_type3)

    # --- R8 : couches temporelles sur Type 3 seulement — Passe 1 (CC-V4-12) ---
    prompt_ext_type1 = grid_prompts.get_extraction_prompt("A.1.1", ["chunk"], document_type=1)
    prompt_ext_type3 = grid_prompts.get_extraction_prompt("A.1.1", ["chunk"], document_type=3)
    _test("R8 : 'Couche 1' présent sur Type 3 (extraction)", "Couche 1" in prompt_ext_type3)
    _test("R8 : 'Couche 2' présent sur Type 3 (extraction)", "Couche 2" in prompt_ext_type3)
    _test("R8 : 'Couche 1' ABSENT sur Type 1 (règle non universelle)",
          "Couche 1" not in prompt_ext_type1)

    # --- R9 : hiérarchie des sources sur Types 2-3 seulement — Passe 2 (CC-V4-12) ---
    _test("R9 : 'auditeur indépendant' présent sur Type 3 (qualification)",
          "auditeur indépendant" in prompt_qual_type3)
    _test("R9 : hiérarchie des sources ABSENTE sur Type 1 (qualification)",
          "HIÉRARCHIE DES SOURCES" not in prompt_qual_type1)

    # --- R10 : filtre de sujet, en Passe 1, universel (CC-V4-12) ---
    _test("R10 : 5 catégories mentionnées dans le prompt d'extraction",
          all(cat in prompt_a11 for cat in ("SPV", "LENDER", "SUBSTITUTION", "AMBIGUOUS", "INDIRECT")))

    # --- Few-shot des 4 dossiers annotés (reconstruit CC-V4-12, cf.
    # grid_prompts.FEW_SHOTS pour le détail) ---
    prompt_b21 = grid_prompts.get_extraction_prompt("B.2.1", ["chunk"])
    _test("Few-shot B.2.1 : cas Mundra ('PM10')", "PM10" in prompt_b21)

    prompt_a22 = grid_prompts.get_extraction_prompt("A.2.2", ["chunk"])
    _test("Few-shot A.2.2 : cas Mundra (retrait IFC 2018)", "IFC" in prompt_a22 and "2018" in prompt_a22)

    prompt_b11_ext = grid_prompts.get_extraction_prompt("B.1.1", ["chunk"])
    _test("Few-shot B.1.1 : cas Indorama (N/A argumenté)", "Indorama Free Zone" in prompt_b11_ext)
    _test("Few-shot B.1.1 : cas Aysha (déplacement économique)", "909 persons" in prompt_b11_ext)

    prompt_b31_ext = grid_prompts.get_extraction_prompt("B.3.1", ["chunk"])
    _test("Few-shot B.3.1 : baseline chiffrée (Aysha/Indorama)",
          "baseline" in prompt_b31_ext.lower())

    prompt_b32_ext = grid_prompts.get_extraction_prompt("B.3.2", ["chunk"])
    _test("Few-shot B.3.2 : refus de partage de données Mundra",
          "share monitoring data" in prompt_b32_ext)

    # Questions sans AUCUN cas annoté (ni OUI ni NON) -> pas de few-shot
    # inventé. A.3.1/A.3.2 ont un cas NON mais pas de cas OUI — pas
    # "aucun cas annoté" au sens strict, cf. FEW_SHOTS.
    for code in ("B.1.2", "B.2.2"):
        fs = grid_prompts.FEW_SHOTS.get(code) or {}
        _test(f"Pas de few-shot pour {code} (aucun cas annoté, CC-V4-12)",
              fs.get("oui") is None and fs.get("non") is None)

    # Codes retirés (absents de la Maquette Vierge) : plus de prompt du tout.
    for code in ("A.1.3", "A.4.1", "B.2.3", "B.4.1"):
        _test(f"{code} retiré : get_extraction_prompt('{code}', ...) -> None",
              grid_prompts.get_extraction_prompt(code, ["chunk"]) is None)

    # --- Parsing JSON — Passe 1 (extraction) ---
    resp_extraction = json_module.dumps({
        "code": "B.1.1", "found": True, "verbatim": "test verbatim", "page": 9,
        "subject": "SPV", "brief": "constat de deplacement",
    })
    parsed_ext = grid_prompts.parse_extraction_response(resp_extraction)
    _test("Parsing extraction : dict valide", isinstance(parsed_ext, dict), f"Obtenu : {parsed_ext!r}")
    if isinstance(parsed_ext, dict):
        _test("Parsing extraction : found=True", parsed_ext["found"] is True)
        _test("Parsing extraction : verbatim correct", parsed_ext["verbatim"] == "test verbatim")
        _test("Parsing extraction : page=9", parsed_ext["page"] == 9)
        _test("Parsing extraction : subject='SPV'", parsed_ext["subject"] == "SPV")

    # Tolérance backticks markdown autour du JSON
    resp_fenced = "```json\n" + resp_extraction + "\n```"
    parsed_fenced = grid_prompts.parse_extraction_response(resp_fenced)
    _test("Parsing extraction : tolère les backticks markdown ```json ... ```",
          isinstance(parsed_fenced, dict) and parsed_fenced["found"] is True)

    # Tolérance texte parasite avant/après le JSON
    resp_wrapped = "Voici le résultat :\n" + resp_extraction + "\nFin de réponse."
    parsed_wrapped = grid_prompts.parse_extraction_response(resp_wrapped)
    _test("Parsing extraction : tolère du texte avant/après le JSON",
          isinstance(parsed_wrapped, dict) and parsed_wrapped["found"] is True)

    # found=false
    resp_not_found = json_module.dumps({
        "code": "B.1.1", "found": False, "verbatim": None, "page": None,
        "subject": None, "brief": "aucun passage pertinent",
    })
    parsed_nf = grid_prompts.parse_extraction_response(resp_not_found)
    _test("Parsing extraction : found=False", parsed_nf["found"] is False)

    # subject absent/invalide -> repli AMBIGUOUS (JAMAIS SPV, cf. R10)
    resp_no_subject = json_module.dumps({
        "code": "B.1.1", "found": True, "verbatim": "x", "page": None, "brief": "test",
    })
    parsed_no_subj = grid_prompts.parse_extraction_response(resp_no_subject)
    _test("Parsing extraction : subject absent -> repli AMBIGUOUS",
          parsed_no_subj["subject"] == "AMBIGUOUS", f"Obtenu : {parsed_no_subj['subject']!r}")

    # found manquant/invalide -> None (parsing échoué)
    resp_no_found = json_module.dumps({"code": "B.1.1", "verbatim": "x"})
    _test("Parsing extraction : 'found' absent -> None (parsing échoué)",
          grid_prompts.parse_extraction_response(resp_no_found) is None)

    # JSON illisible -> None
    _test("Parsing extraction : JSON illisible -> None",
          grid_prompts.parse_extraction_response("ceci n'est pas du JSON") is None)
    _test("Parsing extraction : chaîne vide -> None",
          grid_prompts.parse_extraction_response("") is None)

    # --- Parsing JSON — Passe 2 (qualification) ---
    resp_qual = json_module.dumps({
        "code": "B.1.1", "status": "OUI", "confidence": "HIGH",
        "mitigation_status": "OUI_PROUVEE", "verbatim_r": "risque", "verbatim_a_mesure": "mesure",
        "verbatim_a_defaillance": None, "brief_r": "test risque", "brief_a": "test mitigation",
    })
    parsed_qual = grid_prompts.parse_qualification_response(resp_qual)
    _test("Parsing qualification : dict valide", isinstance(parsed_qual, dict), f"Obtenu : {parsed_qual!r}")
    if isinstance(parsed_qual, dict):
        _test("Parsing qualification : status='OUI'", parsed_qual["status"] == "OUI")
        _test("Parsing qualification : mitigation_status='OUI_PROUVEE'",
              parsed_qual["mitigation_status"] == "OUI_PROUVEE")
        _test("Parsing qualification : confidence='HIGH'", parsed_qual["confidence"] == "HIGH")

    # NA_ARGUMENTE — nouveau statut CC-V4-12 (cf. AUDIT_PERTINENCE_NOTE_CADRAGE.md point 3)
    resp_na = json_module.dumps({
        "code": "B.1.1", "status": "NA_ARGUMENTE", "confidence": "HIGH",
        "mitigation_status": None, "verbatim_r": "PS5 non declenchee", "verbatim_a_mesure": "",
        "verbatim_a_defaillance": None, "brief_r": "motif explicite d'inapplicabilite", "brief_a": "",
    })
    parsed_na = grid_prompts.parse_qualification_response(resp_na)
    _test("Parsing qualification : status='NA_ARGUMENTE' accepté", parsed_na["status"] == "NA_ARGUMENTE")

    # status invalide -> None (parsing échoué)
    resp_bad_status = json_module.dumps({"code": "B.1.1", "status": "PEUT-ETRE"})
    _test("Parsing qualification : status invalide -> None",
          grid_prompts.parse_qualification_response(resp_bad_status) is None)

    # mitigation_status invalide -> None (pas d'exception)
    resp_bad_mit = json_module.dumps({"code": "B.1.1", "status": "OUI", "mitigation_status": "AUTRE_CHOSE"})
    parsed_bad_mit = grid_prompts.parse_qualification_response(resp_bad_mit)
    _test("Parsing qualification : mitigation_status invalide -> None (pas d'exception)",
          parsed_bad_mit["mitigation_status"] is None)

    # confidence absente -> repli LOW (prudent, pas HIGH par défaut)
    resp_no_conf = json_module.dumps({"code": "B.1.1", "status": "NON"})
    parsed_no_conf = grid_prompts.parse_qualification_response(resp_no_conf)
    _test("Parsing qualification : confidence absente -> repli LOW (prudent)",
          parsed_no_conf["confidence"] == "LOW")

    _test("Parsing qualification : JSON illisible -> None",
          grid_prompts.parse_qualification_response("pas du JSON") is None)


def _zero_risk_grid_answers():
    """Jeu de réponses où aucune question ne déclenche de pénalité : NON
    pour les 12 questions (schéma standard uniquement depuis CC-V4-11 —
    plus aucune question inversée, cf. grid_questions.py). Partagée par
    test_grid_scoring et test_grid_result pour construire des cas de base
    à surcharger. La boucle reste générique (dynamique sur
    grid_questions.ESG_QUESTIONS) au cas où une question inversée serait
    réintroduite."""
    import grid_questions

    answers = {}
    for q in grid_questions.ESG_QUESTIONS:
        if q["inverted_polarity"]:
            answers[q["code"]] = {"status": "OUI", "mitigation": None}
        else:
            answers[q["code"]] = {"status": "NON", "mitigation": None}
    return answers


def test_grid_scoring():
    """Tests du moteur de scoring V3 (grid_scoring.py, directive CC-02)."""
    print("\n--- 1.7 Moteur de scoring Grille V3 (grid_scoring.py) ---\n")

    import grid_questions
    import grid_scoring

    # 1. Tout "sans risque" -> score=100, VERT
    result = grid_scoring.compute_grid_score(_zero_risk_grid_answers())
    _test("Test 1 : aucun risque -> score=100", result["score"] == 100, f"Obtenu : {result['score']}")
    _test("Test 1 : aucun risque -> VERT", result["color"] == "VERT", f"Obtenu : {result['color']}")

    # 2. Tout "à risque" (OUI — schéma standard, plus aucune question
    # inversée depuis CC-V4-11), sans mitigation -> plancher 0, ROUGE
    answers_all_risk = {}
    for q in grid_questions.ESG_QUESTIONS:
        if q["inverted_polarity"]:
            answers_all_risk[q["code"]] = {"status": "NON", "mitigation": None}
        else:
            answers_all_risk[q["code"]] = {"status": "OUI", "mitigation": None}
    result = grid_scoring.compute_grid_score(answers_all_risk)
    _test("Test 2 : tout à risque sans mitigation -> plancher 0", result["score"] == 0, f"Obtenu : {result['score']}")
    _test("Test 2 : tout à risque -> ROUGE", result["color"] == "ROUGE", f"Obtenu : {result['color']}")

    # 3. B.3.1 seul OUI (schéma STANDARD depuis CC-V4-11 : OUI = risque,
    # "Absence de données de référence baseline" avérée) -> 100-15=85
    answers = _zero_risk_grid_answers()
    answers["B.3.1"] = {"status": "OUI", "mitigation": None}
    result = grid_scoring.compute_grid_score(answers)
    _test("Test 3 : B.3.1 OUI (standard, CC-V4-11) -> score=85", result["score"] == 85, f"Obtenu : {result['score']}")
    _test("Test 3 : B.3.1 OUI -> VERT", result["color"] == "VERT", f"Obtenu : {result['color']}")

    # 4. B.3.1 seul NON (favorable — schéma standard : baseline établie) -> 100, VERT
    answers = _zero_risk_grid_answers()
    answers["B.3.1"] = {"status": "NON", "mitigation": None}
    result = grid_scoring.compute_grid_score(answers)
    _test("Test 4 : B.3.1 NON (favorable, CC-V4-11) -> score=100", result["score"] == 100, f"Obtenu : {result['score']}")
    _test("Test 4 : B.3.1 NON -> VERT", result["color"] == "VERT", f"Obtenu : {result['color']}")

    # 5. Les 2 questions B.2 (Maquette Vierge, CC-V4-11 : B.2.1 Air/PM10,
    # B.2.2 Eau/thermique — plus 3 comme en V4 précédent) en NA, reste
    # sans risque -> 100 (N/A exclu du calcul)
    answers = _zero_risk_grid_answers()
    for code in ("B.2.1", "B.2.2"):
        answers[code] = {"status": "NA", "mitigation": None}
    result = grid_scoring.compute_grid_score(answers)
    _test("Test 5 : 2 questions NA, reste sans risque -> score=100", result["score"] == 100, f"Obtenu : {result['score']}")
    _test("Test 5 : questions_na=2", result["questions_na"] == 2, f"Obtenu : {result['questions_na']}")
    # 12 questions Maquette Vierge - 2 NA (B.2.x) = 10 actives.
    _test("Test 5 : questions_active=10", result["questions_active"] == 10, f"Obtenu : {result['questions_active']}")

    # 6. A.1.1 OUI + mitigation OUI -> pénalité -25, gain +5 -> score=80
    answers = _zero_risk_grid_answers()
    answers["A.1.1"] = {"status": "OUI", "mitigation": "OUI"}
    result = grid_scoring.compute_grid_score(answers)
    _test("Test 6 : A.1.1 OUI+mitigation OUI -> score=80", result["score"] == 80, f"Obtenu : {result['score']}")

    # 7. Cap d'atténuation : gain brut > 20 -> plafonné à 20.
    # (6 questions Cat A dans la Maquette Vierge, CC-V4-11 — on n'en
    # sollicite que 4 + 2 Cat B en OUI+mitigation OUI : 4*5 + 2*3 = 26 > 20.)
    answers = _zero_risk_grid_answers()
    for code in ("A.1.1", "A.1.2", "A.2.1", "A.2.2"):
        answers[code] = {"status": "OUI", "mitigation": "OUI"}
    for code in ("B.1.1", "B.1.2"):
        answers[code] = {"status": "OUI", "mitigation": "OUI"}
    result = grid_scoring.compute_grid_score(answers)
    _test("Test 7 : gain brut > 20", result["total_gain"] == 26, f"Obtenu : {result['total_gain']}")
    _test("Test 7 : gain plafonné à 20", result["total_gain_capped"] == 20, f"Obtenu : {result['total_gain_capped']}")
    _test("Test 7 : cap_applied=True", result["cap_applied"] is True, f"Obtenu : {result['cap_applied']}")

    # 8. Mundra simplifié (valeurs arbitraires — PAS un ground truth
    # métier, juste un cas de calcul composite avec pénalités + 1 gain).
    # B.3.1 en OUI (schéma standard, CC-V4-11) au lieu de NON (ancien
    # schéma inversé) — même pénalité (-15), la sémantique de risque
    # change (absence de baseline avérée) mais pas l'arithmétique.
    answers = _zero_risk_grid_answers()
    answers["A.2.1"] = {"status": "OUI", "mitigation": "OUI"}   # -25 +5
    answers["B.1.1"] = {"status": "OUI", "mitigation": "NON"}   # -15
    answers["B.2.1"] = {"status": "NON", "mitigation": None}    # 0
    answers["B.2.2"] = {"status": "OUI", "mitigation": "NON"}   # -15
    answers["B.3.1"] = {"status": "OUI", "mitigation": None}    # -15 (standard, CC-V4-11)
    result = grid_scoring.compute_grid_score(answers)
    # pénalités : -25-15-15-15 = -70 ; gain : +5 -> 100-70+5 = 35
    _test("Test 8 : Mundra simplifié -> score=35", result["score"] == 35, f"Obtenu : {result['score']}")
    # get_color V4 (CC-V4-02) : 4 zones (>=75 VERT, 50-74 JAUNE, 25-49
    # ORANGE, <25 ROUGE) au lieu des 3 zones V3 (<50 ROUGE) — 35 est
    # ORANGE en V4, pas ROUGE.
    _test("Test 8 : Mundra simplifié -> ORANGE (zones V4)", result["color"] == "ORANGE", f"Obtenu : {result['color']}")

    # 9. Code invalide -> ValueError
    answers = _zero_risk_grid_answers()
    answers["Z.9.9"] = {"status": "NON", "mitigation": None}
    try:
        grid_scoring.compute_grid_score(answers)
        _test("Test 9 : code invalide -> ValueError", False, "Aucune exception levée")
    except ValueError:
        _test("Test 9 : code invalide -> ValueError", True)

    # 10. Status invalide -> ValueError
    answers = _zero_risk_grid_answers()
    answers["A.1.1"] = {"status": "PEUT-ÊTRE", "mitigation": None}
    try:
        grid_scoring.compute_grid_score(answers)
        _test("Test 10 : status invalide -> ValueError", False, "Aucune exception levée")
    except ValueError:
        _test("Test 10 : status invalide -> ValueError", True)


def test_grid_scoring_v4():
    """Tests du scoring V4 (grid_scoring.py, directive CC-V4-02) : 4 zones
    de couleur, plafond partagé (dormant depuis CC-V4-11), B.3.1 (schéma
    standard depuis CC-V4-11), métadonnées du mode de lecture (document_type)."""
    print("\n--- 1.7b Scoring Grille V4 (grid_scoring.py) ---\n")

    import grid_questions
    import grid_scoring

    # --- 4 zones de couleur ---
    _test("get_color(80) = VERT", grid_scoring.get_color(80) == "VERT", f"Obtenu : {grid_scoring.get_color(80)!r}")
    _test("get_color(75) = VERT", grid_scoring.get_color(75) == "VERT", f"Obtenu : {grid_scoring.get_color(75)!r}")
    _test("get_color(74) = JAUNE", grid_scoring.get_color(74) == "JAUNE", f"Obtenu : {grid_scoring.get_color(74)!r}")
    _test("get_color(50) = JAUNE", grid_scoring.get_color(50) == "JAUNE", f"Obtenu : {grid_scoring.get_color(50)!r}")
    _test("get_color(49) = ORANGE", grid_scoring.get_color(49) == "ORANGE", f"Obtenu : {grid_scoring.get_color(49)!r}")
    _test("get_color(25) = ORANGE", grid_scoring.get_color(25) == "ORANGE", f"Obtenu : {grid_scoring.get_color(25)!r}")
    _test("get_color(24) = ROUGE", grid_scoring.get_color(24) == "ROUGE", f"Obtenu : {grid_scoring.get_color(24)!r}")
    _test("get_color(0) = ROUGE", grid_scoring.get_color(0) == "ROUGE", f"Obtenu : {grid_scoring.get_color(0)!r}")

    # --- Plafond partagé : DORMANT depuis CC-V4-11 (A.1.3 n'existe plus,
    # aucune question ne porte plus shared_cap_group — cf. grid_questions.py) ---
    _test("Plafond partagé dormant : aucune question ne porte shared_cap_group (CC-V4-11)",
          all(q.get("shared_cap_group") is None for q in grid_questions.QUESTIONS))
    answers_a11_only = _make_answers({"A.1.1": {"status": "OUI", "mitigation_status": "NON_INTENTION"}})
    result_a11_only = grid_scoring.compute_grid_score(answers_a11_only)
    detail_a11 = _find_detail(result_a11_only, "A.1.1")
    _test("A.1.1 seul OUI : pénalité pleine -25, jamais plafonnée (dormant)",
          detail_a11["penalty"] == -25, f"Obtenu : {detail_a11['penalty']}")
    _test("A.1.1 seul OUI : _shared_cap_applied jamais déclenché (dormant)",
          not detail_a11.get("_shared_cap_applied"))

    # --- B.3.1 : schéma STANDARD depuis CC-V4-11 (OUI = risque) ---
    # --- B.3.1 : OUI sans mitigation -> -15 ---
    answers_b31_oui = _make_answers({"B.3.1": {"status": "OUI"}})
    result_b31 = grid_scoring.compute_grid_score(answers_b31_oui)
    detail_b31 = _find_detail(result_b31, "B.3.1")
    _test("B.3.1 OUI sans mitigation : pénalité=-15", detail_b31["penalty"] == -15,
          f"Obtenu : {detail_b31['penalty']}")
    _test("B.3.1 OUI sans mitigation : gain=0", detail_b31["gain"] == 0, f"Obtenu : {detail_b31['gain']}")

    # --- B.3.1 : OUI + mitigation prouvée -> -15 + 3 ---
    answers_b31_mit = _make_answers({"B.3.1": {"status": "OUI", "mitigation_status": "OUI_PROUVEE"}})
    result_b31_mit = grid_scoring.compute_grid_score(answers_b31_mit)
    detail_b31_mit = _find_detail(result_b31_mit, "B.3.1")
    _test("B.3.1 OUI + mitigation prouvée : pénalité=-15", detail_b31_mit["penalty"] == -15,
          f"Obtenu : {detail_b31_mit['penalty']}")
    _test("B.3.1 OUI + mitigation prouvée : gain=3", detail_b31_mit["gain"] == 3,
          f"Obtenu : {detail_b31_mit['gain']}")

    # --- B.3.1 : NON (favorable) -> 0 ---
    answers_b31_non = _make_answers({"B.3.1": {"status": "NON"}})
    result_b31_non = grid_scoring.compute_grid_score(answers_b31_non)
    detail_b31_non = _find_detail(result_b31_non, "B.3.1")
    _test("B.3.1 NON : pénalité=0", detail_b31_non["penalty"] == 0, f"Obtenu : {detail_b31_non['penalty']}")
    _test("B.3.1 NON : gain=0 (mitigation non pertinente côté favorable)", detail_b31_non["gain"] == 0,
          f"Obtenu : {detail_b31_non['gain']}")

    # --- Score maximum = 100 (tout NON, y compris B.3.1 — schéma standard) ---
    result_perfect = grid_scoring.compute_grid_score(_make_all_non())
    _test("Score maximum (tout NON) = 100", result_perfect["score"] == 100,
          f"Obtenu : {result_perfect['score']}")
    _test("Score maximum -> saturation=False", result_perfect["saturation"] is False,
          f"Obtenu : {result_perfect['saturation']}")

    # --- Drapeau de saturation (score plancher = 0) ---
    answers_saturation = {q["code"]: {"status": "OUI" if not q["inverted_polarity"] else "NON",
                                       "mitigation_status": None}
                          for q in grid_questions.QUESTIONS}
    result_saturation = grid_scoring.compute_grid_score(answers_saturation)
    _test("Score plancher (tout à risque) = 0", result_saturation["score"] == 0,
          f"Obtenu : {result_saturation['score']}")
    _test("Score plancher -> saturation=True", result_saturation["saturation"] is True,
          f"Obtenu : {result_saturation['saturation']}")
    _test("Score plancher -> color=ROUGE", result_saturation["color"] == "ROUGE",
          f"Obtenu : {result_saturation['color']!r}")

    # --- Statut INCONNU accepté, scoré comme NON ---
    answers_inconnu = _make_answers({"A.2.1": {"status": "INCONNU"}})
    result_inconnu = grid_scoring.compute_grid_score(answers_inconnu)
    detail_inconnu = _find_detail(result_inconnu, "A.2.1")
    _test("Statut INCONNU accepté sans ValueError", True)
    _test("INCONNU : pénalité=0", detail_inconnu["penalty"] == 0, f"Obtenu : {detail_inconnu['penalty']}")
    _test("INCONNU : conservé tel quel dans le détail (pas réécrit en NON)",
          detail_inconnu["status"] == "INCONNU", f"Obtenu : {detail_inconnu['status']!r}")

    # --- Document type dans le résultat (R11) — n'affecte PAS le score ---
    result_type1 = grid_scoring.compute_grid_score(_make_all_non())
    _test("document_type par défaut = 1", result_type1["document_type"] == 1,
          f"Obtenu : {result_type1['document_type']}")
    _test("reading_mode par défaut = 'instruction'", result_type1["reading_mode"] == "instruction",
          f"Obtenu : {result_type1['reading_mode']!r}")

    result_type3 = grid_scoring.compute_grid_score(_make_all_non(), document_type=3)
    _test("document_type=3 reporté dans le résultat", result_type3["document_type"] == 3,
          f"Obtenu : {result_type3['document_type']}")
    _test("document_type=3 : reading_mode='suivi'", result_type3["reading_mode"] == "suivi",
          f"Obtenu : {result_type3['reading_mode']!r}")
    _test("document_type n'affecte pas le score (mêmes réponses)",
          result_type1["score"] == result_type3["score"],
          f"type1={result_type1['score']} vs type3={result_type3['score']}")

    # --- Verrou B.2.1→B.2.2 (CC-08) toujours intégré en V4 ---
    answers_verrou = _make_answers({"B.2.1": {"status": "OUI"}, "B.2.2": {"status": "NON"}})
    result_verrou = grid_scoring.compute_grid_score(answers_verrou)
    detail_b22 = _find_detail(result_verrou, "B.2.2")
    _test("Verrou B.2.1→B.2.2 toujours actif en V4 : B.2.2 forcé à INCONNU",
          detail_b22["status"] == "INCONNU", f"Obtenu : {detail_b22['status']!r}")

    # --- CC-07 non régressé : MITIGATION_STATUTS / garde-fou statut 4 ---
    _test("MITIGATION_STATUTS non modifié par la V4 (4 statuts)",
          len(grid_questions.MITIGATION_STATUTS) == 4,
          f"Obtenu : {len(grid_questions.MITIGATION_STATUTS)}")

    answers_gf = _make_answers({"B.1.1": {
        "status": "OUI", "mitigation_status": "OUI_DEFAILLANTE",
        "evidence_a": {"verbatim_mesure": None, "verbatim_defaillance": None},
    }})
    result_gf = grid_scoring.compute_grid_score(answers_gf)
    detail_gf = _find_detail(result_gf, "B.1.1")
    _test("Garde-fou statut 4 (CC-07) toujours actif en V4 : retombe sur OUI_PROUVEE",
          detail_gf["mitigation_status"] == "OUI_PROUVEE", f"Obtenu : {detail_gf['mitigation_status']!r}")


def test_grid_result():
    """Tests du format de résultat V3 (grid_result.py, directive CC-03)."""
    print("\n--- 1.8 Format de résultat Grille V3 (grid_result.py) ---\n")

    import json

    import grid_scoring
    import grid_result

    def _question_results_from_answers(answers):
        return {
            code: {**ans, "evidence_r": None, "evidence_a": None, "confidence_note": None}
            for code, ans in answers.items()
        }

    # 1. build_grid_result avec des données valides -> validate_grid_result ne lève pas
    answers = _zero_risk_grid_answers()
    answers["A.1.1"] = {"status": "OUI", "mitigation_status": "OUI_PROUVEE"}
    scoring = grid_scoring.compute_grid_score(answers)
    result = grid_result.build_grid_result(_question_results_from_answers(answers), scoring)
    try:
        grid_result.validate_grid_result(result)
        _test("Test 1 : résultat valide -> pas d'exception", True)
    except ValueError as e:
        _test("Test 1 : résultat valide -> pas d'exception", False, str(e))

    # 2. Résultat avec code inconnu -> ValueError
    bad_result = json.loads(json.dumps(result))
    bad_result["questions"][0]["code"] = "Z.9.9"
    try:
        grid_result.validate_grid_result(bad_result)
        _test("Test 2 : code inconnu -> ValueError", False, "Aucune exception levée")
    except ValueError:
        _test("Test 2 : code inconnu -> ValueError", True)

    # 3. Résultat avec status invalide -> ValueError
    bad_result = json.loads(json.dumps(result))
    bad_result["questions"][0]["status"] = "PEUT-ÊTRE"
    try:
        grid_result.validate_grid_result(bad_result)
        _test("Test 3 : status invalide -> ValueError", False, "Aucune exception levée")
    except ValueError:
        _test("Test 3 : status invalide -> ValueError", True)

    # 4. Résultat avec mitigation sur une question NA -> ValueError
    bad_result = json.loads(json.dumps(result))
    for q in bad_result["questions"]:
        if q["code"] == "B.2.1":
            q["status"] = "NA"
            q["mitigation_status"] = "OUI_PROUVEE"
    try:
        grid_result.validate_grid_result(bad_result)
        _test("Test 4 : mitigation sur question NA -> ValueError", False, "Aucune exception levée")
    except ValueError:
        _test("Test 4 : mitigation sur question NA -> ValueError", True)

    # 5. Résultat avec mitigation sur B.3.1 -> ValueError
    bad_result = json.loads(json.dumps(result))
    for q in bad_result["questions"]:
        if q["code"] == "B.3.1":
            q["mitigation_status"] = "OUI_PROUVEE"
    try:
        grid_result.validate_grid_result(bad_result)
        _test("Test 5 : mitigation sur B.3.1 -> ValueError", False, "Aucune exception levée")
    except ValueError:
        _test("Test 5 : mitigation sur B.3.1 -> ValueError", True)

    # 6. json.dumps(result) fonctionne sans erreur
    try:
        json.dumps(result)
        _test("Test 6 : json.dumps sans erreur", True)
    except TypeError as e:
        _test("Test 6 : json.dumps sans erreur", False, str(e))

    # 7. B.3.1 (schéma standard, CC-V4-11) avec status="OUI" -> penalty = -15
    answers_b31 = _zero_risk_grid_answers()
    answers_b31["B.3.1"] = {"status": "OUI"}
    scoring_b31 = grid_scoring.compute_grid_score(answers_b31)
    result_b31 = grid_result.build_grid_result(_question_results_from_answers(answers_b31), scoring_b31)
    b31_entry = next(q for q in result_b31["questions"] if q["code"] == "B.3.1")
    _test("Test 7 : B.3.1 OUI (standard, CC-V4-11) -> penalty=-15", b31_entry["penalty"] == -15,
          f"Obtenu : {b31_entry['penalty']}")

    # 8. OUI_DEFAILLANTE sans les deux verbatims -> ValueError (CC-07 —
    # validate_grid_result est STRICT ici, contrairement à
    # compute_grid_score qui corrige silencieusement, cf. grid_result.py)
    bad_result = json.loads(json.dumps(result))
    for q in bad_result["questions"]:
        if q["code"] == "B.2.2":
            q["mitigation_status"] = "OUI_DEFAILLANTE"
            q["evidence_a"] = {"verbatim_mesure": "filters installed", "verbatim_defaillance": ""}
    try:
        grid_result.validate_grid_result(bad_result)
        _test("Test 8 : OUI_DEFAILLANTE sans double verbatim -> ValueError", False, "Aucune exception levée")
    except ValueError:
        _test("Test 8 : OUI_DEFAILLANTE sans double verbatim -> ValueError", True)

    # 9. evidence_a fournie alors que mitigation_status=None -> ValueError (CC-07)
    bad_result = json.loads(json.dumps(result))
    for q in bad_result["questions"]:
        if q["code"] == "B.2.2":
            q["mitigation_status"] = None
            q["evidence_a"] = {"verbatim_mesure": "filters installed", "verbatim_defaillance": None}
    try:
        grid_result.validate_grid_result(bad_result)
        _test("Test 9 : evidence_a sans mitigation_status -> ValueError", False, "Aucune exception levée")
    except ValueError:
        _test("Test 9 : evidence_a sans mitigation_status -> ValueError", True)

    # 10. status=NA sans verbatim (evidence_r) -> ValueError (CC-08 —
    # cf. grid_questions.SILENCE_VALUES["NA"] : "verbatim obligatoire")
    bad_result = json.loads(json.dumps(result))
    for q in bad_result["questions"]:
        if q["code"] == "B.2.1":
            q["status"] = "NA"
            q["evidence_r"] = None
    try:
        grid_result.validate_grid_result(bad_result)
        _test("Test 10 : status=NA sans verbatim -> ValueError", False, "Aucune exception levée")
    except ValueError:
        _test("Test 10 : status=NA sans verbatim -> ValueError", True)


def test_grid_result_v4():
    """Tests du format de résultat V4 (grid_result.py, directive CC-V4-03) :
    grid_version, document_type/reading_mode, garde-fou mitigation basé sur
    a_condition (B.3.1), champ qualifying optionnel."""
    print("\n--- 1.8b Format de résultat Grille V4 (grid_result.py) ---\n")

    import json

    import grid_questions
    import grid_scoring
    import grid_result

    def _question_results_from_answers(answers):
        return {
            code: {**ans, "evidence_r": None, "evidence_a": None, "confidence_note": None}
            for code, ans in answers.items()
        }

    base_answers = _make_all_non()
    scoring = grid_scoring.compute_grid_score(base_answers, document_type=2)
    result = grid_result.build_grid_result(_question_results_from_answers(base_answers), scoring)

    # 1. grid_version == "V4"
    _test("grid_version == 'V4'", result["grid_version"] == "V4", f"Obtenu : {result['grid_version']!r}")

    # 2. 12 questions dans le résultat
    _test("12 questions dans le résultat", len(result["questions"]) == 12,
          f"Obtenu : {len(result['questions'])}")

    # 3. document_type et reading_mode présents
    _test("document_type=2 reporté dans le résultat", result["document_type"] == 2,
          f"Obtenu : {result['document_type']}")
    _test("reading_mode='suivi' (document_type=2)", result["reading_mode"] == "suivi",
          f"Obtenu : {result['reading_mode']!r}")
    _test("reading_mode_label présent", bool(result["reading_mode_label"]))
    try:
        grid_result.validate_grid_result(result)
        _test("Résultat de base -> validate_grid_result ne lève pas", True)
    except ValueError as e:
        _test("Résultat de base -> validate_grid_result ne lève pas", False, str(e))

    # 4. B.3.1 (schéma standard depuis CC-V4-11) avec status=OUI et
    # mitigation_status=OUI_PROUVEE -> valide
    answers_b31_valide = _make_answers({"B.3.1": {"status": "OUI", "mitigation_status": "OUI_PROUVEE"}})
    scoring_b31_valide = grid_scoring.compute_grid_score(answers_b31_valide)
    result_b31_valide = grid_result.build_grid_result(
        _question_results_from_answers(answers_b31_valide), scoring_b31_valide
    )
    try:
        grid_result.validate_grid_result(result_b31_valide)
        _test("B.3.1 OUI + mitigation OUI_PROUVEE -> valide (standard, CC-V4-11)", True)
    except ValueError as e:
        _test("B.3.1 OUI + mitigation OUI_PROUVEE -> valide (standard, CC-V4-11)", False, str(e))
    b31_entry = next(q for q in result_b31_valide["questions"] if q["code"] == "B.3.1")
    _test("B.3.1 : a_condition='r_oui' reporté dans le résultat (standard, CC-V4-11)",
          b31_entry["a_condition"] == "r_oui", f"Obtenu : {b31_entry['a_condition']!r}")

    # 5. B.3.1 (schéma standard) avec status=NON et mitigation_status
    # fourni -> ValueError (mitigation illégale du côté "sans risque")
    bad_result = json.loads(json.dumps(result))
    for q in bad_result["questions"]:
        if q["code"] == "B.3.1":
            q["status"] = "NON"
            q["mitigation_status"] = "OUI_PROUVEE"
    try:
        grid_result.validate_grid_result(bad_result)
        _test("B.3.1 NON + mitigation -> ValueError (standard, CC-V4-11)", False, "Aucune exception levée")
    except ValueError:
        _test("B.3.1 NON + mitigation -> ValueError (standard, CC-V4-11)", True)

    # 6. json.dumps(result) fonctionne
    try:
        json.dumps(result)
        _test("json.dumps(result) sans erreur", True)
    except TypeError as e:
        _test("json.dumps(result) sans erreur", False, str(e))

    # 7. qualifying peut être None
    _test("qualifying=None par défaut (non fourni en entrée)",
          all(q["qualifying"] is None for q in result["questions"]))

    # 8. qualifying avec allegation remplie -> valide
    answers_qual = _make_answers({"A.2.1": {"status": "NON"}})
    question_results_qual = _question_results_from_answers(answers_qual)
    question_results_qual["A.2.1"]["qualifying"] = {
        "allegation": "ONG locale allègue un déversement, non confirmé par le rapport.",
        "access_denied": False,
    }
    scoring_qual = grid_scoring.compute_grid_score(answers_qual)
    result_qual = grid_result.build_grid_result(question_results_qual, scoring_qual)
    a21_entry = next(q for q in result_qual["questions"] if q["code"] == "A.2.1")
    _test("qualifying avec allegation remplie -> reporté tel quel",
          a21_entry["qualifying"] is not None and "allegation" in a21_entry["qualifying"],
          f"Obtenu : {a21_entry['qualifying']!r}")
    try:
        grid_result.validate_grid_result(result_qual)
        _test("qualifying avec allegation remplie -> validate_grid_result ne lève pas", True)
    except ValueError as e:
        _test("qualifying avec allegation remplie -> validate_grid_result ne lève pas", False, str(e))

    # --- document_type invalide -> ValueError ---
    bad_result = json.loads(json.dumps(result))
    bad_result["document_type"] = 99
    try:
        grid_result.validate_grid_result(bad_result)
        _test("document_type invalide (99) -> ValueError", False, "Aucune exception levée")
    except ValueError:
        _test("document_type invalide (99) -> ValueError", True)

    # --- color V4 (JAUNE) accepté ---
    _test("_VALID_COLORS accepte JAUNE (zones V4)", "JAUNE" in grid_result._VALID_COLORS,
          f"Obtenu : {grid_result._VALID_COLORS}")

    # --- qualifying de type invalide -> ValueError ---
    bad_result = json.loads(json.dumps(result))
    bad_result["questions"][0]["qualifying"] = "pas un dict"
    try:
        grid_result.validate_grid_result(bad_result)
        _test("qualifying non-dict -> ValueError", False, "Aucune exception levée")
    except ValueError:
        _test("qualifying non-dict -> ValueError", True)


def _make_all_non():
    """Retourne un dict answers avec toutes les questions à NON — "sans
    risque" pour les 12 questions depuis CC-V4-11 (schéma standard
    uniquement, plus de polarité inversée, cf. grid_questions.py)."""
    import grid_questions
    return {q["code"]: {"status": "NON"} for q in grid_questions.ESG_QUESTIONS}


def _make_answers(overrides):
    """Retourne un dict answers avec toutes les questions à NON,
    sauf celles spécifiées dans overrides."""
    base = _make_all_non()
    base.update(overrides)
    return base


def _find_detail(result, code):
    """Trouve le détail d'une question dans le résultat."""
    for d in result["details"]:
        if d["code"] == code:
            return d
    raise ValueError(f"Code {code} non trouvé dans les détails")


def test_mitigation_4_statuts():
    """Tests des 4 statuts de mitigation (passe CBG, directive CC-07)."""
    print("\n--- 1.10 Mitigation à 4 statuts (passe CBG) ---\n")

    import grid_questions
    import grid_scoring

    # --- Structure ---
    _test("MITIGATION_STATUTS contient exactement 4 clés",
          len(grid_questions.MITIGATION_STATUTS) == 4,
          f"Obtenu : {len(grid_questions.MITIGATION_STATUTS)}")

    _test("Chaque statut a label/points_multiplier/description",
          all("label" in v and "points_multiplier" in v and "description" in v
              for v in grid_questions.MITIGATION_STATUTS.values()))

    _test("points_multiplier ∈ {0, 1} pour tous les statuts",
          all(v["points_multiplier"] in (0, 1) for v in grid_questions.MITIGATION_STATUTS.values()))

    _test("Seul OUI_PROUVEE a points_multiplier == 1",
          grid_questions.MITIGATION_STATUTS["OUI_PROUVEE"]["points_multiplier"] == 1)
    _test("NON_INTENTION/NON_FORME_INSUFFISANTE/OUI_DEFAILLANTE ont points_multiplier == 0",
          all(grid_questions.MITIGATION_STATUTS[k]["points_multiplier"] == 0
              for k in ("NON_INTENTION", "NON_FORME_INSUFFISANTE", "OUI_DEFAILLANTE")))

    # --- Scoring statut 1 : intention ---
    # "The company will implement..." -> NON_INTENTION
    answers_s1 = _make_answers({"B.1.1": {"status": "OUI", "mitigation_status": "NON_INTENTION"}})
    result_s1 = grid_scoring.compute_grid_score(answers_s1)
    detail_s1 = _find_detail(result_s1, "B.1.1")
    _test("Statut 1 (intention) : pénalité -15", detail_s1["penalty"] == -15, f"Obtenu : {detail_s1['penalty']}")
    _test("Statut 1 (intention) : gain 0", detail_s1["gain"] == 0, f"Obtenu : {detail_s1['gain']}")

    # --- Scoring statut 2 : forme insuffisante ---
    # "A complaints register was opened in 2015" -> NON_FORME_INSUFFISANTE
    answers_s2 = _make_answers({"B.1.2": {"status": "OUI", "mitigation_status": "NON_FORME_INSUFFISANTE"}})
    result_s2 = grid_scoring.compute_grid_score(answers_s2)
    detail_s2 = _find_detail(result_s2, "B.1.2")
    _test("Statut 2 (forme insuffisante) : gain 0", detail_s2["gain"] == 0, f"Obtenu : {detail_s2['gain']}")

    # --- Scoring statut 3 : prouvée ---
    # "24 of 31 complaints have been processed and closed" -> OUI_PROUVEE
    answers_s3 = _make_answers({"B.1.2": {"status": "OUI", "mitigation_status": "OUI_PROUVEE"}})
    result_s3 = grid_scoring.compute_grid_score(answers_s3)
    detail_s3 = _find_detail(result_s3, "B.1.2")
    _test("Statut 3 (prouvée) : gain +3 (Cat B)", detail_s3["gain"] == 3, f"Obtenu : {detail_s3['gain']}")

    # --- Scoring statut 4 : défaillante ---
    # Filtres installés MAIS poussières persistantes -> OUI_DEFAILLANTE
    answers_s4 = _make_answers({"B.2.2": {
        "status": "OUI",
        "mitigation_status": "OUI_DEFAILLANTE",
        "evidence_a": {
            "verbatim_mesure": "additional filters and scrubbers were installed",
            "verbatim_defaillance": "generated significant quantities of fugitive dust",
            "page": None,
        },
    }})
    result_s4 = grid_scoring.compute_grid_score(answers_s4)
    detail_s4 = _find_detail(result_s4, "B.2.2")
    _test("Statut 4 (défaillante, avec double verbatim) : gain 0", detail_s4["gain"] == 0, f"Obtenu : {detail_s4['gain']}")

    # --- Garde-fou : statut 4 sans double verbatim -> retour en statut 3 ---
    answers_gf = _make_answers({"B.2.2": {
        "status": "OUI",
        "mitigation_status": "OUI_DEFAILLANTE",
        "evidence_a": {
            "verbatim_mesure": "filters installed",
            "verbatim_defaillance": "",  # vide -> garde-fou
            "page": None,
        },
    }})
    result_gf = grid_scoring.compute_grid_score(answers_gf)
    detail_gf = _find_detail(result_gf, "B.2.2")
    _test("Garde-fou statut 4 sans double verbatim -> retombe OUI_PROUVEE (+3)",
          detail_gf["gain"] == 3, f"Obtenu : {detail_gf['gain']}")
    _test("Garde-fou : mitigation_status résolu = OUI_PROUVEE dans le détail",
          detail_gf["mitigation_status"] == "OUI_PROUVEE", f"Obtenu : {detail_gf['mitigation_status']!r}")

    # --- Scoring Cat A avec mitigation prouvée ---
    answers_a = _make_answers({"A.1.1": {"status": "OUI", "mitigation_status": "OUI_PROUVEE"}})
    result_a = grid_scoring.compute_grid_score(answers_a)
    detail_a = _find_detail(result_a, "A.1.1")
    _test("Cat A prouvée : pénalité -25", detail_a["penalty"] == -25, f"Obtenu : {detail_a['penalty']}")
    _test("Cat A prouvée : gain +5", detail_a["gain"] == 5, f"Obtenu : {detail_a['gain']}")

    # --- Score global CBG (vérifié manuellement, CC-V4-11) ---
    # B.1.1 OUI, mit=NON_INTENTION -> -15, +0
    # B.1.2 OUI, mit=OUI_PROUVEE -> -15, +3
    # B.2.1 OUI, mit=NON_FORME_INSUFFISANTE -> -15, +0
    # B.2.2 OUI, mit=OUI_DEFAILLANTE (avec 2 verbatims) -> -15, +0
    # B.3.1 OUI (schéma standard, CC-V4-11) -> -15
    # Reste = NON (sans risque) -> 0
    # Total pénalités = -75, gains = +3, score = 100 - 75 + 3 = 28
    answers_cbg = _make_all_non()
    answers_cbg["B.1.1"] = {"status": "OUI", "mitigation_status": "NON_INTENTION"}
    answers_cbg["B.1.2"] = {"status": "OUI", "mitigation_status": "OUI_PROUVEE"}
    answers_cbg["B.2.1"] = {"status": "OUI", "mitigation_status": "NON_FORME_INSUFFISANTE"}
    answers_cbg["B.2.2"] = {
        "status": "OUI", "mitigation_status": "OUI_DEFAILLANTE",
        "evidence_a": {
            "verbatim_mesure": "additional filters and scrubbers were installed on the dryer stack in recent years",
            "verbatim_defaillance": "generated significant quantities of fugitive dust as well as particulate and gaseous emissions",
            "page": None,
        },
    }
    answers_cbg["B.3.1"] = {"status": "OUI"}  # standard (CC-V4-11)
    result_cbg = grid_scoring.compute_grid_score(answers_cbg)
    _test("Score global CBG = 28/100", result_cbg["score"] == 28, f"Obtenu : {result_cbg['score']}")
    # get_color V4 (CC-V4-02) : 28 est ORANGE (25-49) en V4, pas ROUGE (V3: <50).
    _test("Score global CBG -> ORANGE (zones V4)", result_cbg["color"] == "ORANGE", f"Obtenu : {result_cbg['color']}")

    # --- Rétro-compatibilité CC-02 : ancien champ "mitigation" OUI/NON ---
    answers_legacy = _make_answers({"A.1.1": {"status": "OUI", "mitigation": "OUI"}})
    result_legacy = grid_scoring.compute_grid_score(answers_legacy)
    detail_legacy = _find_detail(result_legacy, "A.1.1")
    _test("Rétro-compat : mitigation='OUI' (ancien format) -> converti en OUI_PROUVEE (+5)",
          detail_legacy["gain"] == 5, f"Obtenu : {detail_legacy['gain']}")
    _test("Rétro-compat : mitigation_status résolu = OUI_PROUVEE",
          detail_legacy["mitigation_status"] == "OUI_PROUVEE", f"Obtenu : {detail_legacy['mitigation_status']!r}")


def test_silence_evenement_etat():
    """Tests des règles de silence événement/état (passe CBG, directive CC-08)."""
    print("\n--- 1.11 Règles de silence événement/état (passe CBG) ---\n")

    import grid_questions
    import grid_scoring

    # --- Classification correcte (refaite pour les 12 codes CC-V4-11,
    # par analogie avec la logique CC-08 — cf. FRAGILE grid_questions.py :
    # pas encore re-validée par Elisa question par question) ---
    _test("A.1.1 silence_type='evenement'",
          grid_questions.get_question("A.1.1")["silence_type"] == "evenement")
    _test("B.1.2 silence_type='evenement' (déplacement non réinstallé, CC-V4-11)",
          grid_questions.get_question("B.1.2")["silence_type"] == "evenement")
    _test("B.2.1 silence_type='etat' (dépassements Air, mesure requise)",
          grid_questions.get_question("B.2.1")["silence_type"] == "etat")
    _test("B.2.2 silence_type='etat' (modélisation Eau, mesure requise)",
          grid_questions.get_question("B.2.2")["silence_type"] == "etat")
    _test("B.3.1 silence_type='etat' (absence de baseline, CC-V4-11)",
          grid_questions.get_question("B.3.1")["silence_type"] == "etat")
    _test("B.3.2 silence_type='etat' (suivi RSE périodique, CC-V4-11)",
          grid_questions.get_question("B.3.2")["silence_type"] == "etat")
    _test("Toutes les questions ont silence_type ∈ {evenement, etat}",
          all(q["silence_type"] in ("evenement", "etat") for q in grid_questions.ESG_QUESTIONS))

    # --- Silence événement -> NON, aucune pénalité -> score=100.
    # Depuis CC-V4-11, B.3.1 utilise le schéma standard (NON = sans
    # risque, comme les 11 autres questions) — _make_all_non() suffit
    # seule, plus besoin de forcer B.3.1 séparément (l'ancien schéma
    # inversé qui nécessitait ce contournement n'existe plus).
    result_evt = grid_scoring.compute_grid_score(_make_all_non())
    _test("Silence événement (NON sur les 12 questions) -> score=100",
          result_evt["score"] == 100, f"Obtenu : {result_evt['score']}")

    # --- Silence état -> INCONNU, pénalité 0 ---
    answers_etat = _make_answers({"B.2.1": {"status": "INCONNU"}})
    result_etat = grid_scoring.compute_grid_score(answers_etat)
    detail_etat = _find_detail(result_etat, "B.2.1")
    _test("Silence état (B.2.1=INCONNU) -> pénalité 0",
          detail_etat["penalty"] == 0, f"Obtenu : {detail_etat['penalty']}")

    # --- INCONNU ne se convertit jamais en NON (ni l'inverse) ---
    answers_inconnu = _make_answers({"B.1.2": {"status": "INCONNU"}})
    result_inconnu = grid_scoring.compute_grid_score(answers_inconnu)
    detail_inconnu = _find_detail(result_inconnu, "B.1.2")
    _test("INCONNU n'est jamais reconverti en NON",
          detail_inconnu["status"] == "INCONNU", f"Obtenu : {detail_inconnu['status']!r}")


def test_verrou_b21_b22():
    """Test du verrou logique B.2.1 -> B.2.2 (passe CBG, directive CC-08)."""
    print("\n--- 1.12 Verrou B.2.1→B.2.2 (passe CBG) ---\n")

    import grid_scoring

    # --- B.2.1 = OUI + B.2.2 = NON -> B.2.2 forcé à INCONNU ---
    answers_verrou = _make_all_non()
    answers_verrou["B.2.1"] = {"status": "OUI", "mitigation_status": "NON_INTENTION"}
    answers_verrou["B.2.2"] = {"status": "NON"}  # devrait être forcé à INCONNU
    result_verrou = grid_scoring.compute_grid_score(answers_verrou)
    detail_b22 = _find_detail(result_verrou, "B.2.2")
    _test("Verrou : B.2.1=OUI + B.2.2=NON -> B.2.2 forcé à INCONNU",
          detail_b22["status"] == "INCONNU", f"Obtenu : {detail_b22['status']!r}")
    _test("Verrou : B.2.2 forcé -> pénalité 0",
          detail_b22["penalty"] == 0, f"Obtenu : {detail_b22['penalty']}")
    _test("Verrou : verrou_applique=True pour B.2.2",
          detail_b22["verrou_applique"] is True, f"Obtenu : {detail_b22['verrou_applique']!r}")

    # --- B.2.1 = NON + B.2.2 = NON -> pas de verrou ---
    answers_no_verrou = _make_all_non()
    result_nv = grid_scoring.compute_grid_score(answers_no_verrou)
    detail_b22_nv = _find_detail(result_nv, "B.2.2")
    _test("Pas de verrou si B.2.1=NON : B.2.2 reste NON",
          detail_b22_nv["status"] == "NON", f"Obtenu : {detail_b22_nv['status']!r}")
    _test("Pas de verrou si B.2.1=NON : verrou_applique=False",
          detail_b22_nv["verrou_applique"] is False, f"Obtenu : {detail_b22_nv['verrou_applique']!r}")

    # --- B.2.1 = OUI + B.2.2 = OUI -> pas de verrou (le risque est documenté) ---
    answers_oui = _make_all_non()
    answers_oui["B.2.1"] = {"status": "OUI", "mitigation_status": "NON_INTENTION"}
    answers_oui["B.2.2"] = {"status": "OUI", "mitigation_status": "NON_INTENTION"}
    result_oui = grid_scoring.compute_grid_score(answers_oui)
    detail_b22_oui = _find_detail(result_oui, "B.2.2")
    _test("Pas de verrou si B.2.2=OUI (risque déjà documenté)",
          detail_b22_oui["status"] == "OUI", f"Obtenu : {detail_b22_oui['status']!r}")
    _test("Pas de verrou si B.2.2=OUI : pénalité -15 conservée",
          detail_b22_oui["penalty"] == -15, f"Obtenu : {detail_b22_oui['penalty']}")


def test_flag_atteste():
    """Test du flag attesté (NON + verbatim explicite), directive CC-08.
    Le flag est calculé dans build_grid_result (pas dans le scoring) —
    atteste = (status == "NON" and evidence_r is not None), cf. sa
    docstring. Ne change PAS le score, change seulement la lisibilité."""
    print("\n--- 1.13 Flag attesté (passe CBG) ---\n")

    import grid_scoring
    import grid_result

    def _question_results_from_answers(answers):
        return {
            code: {**ans, "evidence_a": None, "confidence_note": None}
            for code, ans in answers.items()
        }

    # --- NON + verbatim explicite -> atteste=True ---
    answers = _make_answers({"B.3.1": {"status": "OUI"}, "A.1.1": {"status": "NON"}})
    question_results = _question_results_from_answers(answers)
    question_results["A.1.1"]["evidence_r"] = {
        "passage": "Fieldwork confirmed the absence of any labor dispute.", "page": 12,
    }
    scoring = grid_scoring.compute_grid_score(answers)
    result = grid_result.build_grid_result(question_results, scoring)
    a11 = next(q for q in result["questions"] if q["code"] == "A.1.1")
    _test("NON + evidence_r non-None -> atteste=True", a11["atteste"] is True, f"Obtenu : {a11['atteste']!r}")

    try:
        grid_result.validate_grid_result(result)
        _test("Résultat avec atteste=True -> validate_grid_result ne lève pas", True)
    except ValueError as e:
        _test("Résultat avec atteste=True -> validate_grid_result ne lève pas", False, str(e))

    # --- NON sans verbatim (silence) -> atteste=False ---
    question_results_silence = _question_results_from_answers(answers)
    question_results_silence["A.1.1"]["evidence_r"] = None
    scoring_silence = grid_scoring.compute_grid_score(answers)
    result_silence = grid_result.build_grid_result(question_results_silence, scoring_silence)
    a11_silence = next(q for q in result_silence["questions"] if q["code"] == "A.1.1")
    _test("NON sans evidence_r -> atteste=False", a11_silence["atteste"] is False, f"Obtenu : {a11_silence['atteste']!r}")

    # --- OUI + verbatim -> atteste=False (atteste exige status=NON, pas juste une preuve) ---
    answers_oui = _make_answers({
        "B.3.1": {"status": "OUI"},
        "A.1.1": {"status": "OUI", "mitigation_status": "NON_INTENTION"},
    })
    question_results_oui = _question_results_from_answers(answers_oui)
    question_results_oui["A.1.1"]["evidence_r"] = {"passage": "A strike occurred on site in March.", "page": 3}
    scoring_oui = grid_scoring.compute_grid_score(answers_oui)
    result_oui = grid_result.build_grid_result(question_results_oui, scoring_oui)
    a11_oui = next(q for q in result_oui["questions"] if q["code"] == "A.1.1")
    _test("OUI + evidence_r -> atteste=False (atteste exige status=NON)",
          a11_oui["atteste"] is False, f"Obtenu : {a11_oui['atteste']!r}")


def test_esap_exclusion():
    """Test de l'exclusion ESAP (passe CBG, directive CC-09)."""
    print("\n--- 1.14a Exclusion ESAP (grid_sections.py) ---\n")

    import grid_sections

    chunks_with_esap = [
        {"text": "The project has significant impacts on biodiversity.", "page": 1},
        {"text": "Environmental and Social Action Plan\nItem 1: Prepare a Livelihood Restoration Plan", "page": 5},
        {"text": "Item 2: Finalize the Past Compensation Report (2010-2015)", "page": 5},
        {"text": "Item 3: Complete establishment of a data management system", "page": 6},
        {"text": "The project area includes critical habitat for chimpanzees.", "page": 8},
    ]
    classified = grid_sections.classify_chunks(chunks_with_esap)
    _test("Chunk avant ESAP : section_tag=None", classified[0]["section_tag"] is None,
          f"Obtenu : {classified[0]['section_tag']!r}")
    _test("Chunk titre ESAP (Item 1) : section_tag='esap'", classified[1]["section_tag"] == "esap",
          f"Obtenu : {classified[1]['section_tag']!r}")
    _test("Chunk ESAP (Item 2) : section_tag='esap'", classified[2]["section_tag"] == "esap",
          f"Obtenu : {classified[2]['section_tag']!r}")
    _test("Chunk ESAP (Item 3) : section_tag='esap'", classified[3]["section_tag"] == "esap",
          f"Obtenu : {classified[3]['section_tag']!r}")
    _test("Chunk après ESAP : section_tag=None", classified[4]["section_tag"] is None,
          f"Obtenu : {classified[4]['section_tag']!r}")

    # --- Question de risque (R) : les chunks ESAP sont INCLUS ---
    r_chunks = grid_sections.get_chunks_for_question(classified, "B.2.1")
    _test("Question R : chunks ESAP inclus (renseignent le risque)",
          len([c for c in r_chunks if c["section_tag"] == "esap"]) > 0)

    # --- Question de mitigation (A) : les chunks ESAP sont EXCLUS ---
    a_chunks = grid_sections.get_chunks_for_question(classified, "B.2.1", for_mitigation=True)
    _test("Question A (mitigation) : chunks ESAP exclus",
          len([c for c in a_chunks if c["section_tag"] == "esap"]) == 0)


def test_ifc_complaints_exclusion():
    """Test de l'exclusion de la section plaintes IFC (directive CC-09)."""
    print("\n--- 1.14b Exclusion section plaintes IFC (grid_sections.py) ---\n")

    import grid_sections

    chunks_with_complaints = [
        {"text": "The project has 31 unresolved complaints.", "page": 10},
        {"text": "Information on E&S Complaints and Grievance Mechanisms\nAffected communities have unrestricted access to the Compliance Advisor Ombudsman (CAO).", "page": 15},
        {"text": "The CAO is mandated to address complaints from people affected by IFC projects.", "page": 15},
    ]
    classified = grid_sections.classify_chunks(chunks_with_complaints)
    _test("Chunk avant section plaintes : section_tag=None", classified[0]["section_tag"] is None,
          f"Obtenu : {classified[0]['section_tag']!r}")
    _test("Chunk titre section plaintes : section_tag='ifc_complaints'",
          classified[1]["section_tag"] == "ifc_complaints", f"Obtenu : {classified[1]['section_tag']!r}")
    _test("Chunk suite section plaintes : section_tag='ifc_complaints'",
          classified[2]["section_tag"] == "ifc_complaints", f"Obtenu : {classified[2]['section_tag']!r}")

    # --- Chunks ifc_complaints exclus de TOUT (R et A) ---
    r_chunks = grid_sections.get_chunks_for_question(classified, "A.2.1")
    _test("Question R : chunks ifc_complaints exclus",
          all(c["section_tag"] != "ifc_complaints" for c in r_chunks))
    a_chunks = grid_sections.get_chunks_for_question(classified, "A.2.1", for_mitigation=True)
    _test("Question A : chunks ifc_complaints exclus",
          all(c["section_tag"] != "ifc_complaints" for c in a_chunks))


def test_no_sections_detected():
    """Si aucune section spéciale n'est détectée, tous les chunks passent
    inchangés (fail-open, cf. ADR-002 et grid_sections.py)."""
    print("\n--- 1.14c Aucune section détectée -> fail-open (grid_sections.py) ---\n")

    import grid_sections

    chunks_normal = [
        {"text": "Normal project description.", "page": 1},
        {"text": "Some environmental data.", "page": 2},
    ]
    classified = grid_sections.classify_chunks(chunks_normal)
    _test("Aucun chunk taggé", all(c["section_tag"] is None for c in classified))
    _test("get_chunks_for_question renvoie tous les chunks",
          len(grid_sections.get_chunks_for_question(classified, "A.1.1")) == 2,
          f"Obtenu : {len(grid_sections.get_chunks_for_question(classified, 'A.1.1'))}")


def test_temporal_layer_marking():
    """Test du marquage best-effort des couches temporelles (grid_sections.py,
    addition V4, directive CC-V4-04) — bonus non bloquant, cf. R8."""
    print("\n--- 1.14d Marquage couches temporelles (grid_sections.py, bonus V4) ---\n")

    import grid_sections

    # --- Détection + persistance jusqu'au titre suivant ---
    chunks = [
        {"text": "Introduction generale du rapport de suivi.", "page": 1},
        {"text": "Summary of First Monitoring Period\nNo major incidents reported.", "page": 3},
        {"text": "Water quality remained within permitted limits during this period.", "page": 4},
        {"text": "Third Monitoring Period\nA grievance was filed by local residents.", "page": 20},
        {"text": "The grievance was resolved through the community mechanism.", "page": 21},
    ]
    classified = grid_sections.classify_chunks(chunks)
    _test("Avant tout titre de période : section_tag=None",
          classified[0]["section_tag"] is None, f"Obtenu : {classified[0]['section_tag']!r}")
    _test("Titre 'First Monitoring Period' -> temporal_layer_1",
          classified[1]["section_tag"] == "temporal_layer_1", f"Obtenu : {classified[1]['section_tag']!r}")
    _test("Chunk suivant (même période) -> temporal_layer_1 persiste",
          classified[2]["section_tag"] == "temporal_layer_1", f"Obtenu : {classified[2]['section_tag']!r}")
    _test("Titre 'Third Monitoring Period' -> temporal_layer_3",
          classified[3]["section_tag"] == "temporal_layer_3", f"Obtenu : {classified[3]['section_tag']!r}")
    _test("Chunk suivant (période 3) -> temporal_layer_3 persiste",
          classified[4]["section_tag"] == "temporal_layer_3", f"Obtenu : {classified[4]['section_tag']!r}")

    # --- Fail-open : pas de titre de période -> tous None ---
    chunks_no_period = [
        {"text": "Le projet respecte les normes environnementales en vigueur.", "page": 1},
        {"text": "Aucune plainte n'a été enregistrée sur la période.", "page": 2},
    ]
    classified_no_period = grid_sections.classify_chunks(chunks_no_period)
    _test("Pas de titre de période détecté -> tous les chunks à None",
          all(c["section_tag"] is None for c in classified_no_period))

    # --- Priorité : exclusion ESAP/plaintes IFC toujours prioritaire ---
    chunks_esap_wins = [
        {"text": "Third Monitoring Period", "page": 20},
        {"text": "Environmental and Social Action Plan\nItem 1: Prepare a Livelihood Restoration Plan", "page": 21},
    ]
    classified_esap_wins = grid_sections.classify_chunks(chunks_esap_wins)
    _test("Titre ESAP après un titre de période -> 'esap' prioritaire (pas temporal_layer)",
          classified_esap_wins[1]["section_tag"] == "esap",
          f"Obtenu : {classified_esap_wins[1]['section_tag']!r}")

    # --- get_chunks_for_question n'exclut PAS les couches temporelles ---
    r_chunks = grid_sections.get_chunks_for_question(classified, "A.1.1")
    _test("Couches temporelles non exclues des chunks R",
          len([c for c in r_chunks if c["section_tag"] == "temporal_layer_1"]) > 0)
    a_chunks = grid_sections.get_chunks_for_question(classified, "A.1.1", for_mitigation=True)
    _test("Couches temporelles non exclues des chunks A (mitigation)",
          len([c for c in a_chunks if c["section_tag"] == "temporal_layer_3"]) > 0)


def test_grid_analyze_v4():
    """Tests d'intégration de l'orchestrateur V4, architecture 2 passes
    (grid_analyze.py, directive CC-V4-12). Backend LLM monkey-patché —
    pas d'appel réseau réel."""
    print("\n--- 1.15 Orchestrateur Grille V4 — 2 passes (grid_analyze.py, CC-V4-12) ---\n")

    import json as json_module

    import config
    import llm_backend
    import grid_result
    import grid_analyze

    # 1. GRID_V4_ENABLED=False -> None (renommé depuis GRID_V3_ENABLED, CC-V4-08)
    _orig_enabled = config.GRID_V4_ENABLED
    config.GRID_V4_ENABLED = False
    try:
        _test("GRID_V4_ENABLED=False -> analyze_grid() retourne None",
              grid_analyze.analyze_grid([{"text": "peu importe", "page": 1}]) is None)
    finally:
        config.GRID_V4_ENABLED = _orig_enabled

    # --- Backend LLM simulé, déterministe, 2 passes JSON : found=False
    # partout (Passe 2 jamais appelée), SAUF B.3.1 (OUI = risque, schéma
    # STANDARD depuis CC-V4-11) avec une mitigation prouvée en Passe 2.
    # Distinction Passe 1/Passe 2 : "VERBATIM EXTRAIT" n'apparaît que dans
    # le prompt de qualification (cf. grid_prompts._QUALIFICATION_PROMPT).
    # Distinction B.3.1 : sa formulation R ("Absence de données de
    # référence") apparaît dans les DEUX passes (question_r est injecté
    # dans les deux templates).
    def _fake_dispatch(backend, prompt, model, options, timeout, response_format=None):
        # Synthèse finale (directive "évolutions pipeline ESG/risk",
        # 2026-08-20) : analyze_grid() appelle désormais aussi cette passe
        # à la fin — cf. test_grid_analyze_synthesis pour les tests dédiés,
        # ici on répond juste un texte plausible pour ne pas polluer
        # result["synthesis"] avec le JSON destiné aux 2 autres passes.
        if "analyste risque chargé de rédiger une synthèse" in prompt:
            return "Synthèse factice (test 1.15, non vérifiée ici)."

        is_qualification = "VERBATIM EXTRAIT" in prompt
        is_b31 = "Absence de données de référence" in prompt

        if not is_qualification:
            if is_b31:
                return json_module.dumps({
                    "code": "B.3.1", "found": True,
                    "verbatim": "aucune etude de reference socio-economique disponible",
                    "page": 10, "subject": "SPV", "brief": "absence de baseline",
                })
            return json_module.dumps({
                "code": "X", "found": False, "verbatim": None, "page": None,
                "subject": None, "brief": "rien trouve",
            })

        # Passe 2 — uniquement atteinte si found=True (donc B.3.1 ici)
        return json_module.dumps({
            "code": "B.3.1", "status": "OUI", "confidence": "HIGH",
            "mitigation_status": "OUI_PROUVEE",
            "verbatim_r": "aucune etude de reference socio-economique disponible",
            "verbatim_a_mesure": "verification tierce independante realisee",
            "verbatim_a_defaillance": None,
            "brief_r": "absence de baseline", "brief_a": "verification tierce",
        })

    chunks = [{"text": "Passage générique pertinent pour l'analyse du projet.", "page": 5}]

    _orig_dispatch = llm_backend._dispatch
    config.GRID_V4_ENABLED = True
    llm_backend._dispatch = _fake_dispatch
    try:
        result = grid_analyze.analyze_grid(chunks, document_type=2)

        # 2. document_type et reading_mode dans le résultat
        _test("document_type=2 reporté dans le résultat", result["document_type"] == 2,
              f"Obtenu : {result.get('document_type')}")
        _test("reading_mode='suivi' (document_type=2)", result["reading_mode"] == "suivi",
              f"Obtenu : {result.get('reading_mode')}")

        # 3. 12 questions dans le résultat
        _test("12 questions dans le résultat", len(result["questions"]) == 12,
              f"Obtenu : {len(result['questions'])}")

        # 5. B.3.1 OUI (standard, CC-V4-11) -> sous-question A active dans le résultat
        b31 = next(q for q in result["questions"] if q["code"] == "B.3.1")
        _test("B.3.1 OUI -> status='OUI' dans le résultat", b31["status"] == "OUI",
              f"Obtenu : {b31['status']!r}")
        _test("B.3.1 OUI -> sous-question A active (mitigation_status='OUI_PROUVEE')",
              b31["mitigation_status"] == "OUI_PROUVEE", f"Obtenu : {b31['mitigation_status']!r}")
        _test("B.3.1 OUI + mitigation prouvée -> gain=3", b31["gain"] == 3, f"Obtenu : {b31['gain']}")

        try:
            grid_result.validate_grid_result(result)
            _test("Résultat de analyze_grid() -> validate_grid_result ne lève pas", True)
        except ValueError as e:
            _test("Résultat de analyze_grid() -> validate_grid_result ne lève pas", False, str(e))

        # 4. na_modules=["B.2"] -> 2 questions NA (B.2.1/B.2.2, CC-V4-11)
        result_na = grid_analyze.analyze_grid(chunks, na_modules=["B.2"], document_type=1)
        na_count = sum(1 for q in result_na["questions"] if q["status"] == "NA")
        _test("na_modules=['B.2'] -> 2 questions NA", na_count == 2, f"Obtenu : {na_count}")
        _test("na_modules=['B.2'] -> scoring.questions_na=2",
              result_na["scoring"]["questions_na"] == 2, f"Obtenu : {result_na['scoring']['questions_na']}")

        # --- Silence R5 : aucun chunk candidat -> repli sans appel LLM ---
        result_empty = grid_analyze.analyze_grid([], document_type=1)
        a11_empty = next(q for q in result_empty["questions"] if q["code"] == "A.1.1")
        _test("Aucun chunk -> repli silence (A.1.1, événement) -> status='NON'",
              a11_empty["status"] == "NON", f"Obtenu : {a11_empty['status']!r}")
        _test("Aucun chunk -> silence_applied=True", a11_empty["silence_applied"] is True)

        b21_empty = next(q for q in result_empty["questions"] if q["code"] == "B.2.1")
        _test("Aucun chunk -> repli silence (B.2.1, état) -> status='INCONNU'",
              b21_empty["status"] == "INCONNU", f"Obtenu : {b21_empty['status']!r}")
        # Directive "gestion INCONNU" (2026-08-20) : texte d'affichage
        # canonique, jamais de justification inventée sur une absence.
        _test("Aucun chunk -> confidence_note = texte canonique 'Aucun élément n'a été trouvé.'",
              b21_empty["confidence_note"] == "Aucun élément n'a été trouvé.",
              f"Obtenu : {b21_empty['confidence_note']!r}")
        _test("INCONNU (B.2.1) -> pénalité 0 et gain 0 dans le résultat assemblé",
              b21_empty["penalty"] == 0 and b21_empty.get("gain", 0) == 0,
              f"Obtenu : penalty={b21_empty['penalty']}, gain={b21_empty.get('gain')}")

        # --- _SILENCE_CONFIRMS_ABSENCE (B.3.1/B.3.2) : silence -> OUI
        # pénalisant, mais avec un texte DIFFÉRENT du texte canonique
        # générique (retour Elisa 2026-08-21, PDF Vorotan : une pénalité
        # affichée à côté du même texte "Aucun élément trouvé" que des
        # questions à 0 pt ne s'explique pas). ---
        b31_empty = next(q for q in result_empty["questions"] if q["code"] == "B.3.1")
        _test("Aucun chunk -> repli silence (B.3.1, confirms_absence) -> status='OUI'",
              b31_empty["status"] == "OUI", f"Obtenu : {b31_empty['status']!r}")
        _test("Aucun chunk -> B.3.1 silence_applied=True", b31_empty["silence_applied"] is True)
        _test("Aucun chunk -> B.3.1 pénalité=-15 (silence = risque confirmé)",
              b31_empty["penalty"] == -15, f"Obtenu : {b31_empty['penalty']}")
        _test("B.3.1 (confirms_absence) -> confidence_note DIFFÉRENT du texte canonique générique",
              b31_empty["confidence_note"] != "Aucun élément n'a été trouvé.",
              f"Obtenu : {b31_empty['confidence_note']!r}")
        _test("B.3.1 (confirms_absence) -> confidence_note explique le mécanisme en clair (pas de jargon 'R5')",
              "pénalité" in (b31_empty["confidence_note"] or "")
              and "R5" not in (b31_empty["confidence_note"] or ""),
              f"Obtenu : {b31_empty['confidence_note']!r}")

        try:
            grid_result.validate_grid_result(result_empty)
            _test("Résultat 100% silence -> validate_grid_result ne lève pas", True)
        except ValueError as e:
            _test("Résultat 100% silence -> validate_grid_result ne lève pas", False, str(e))

    finally:
        llm_backend._dispatch = _orig_dispatch
        config.GRID_V4_ENABLED = _orig_enabled


def test_grid_synthesis_prompt():
    """Tests du prompt de synthèse finale (grid_prompts.get_synthesis_prompt),
    directive "évolutions pipeline ESG/risk" (2026-08-20) — purement
    déterministe (construction de chaîne), pas de LLM. Réutilise
    _build_fake_v4_result() (test 1.17) plutôt qu'un nouveau générateur."""
    print("\n--- 1.23a Prompt de synthèse finale (grid_prompts.py) ---\n")

    import grid_prompts

    result = _build_fake_v4_result()
    prompt = grid_prompts.get_synthesis_prompt(result)

    _test("Prompt synthèse : rappelle qu'INCONNU ne signifie jamais NON",
          "INCONNU NE SIGNIFIE JAMAIS NON" in prompt)
    _test("Prompt synthèse : contient le score final",
          f"{result['scoring']['score']}/100" in prompt,
          f"score={result['scoring']['score']}")
    _test("Prompt synthèse : contient la couleur finale",
          result["scoring"]["color"] in prompt)

    # A.1.1 = OUI dans _build_fake_v4_result -> doit apparaître avec son verbatim.
    a11 = next(q for q in result["questions"] if q["code"] == "A.1.1")
    _test("Prompt synthèse : code OUI (A.1.1) présent", "A.1.1" in prompt)
    _test("Prompt synthèse : verbatim du OUI présent",
          a11["evidence_r"]["passage"] in prompt)

    # B.2.1 = INCONNU (silence, evidence_r=None) dans _build_fake_v4_result
    # -> texte canonique, jamais présenté comme une conclusion.
    b21 = next(q for q in result["questions"] if q["code"] == "B.2.1")
    _test("B.2.1 (résultat factice) -> status=INCONNU",
          b21["status"] == "INCONNU", f"Obtenu : {b21['status']!r}")
    _test("Prompt synthèse : B.2.1 INCONNU présent avec le texte canonique",
          "B.2.1" in prompt and "Aucun élément n'a été trouvé." in prompt)

    # Les NON ne sont jamais détaillés dans le prompt (hors sujet pour une
    # synthèse de risque, cf. directive "coût / taille du contexte") —
    # B.1.1 = NON dans _build_fake_v4_result, ne doit apparaître nulle part.
    _test("Prompt synthèse : ne détaille pas les NON (hors sujet)",
          "B.1.1" not in prompt)


def test_grid_analyze_synthesis():
    """Tests d'intégration de la passe de synthèse finale dans l'orchestrateur
    (grid_analyze.analyze_grid -> _generate_synthesis), directive "évolutions
    pipeline ESG/risk" (2026-08-20). Backend LLM monkey-patché, pas d'appel
    réseau réel — même idiome que test_grid_analyze_v4 (1.15)."""
    print("\n--- 1.23b Intégration synthèse finale (grid_analyze.py) ---\n")

    import json as json_module

    import config
    import llm_backend
    import grid_analyze

    _SYNTHESIS_MARKER = "analyste risque chargé de rédiger une synthèse"
    _orig_dispatch = llm_backend._dispatch
    _orig_v4_enabled = config.GRID_V4_ENABLED
    _orig_synth_enabled = config.GRID_SYNTHESIS_ENABLED
    config.GRID_V4_ENABLED = True

    # --- 6. Échec de l'appel LLM de synthèse -> analyse et score restent
    # disponibles (fail-open, ne modifie ni les questions ni le score déjà
    # calculés avant cet appel). ---
    def _fake_dispatch_synthesis_fails(backend, prompt, model, options, timeout, response_format=None):
        if _SYNTHESIS_MARKER in prompt:
            raise RuntimeError("simulated LLM outage (synthesis)")
        return json_module.dumps({
            "code": "X", "found": False, "verbatim": None, "page": None,
            "subject": None, "brief": None,
        })

    config.GRID_SYNTHESIS_ENABLED = True
    llm_backend._dispatch = _fake_dispatch_synthesis_fails
    try:
        result_fail = grid_analyze.analyze_grid([], document_type=1)
        _test("Échec appel LLM synthèse -> result['synthesis'] est None (fail-open)",
              result_fail["synthesis"] is None)
        # Aucun chunk -> silence sur les 12 questions : B.3.1/B.3.2 = OUI
        # (_SILENCE_CONFIRMS_ABSENCE, le silence EST le risque qu'elles
        # posent), -15 chacune, le reste = NON/INCONNU sans pénalité ->
        # score = 100 - 15 - 15 = 70. Valeur précise plutôt qu'un simple
        # "not None" : prouve que le scoring a bien tourné normalement,
        # pas juste qu'un champ existe.
        _test("Échec appel LLM synthèse -> le score reste disponible et correct",
              result_fail["scoring"]["score"] == 70, f"Obtenu : {result_fail['scoring']['score']}")
        _test("Échec appel LLM synthèse -> les 12 questions restent disponibles",
              len(result_fail["questions"]) == 12, f"Obtenu : {len(result_fail['questions'])}")
    finally:
        llm_backend._dispatch = _orig_dispatch

    # --- GRID_SYNTHESIS_ENABLED=False -> aucun appel de synthèse (le flag
    # coupe la passe indépendamment de GRID_V4_ENABLED). ---
    _synthesis_calls_when_disabled = {"n": 0}

    def _fake_dispatch_count_synthesis(backend, prompt, model, options, timeout, response_format=None):
        if _SYNTHESIS_MARKER in prompt:
            _synthesis_calls_when_disabled["n"] += 1
            return "ne devrait jamais être appelé"
        return json_module.dumps({
            "code": "X", "found": False, "verbatim": None, "page": None,
            "subject": None, "brief": None,
        })

    config.GRID_SYNTHESIS_ENABLED = False
    llm_backend._dispatch = _fake_dispatch_count_synthesis
    try:
        result_disabled = grid_analyze.analyze_grid([], document_type=1)
        _test("GRID_SYNTHESIS_ENABLED=False -> result['synthesis'] est None",
              result_disabled["synthesis"] is None)
        _test("GRID_SYNTHESIS_ENABLED=False -> aucun appel LLM de synthèse",
              _synthesis_calls_when_disabled["n"] == 0,
              f"Obtenu : {_synthesis_calls_when_disabled['n']} appel(s)")
    finally:
        llm_backend._dispatch = _orig_dispatch

    # --- 4/7/8 : dossier avec un OUI -> synthèse générée, UN SEUL appel,
    # généré APRÈS le scoring (le score capturé dans le prompt doit
    # correspondre au score réellement calculé pour ce résultat, pas à un
    # score par défaut/vide — la seule façon d'avoir le bon score dans le
    # prompt est que le scoring ait déjà tourné). ---
    _captured_synthesis_prompts = []

    def _fake_dispatch_ok(backend, prompt, model, options, timeout, response_format=None):
        if _SYNTHESIS_MARKER in prompt:
            _captured_synthesis_prompts.append(prompt)
            return "Synthèse factice générée pour le test."

        is_qualification = "VERBATIM EXTRAIT" in prompt
        if not is_qualification:
            if "Blocage physique du site" in prompt:  # A.1.1 (Passe 1)
                return json_module.dumps({
                    "code": "A.1.1", "found": True, "verbatim": "greve active sur le site",
                    "page": 5, "subject": "SPV", "brief": "greve en cours",
                })
            return json_module.dumps({
                "code": "X", "found": False, "verbatim": None, "page": None,
                "subject": None, "brief": None,
            })
        # Passe 2 — uniquement atteinte pour A.1.1 (found=True ci-dessus)
        return json_module.dumps({
            "code": "A.1.1", "status": "OUI", "confidence": "HIGH",
            "mitigation_status": "NON_INTENTION",
            "verbatim_r": "greve active sur le site",
            "verbatim_a_mesure": None, "verbatim_a_defaillance": None,
            "brief_r": "greve en cours", "brief_a": None,
        })

    chunks = [{"text": "Passage générique pertinent pour l'analyse du projet.", "page": 5}]

    config.GRID_SYNTHESIS_ENABLED = True
    llm_backend._dispatch = _fake_dispatch_ok
    try:
        result_ok = grid_analyze.analyze_grid(chunks, document_type=1)

        _test("Dossier avec un OUI -> synthèse générée (non None)",
              result_ok["synthesis"] == "Synthèse factice générée pour le test.",
              f"Obtenu : {result_ok['synthesis']!r}")
        _test("Un seul appel de synthèse par dossier",
              len(_captured_synthesis_prompts) == 1,
              f"Obtenu : {len(_captured_synthesis_prompts)} appel(s)")

        a11_ok = next(q for q in result_ok["questions"] if q["code"] == "A.1.1")
        _test("A.1.1 (fake dispatch) -> status=OUI",
              a11_ok["status"] == "OUI", f"Obtenu : {a11_ok['status']!r}")

        prompt_used = _captured_synthesis_prompts[0]
        _test("Synthèse générée APRÈS le scoring (score du résultat = score dans le prompt)",
              f"{result_ok['scoring']['score']}/100" in prompt_used,
              f"Score attendu dans le prompt : {result_ok['scoring']['score']}/100")

        # --- 2. Un critère NON avec preuve (silence événement) ne doit pas
        # devenir INCONNU. ---
        a12 = next(q for q in result_ok["questions"] if q["code"] == "A.1.2")
        _test("A.1.2 (evenement, silence) -> NON, jamais INCONNU",
              a12["status"] == "NON", f"Obtenu : {a12['status']!r}")
    finally:
        llm_backend._dispatch = _orig_dispatch
        config.GRID_V4_ENABLED = _orig_v4_enabled
        config.GRID_SYNTHESIS_ENABLED = _orig_synth_enabled


# ============================================================================
# Tests de calibration V4 — 4 dossiers de référence (directive CC-V4-07,
# RECONSTRUITS pour CC-V4-11 sur les 12 codes de la Maquette Vierge)
#
# B.3.1 : POLARITÉ STANDARD (OUI = risque), pas la polarité inversée de
# l'ancien B.3.1 biodiversité — cf. grid_questions.py, vérifié directement
# contre `1_Maquette_Vierge_Grille_ESG (1).pdf` (page 2 : B.3.1 formaté
# EXACTEMENT comme les 11 autres questions, "Si OUI : -15 pts", mitigation
# "si R = OUI"). La directive CC-V4-11 donnait ce cas "B.3.1=NON
# (inversé, -15)" pour CBG/Aysha/Indorama mais "B.3.1=OUI mit=NON" (SANS
# inversion) pour Mundra — incohérence interne de la directive. Réglée en
# faveur de la Maquette Vierge (source de vérité, cf. AUDIT_PERTINENCE_
# NOTE_CADRAGE.md) : les 3 dossiers "(inversé)" sont réencodés NON->OUI
# pour préserver la MÊME pénalité (-15) avec la polarité correcte —
# l'arithmétique ne change pas, seul le sens de l'encodage change.
# ============================================================================

def _cbg_answers_v4():
    """CBG Expansion — Type 1. Score attendu : 28-31."""
    import grid_questions
    base = {q["code"]: {"status": "NON"} for q in grid_questions.QUESTIONS}
    base["B.1.1"] = {"status": "OUI", "mitigation_status": "NON_INTENTION"}
    base["B.1.2"] = {"status": "OUI", "mitigation_status": "OUI_PROUVEE"}
    base["B.2.1"] = {"status": "OUI", "mitigation_status": "NON_FORME_INSUFFISANTE"}
    base["B.2.2"] = {
        "status": "OUI", "mitigation_status": "OUI_DEFAILLANTE",
        "evidence_a": {
            "verbatim_mesure": "additional filters and scrubbers were installed on the dryer stack in recent years",
            "verbatim_defaillance": "generated significant quantities of fugitive dust as well as particulate and gaseous emissions",
            "page": None,
        },
    }
    base["B.3.1"] = {"status": "OUI"}  # standard (CC-V4-11) : absence de baseline avérée, -15
    return base
    # Calcul : -15 -15 -15 -15 -15 = -75, gains = +3 (B.1.2), score = 28


def _mundra_answers_v4():
    """CGPL Mundra — Type 3.

    FRAGILE (CC-V4-11, PENDING_ELISA) : réponses reprises TELLES QUELLES
    depuis la directive CC-V4-11 (BLOC A, point 3, "issues de
    l'annotation manuelle Stacy"). Avec les 12 codes de la Maquette
    Vierge (B.4.1 "impacts sanitaires" n'existe plus, cf. BLOC A), le
    calcul de CE jeu de réponses donne score=29 (ORANGE), PAS 16 (ROUGE)
    comme l'affirment le reste de la directive (BLOC A intro, BLOC E) —
    la directive elle-même reconnaît cet écart dans son BLOC C ("Mais la
    calibration V4 dit 16 (...) revérifie que les réponses Mundra
    donnent le bon score") sans le résoudre. La pénalité manquante
    (-15, jusqu'à 16) correspondait à l'ancien B.4.1 (irritations
    cutanées liées au rejet de la centrale) — fait réel du dossier CGPL/
    Tata Mundra qui n'a PLUS de question d'accueil évidente dans les 12
    codes Maquette Vierge (B.2.2 "Eau" porte sur un défaut de
    MODÉLISATION, pas sur un dommage sanitaire constaté). Cf. BLOC E de
    la directive : "NE corrige PAS automatiquement — rapporte l'écart."
    Décision requise d'Elisa/Stacy : quelle question de la Maquette
    Vierge doit porter ce fait (B.1.1 ? B.2.2 ? aucune ?) avant de
    pouvoir recalibrer ce jeu de réponses sur 16.
    """
    import grid_questions
    base = {q["code"]: {"status": "NON"} for q in grid_questions.QUESTIONS}
    base["A.1.1"] = {"status": "OUI", "mitigation_status": "OUI_PROUVEE"}
    base["B.1.1"] = {"status": "OUI", "mitigation_status": "OUI_PROUVEE"}
    base["B.2.1"] = {"status": "OUI", "mitigation_status": "OUI_PROUVEE"}
    base["B.2.2"] = {"status": "OUI", "mitigation_status": "OUI_PROUVEE"}
    base["B.3.1"] = {"status": "OUI", "mitigation_status": "NON_INTENTION"}  # standard (CC-V4-11)
    return base
    # Pénalités : A.1.1(-25) B.1.1(-15) B.2.1(-15) B.2.2(-15) B.3.1(-15) = -85
    # Gains : A.1.1(+5) B.1.1(+3) B.2.1(+3) B.2.2(+3) = +14 ; score = 100 - 85 + 14 = 29
    # (directive : 16 attendu — écart non résolu, cf. docstring PENDING_ELISA ci-dessus)


def _aysha_answers_v4():
    """Aysha I Wind — Type 1. Score attendu : 73 (panachage R11)."""
    import grid_questions
    base = {q["code"]: {"status": "NON"} for q in grid_questions.QUESTIONS}
    base["B.1.1"] = {"status": "OUI", "mitigation_status": "OUI_PROUVEE"}  # date butoir = statut 3 (R11 adapté)
    base["B.3.1"] = {"status": "OUI"}  # standard (CC-V4-11) : BAP absent = strict, -15
    # B.1.2 = NON ; B.2.x = NON (pas d'installation)
    return base
    # Pénalités : B.1.1(-15) B.3.1(-15) = -30 ; Gains : B.1.1(+3) = +3 ; score = 100 - 30 + 3 = 73


def _indorama_answers_v4():
    """Indorama IFF Line 2 — Type 1. Score attendu : 85 (lecture stricte B.3.1)."""
    import grid_questions
    base = {q["code"]: {"status": "NON"} for q in grid_questions.QUESTIONS}
    base["B.1.1"] = {"status": "NA"}  # N/A argumenté (PS5 non déclenchée)
    base["B.3.1"] = {"status": "OUI"}  # standard (CC-V4-11) : CHA aboutie mais lecture stricte encodée -> OUI
    return base
    # Pénalités : B.3.1(-15) = -15 ; score = 100 - 15 = 85


def test_calibration_cbg():
    """CBG Expansion — Type 1, score attendu 28-31, ORANGE/ROUGE."""
    print("\n--- 1.16a Calibration CBG Expansion (Type 1) ---\n")
    import grid_scoring
    answers = _cbg_answers_v4()
    result = grid_scoring.compute_grid_score(answers, document_type=1)
    _test("CBG : score entre 28 et 31", 28 <= result["score"] <= 31, f"Obtenu : {result['score']}")
    _test("CBG : color = ORANGE ou ROUGE (zone critique)",
          result["color"] in ("ORANGE", "ROUGE"), f"Obtenu : {result['color']!r}")
    _test("CBG : reading_mode='instruction'", result["reading_mode"] == "instruction",
          f"Obtenu : {result['reading_mode']!r}")


def test_calibration_mundra():
    """CGPL Mundra — Type 3.

    PENDING_ELISA (CC-V4-11) : la directive vise 16/ROUGE, mais avec les
    12 codes de la Maquette Vierge (B.4.1 disparu, cf. BLOC A) les
    réponses données par la directive elle-même calculent 29/ORANGE —
    cf. docstring de _mundra_answers_v4() pour le détail. Ce test
    vérifie le résultat RÉEL du calcul (29/ORANGE), pas le chiffre visé
    par la directive, conformément à sa propre consigne BLOC E ("NE
    corrige PAS automatiquement — rapporte l'écart"). À corriger dès
    qu'Elisa tranche où reloger le fait sanitaire de l'ancien B.4.1.
    """
    print("\n--- 1.16b Calibration Mundra CGPL (Type 3) — PENDING_ELISA, cf. _mundra_answers_v4 ---\n")
    import grid_scoring
    answers = _mundra_answers_v4()
    result = grid_scoring.compute_grid_score(answers, document_type=3)
    _test("Mundra : score = 29 (PAS 16 — écart connu, cf. PENDING_ELISA)",
          result["score"] == 29, f"Obtenu : {result['score']}")
    _test("Mundra : color = ORANGE (PAS ROUGE — écart connu, cf. PENDING_ELISA)",
          result["color"] == "ORANGE", f"Obtenu : {result['color']!r}")
    _test("Mundra : reading_mode='suivi'", result["reading_mode"] == "suivi",
          f"Obtenu : {result['reading_mode']!r}")


def test_calibration_aysha():
    """Aysha I Wind — Type 1, score attendu 73, JAUNE."""
    print("\n--- 1.16c Calibration Aysha I Wind (Type 1) ---\n")
    import grid_scoring
    answers = _aysha_answers_v4()
    result = grid_scoring.compute_grid_score(answers, document_type=1)
    _test("Aysha : score = 73", result["score"] == 73, f"Obtenu : {result['score']}")
    _test("Aysha : color = JAUNE", result["color"] == "JAUNE", f"Obtenu : {result['color']!r}")


def test_calibration_indorama():
    """Indorama IFF Line 2 — Type 1, score attendu 85, VERT."""
    print("\n--- 1.16d Calibration Indorama IFF Line 2 (Type 1) ---\n")
    import grid_scoring
    answers = _indorama_answers_v4()
    result = grid_scoring.compute_grid_score(answers, document_type=1)
    _test("Indorama : score = 85", result["score"] == 85, f"Obtenu : {result['score']}")
    _test("Indorama : color = VERT", result["color"] == "VERT", f"Obtenu : {result['color']!r}")


def test_calibration_ordering():
    """L'ordre Indorama > Aysha > (CBG, Mundra) doit être respecté.

    PENDING_ELISA (CC-V4-11) : l'ordre historique complet (Indorama >
    Aysha > CBG > Mundra, Mundra le pire des 4) ne tient PLUS avec les
    12 codes Maquette Vierge : Mundra calcule 29 > CBG 28 (cf.
    docstring de _mundra_answers_v4 — écart connu, pas silencieusement
    corrigé). On vérifie donc seulement ce qui reste vrai (Indorama et
    Aysha restent nettement devant les deux dossiers à risque), pas
    l'ordre strict CBG/Mundra tant que la question PENDING_ELISA n'est
    pas tranchée."""
    print("\n--- 1.16e Calibration : ordre relatif des 4 dossiers (partiel, cf. PENDING_ELISA) ---\n")
    import grid_scoring
    s_ind = grid_scoring.compute_grid_score(_indorama_answers_v4(), document_type=1)["score"]
    s_ays = grid_scoring.compute_grid_score(_aysha_answers_v4(), document_type=1)["score"]
    s_cbg = grid_scoring.compute_grid_score(_cbg_answers_v4(), document_type=1)["score"]
    s_mun = grid_scoring.compute_grid_score(_mundra_answers_v4(), document_type=3)["score"]
    _test("Ordre Indorama > Aysha", s_ind > s_ays, f"Obtenu : {s_ind}, {s_ays}")
    _test("Ordre Aysha > CBG et Aysha > Mundra (les 2 dossiers à risque)",
          s_ays > s_cbg and s_ays > s_mun, f"Obtenu : Aysha={s_ays}, CBG={s_cbg}, Mundra={s_mun}")


def test_calibration_shared_cap_mundra():
    """Plafond partagé DORMANT depuis CC-V4-11 (A.1.3 n'existe plus, aucune
    question ne porte plus shared_cap_group) — vérifié sur les 4 dossiers
    de calibration, pas seulement Mundra."""
    print("\n--- 1.16f Calibration : plafond partagé dormant sur les 4 dossiers (CC-V4-11) ---\n")
    import grid_scoring
    for fn, doc_type, label in [
        (_cbg_answers_v4, 1, "CBG"),
        (_mundra_answers_v4, 3, "Mundra"),
        (_aysha_answers_v4, 1, "Aysha"),
        (_indorama_answers_v4, 1, "Indorama"),
    ]:
        result = grid_scoring.compute_grid_score(fn(), document_type=doc_type)
        _test(f"{label} : aucun _shared_cap_applied (mécanisme dormant, CC-V4-11)",
              not any(d.get("_shared_cap_applied") for d in result["details"]))


def test_calibration_b31_inverted():
    """B.3.1 : polarité STANDARD sur les 4 dossiers depuis CC-V4-11 (OUI =
    risque, cf. grid_questions.py — vérifié contre la Maquette Vierge,
    plus de schéma inversé). Renommé conceptuellement (le nom de la
    fonction reste pour ne pas casser l'appel dans run_all_tests, mais le
    contenu teste désormais l'ABSENCE d'inversion)."""
    print("\n--- 1.16g Calibration : B.3.1 polarité standard sur les 4 dossiers (CC-V4-11) ---\n")
    for fn, expected, label in [
        (_cbg_answers_v4, "OUI", "CBG"),
        (_mundra_answers_v4, "OUI", "Mundra"),
        (_aysha_answers_v4, "OUI", "Aysha (strict : BAP absent)"),
        (_indorama_answers_v4, "OUI", "Indorama (lecture stricte encodée pour score=85)"),
    ]:
        _test(f"{label} : B.3.1 = {expected} (standard, CC-V4-11)", fn()["B.3.1"]["status"] == expected,
              f"Obtenu : {fn()['B.3.1']['status']!r}")


def _build_fake_v4_result():
    """Construit un result_v4 réaliste via grid_scoring + grid_result (pas
    un dict tapé à la main) — pour tester export.py sans jamais dériver du
    contrat réel de grid_result.build_grid_result (CC-V4-03). Couvre les
    cas intéressants pour l'export : OUI_DEFAILLANTE (double verbatim),
    B.3.1 avec mitigation prouvée, INCONNU (silence), qualifying (R10)."""
    import grid_questions
    import grid_scoring
    import grid_result

    answers = {q["code"]: {"status": "NON"} for q in grid_questions.QUESTIONS}
    answers["A.1.1"] = {"status": "OUI", "mitigation_status": "NON_INTENTION"}
    answers["A.2.1"] = {"status": "OUI", "mitigation_status": "OUI_PROUVEE"}
    answers["B.2.1"] = {"status": "INCONNU"}
    answers["B.2.2"] = {
        "status": "OUI", "mitigation_status": "OUI_DEFAILLANTE",
        "evidence_a": {
            "verbatim_mesure": "des filtres ont ete installes sur la cheminee",
            "verbatim_defaillance": "des poussieres persistent malgre les filtres",
            "page": 34,
        },
    }
    answers["B.3.1"] = {"status": "OUI", "mitigation_status": "OUI_PROUVEE"}  # standard (CC-V4-11)

    question_results = {}
    for code, ans in answers.items():
        entry = dict(ans)
        entry["evidence_r"] = (
            {"passage": f"Passage justificatif pour {code}.", "page": 12}
            if ans["status"] in ("OUI", "NON") else None
        )
        entry["confidence_note"] = "Formulation ambigue dans le rapport." if code == "A.1.1" else None
        entry["qualifying"] = (
            {"subject_filter": "lender", "lender_supervision": "Verbatim vise le preteur, pas la SPV."}
            if code == "A.2.1" else None
        )
        question_results[code] = entry

    scoring = grid_scoring.compute_grid_score(answers, document_type=1)
    return grid_result.build_grid_result(question_results, scoring)


def test_grid_v4_export():
    """Tests des exports PDF/Excel de la Grille V4 (export.py, directive
    CC-V4-10). Purement déterministe (fpdf2/openpyxl), pas de LLM."""
    print("\n--- 1.17 Export PDF/Excel Grille V4 (export.py) ---\n")

    import export
    import grid_result

    result = _build_fake_v4_result()
    try:
        grid_result.validate_grid_result(result)
        _test("Résultat factice V4 -> validate_grid_result ne lève pas", True)
    except ValueError as e:
        _test("Résultat factice V4 -> validate_grid_result ne lève pas", False, str(e))

    # --- PDF ---
    pdf_bytes = export.build_grid_v4_pdf(result, project_name="Test")
    _test("build_grid_v4_pdf : retourne des bytes non vides",
          pdf_bytes is not None and len(pdf_bytes) > 0)
    _test("build_grid_v4_pdf : signature %PDF", pdf_bytes[:4] == b"%PDF",
          f"Obtenu : {pdf_bytes[:4]!r}")

    pdf_bytes_no_name = export.build_grid_v4_pdf(result)
    _test("build_grid_v4_pdf : fonctionne aussi sans project_name",
          pdf_bytes_no_name[:4] == b"%PDF")

    # --- Excel ---
    xlsx_bytes = export.build_grid_v4_excel(result, project_name="Test")
    _test("build_grid_v4_excel : retourne des bytes non vides",
          xlsx_bytes is not None and len(xlsx_bytes) > 0)
    # Les fichiers .xlsx sont des archives ZIP -> signature "PK"
    _test("build_grid_v4_excel : signature PK (format ZIP/xlsx)",
          xlsx_bytes[:2] == b"PK", f"Obtenu : {xlsx_bytes[:2]!r}")

    # --- Contenu Excel : 3 feuilles (Synthese/Grille/Detail), remplissage
    # conditionnel — structure maquette Elisa 2026-08-20 (remplace l'ancien
    # découpage à 4 feuilles Synthese/Grille/Evidence/Qualifiants, cf.
    # directive refonte exports).
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    _test("Excel : 3 feuilles (Synthese, Grille, Detail)",
          wb.sheetnames == ["Synthese", "Grille", "Detail"],
          f"Obtenu : {wb.sheetnames}")

    ws_grille = wb["Grille"]
    _test("Excel/Grille : 12 lignes de données + 1 en-tête",
          ws_grille.max_row == 13, f"Obtenu : {ws_grille.max_row}")

    # A.1.1 est en ligne 2 (ordre de grid_questions.QUESTIONS), status=OUI
    # -> la cellule Statut (colonne 3) doit avoir le fond rouge clair
    # (remplissage conditionnel sur le Statut, pas sur la Pénalité).
    statut_cell = ws_grille.cell(row=2, column=3)
    _test("Excel/Grille : fond rouge clair si Statut=OUI (A.1.1)",
          statut_cell.fill.start_color.rgb in ("00F8D7DA", "F8D7DA"),
          f"Obtenu : {statut_cell.fill.start_color.rgb!r}")

    ws_detail = wb["Detail"]
    _test("Excel/Detail : au moins une ligne de données",
          ws_detail.max_row > 1, f"Obtenu : {ws_detail.max_row}")
    _test("Excel/Detail : colonnes attendues",
          [c.value for c in ws_detail[1]] == [
              "Code", "Statut", "Verbatim risque", "Page",
              "Verbatim mitigation", "Defaillance", "Confiance",
          ])


def test_pdf_question_sentence():
    """Phrase narrative statut/mitigation/impact (export._question_status_
    sentence), remplace les lignes brutes 'Statut : OUI' / 'Mitigation :
    NON - forme insuffisante' / 'Penalite : -15 | Gain : +0 | Net : -15'
    dans la section 'Detail par question' du PDF (cf. retour Elisa
    2026-08-20, "en faire un vrai rapport" — même démarche que la phrase
    de contexte dossier, grid_display._format_context_sentence). N'est
    appelée que pour status OUI/INCONNU (cf. docstring de la fonction :
    NA/NON n'apparaissent jamais dans cette section du PDF). Purement
    déterministe (segments testés directement), pas de LLM."""
    print("\n--- 1.17a Phrase narrative statut/mitigation (export.py, PDF) ---\n")

    import export

    def _seg_text(segments):
        return "".join(t for t, _ in segments)

    # --- INCONNU : pas de détail mitigation, pénalité nulle ---
    q_inconnu = {"status": "INCONNU", "penalty": 0, "gain": 0, "mitigation_label": None}
    text = _seg_text(export._question_status_sentence(q_inconnu))
    _test("INCONNU : phrase mentionne l'absence de conclusion", "ne permettent pas de conclure" in text)
    _test("INCONNU : aucune pénalité mentionnée", "Aucune penalite appliquee" in text)

    # --- OUI, mitigation jamais évaluée (mitigation_label=None) ---
    q_oui_sans_mitigation = {"status": "OUI", "penalty": -25, "gain": 0, "mitigation_label": None}
    segs = export._question_status_sentence(q_oui_sans_mitigation)
    text = _seg_text(segs)
    _test("OUI sans mitigation : phrase mentionne l'absence d'évaluation",
          "Aucune mitigation n'a ete evaluee" in text)
    _test("OUI sans mitigation : pénalité en gras", any(b and "-25" in t for t, b in segs))

    # --- OUI, mitigation NON - forme insuffisante (gain=0) ---
    q_oui_mitigation_non = {
        "status": "OUI", "penalty": -15, "gain": 0,
        "mitigation_label": "NON — forme insuffisante",
    }
    segs = export._question_status_sentence(q_oui_mitigation_non)
    text = _seg_text(segs)
    _test("OUI + mitigation insuffisante : label mentionné", "NON — forme insuffisante" in text)
    _test("OUI + mitigation insuffisante : 'aucun gain' mentionné", "aucun gain d'attenuation" in text)
    _test("OUI + mitigation insuffisante : impact net = -15 en gras",
          any(b and "-15" in t for t, b in segs))
    _test("OUI + mitigation insuffisante : mitigation_label rendu en gras",
          any(b and t == "NON — forme insuffisante" for t, b in segs))

    # --- OUI, mitigation OUI - prouvée (gain>0) ---
    q_oui_mitigation_ok = {
        "status": "OUI", "penalty": -25, "gain": 5,
        "mitigation_label": "OUI — prouvee",
    }
    segs = export._question_status_sentence(q_oui_mitigation_ok)
    text = _seg_text(segs)
    _test("OUI + mitigation prouvée : phrase mentionne la mise en place", "mise en place" in text)
    _test("OUI + mitigation prouvée : gain positif en gras", any(b and "+5" in t for t, b in segs))
    _test("OUI + mitigation prouvée : impact net = -20 en gras", any(b and "-20" in t for t, b in segs))

    # --- Smoke test : écriture réelle sur un objet fpdf2, ne lève pas ---
    from fpdf import FPDF
    pdf = FPDF()
    pdf.add_page()
    for q in (q_inconnu, q_oui_sans_mitigation, q_oui_mitigation_non, q_oui_mitigation_ok):
        export._write_question_sentence(pdf, q)
    _test("_write_question_sentence : aucune exception sur les 4 scénarios", True)


def test_grid_display_executive_summary():
    """Groupes de l'Executive Risk Summary (grid_display.py, audit UI
    2026-08-20) : _top_risk_drivers/_needs_verification/_favorable_points
    sont des fonctions pures (pas de Streamlit) qui dérivent 3 groupes des
    mêmes 12 questions — aucune nouvelle donnée, juste un tri/filtre sur
    penalty/gain/status/confidence_note/silence_applied/atteste/
    mitigation_status déjà présents dans le contrat grid_result.py.
    Fixture synthétique à 6 questions couvrant les 3 groupes + les
    exclusions attendues (NA nulle part, NON par silence hors favorable,
    OUI sans doute hors vérification)."""
    print("\n--- 1.17b Groupes Executive Summary (grid_display.py, audit UI) ---\n")

    import grid_display

    q_a11 = {
        "code": "A.1.1", "sous_theme": "Communautaire", "status": "OUI",
        "penalty": -25, "gain": 5, "mitigation_status": "OUI_PROUVEE",
        "mitigation_label": "OUI — prouvée",
        "evidence_r": {"passage": "x" * 250, "page": 3},
        "confidence_note": None, "silence_applied": False, "atteste": False,
    }
    q_b21 = {
        "code": "B.2.1", "sous_theme": "Pollution Air", "status": "OUI",
        "penalty": -15, "gain": 0, "mitigation_status": None, "mitigation_label": None,
        "evidence_r": {"passage": "dust emissions observed", "page": 9},
        "confidence_note": "Le LLM hésite sur la portée exacte du passage.",
        "silence_applied": False, "atteste": False,
    }
    q_b22 = {
        "code": "B.2.2", "sous_theme": "Pollution Eau", "status": "INCONNU",
        "penalty": 0, "gain": 0, "mitigation_status": None, "mitigation_label": None,
        "evidence_r": None, "confidence_note": "Aucun élément n'a été trouvé.",
        "silence_applied": True, "atteste": False,
    }
    q_b31 = {
        "code": "B.3.1", "sous_theme": "Gouvernance", "status": "NON",
        "penalty": 0, "gain": 0, "mitigation_status": None, "mitigation_label": None,
        "evidence_r": {"passage": "no baseline gap identified", "page": 10},
        "confidence_note": None, "silence_applied": False, "atteste": True,
    }
    q_b12 = {
        "code": "B.1.2", "sous_theme": "Griefs", "status": "NON",
        "penalty": 0, "gain": 0, "mitigation_status": None, "mitigation_label": None,
        "evidence_r": None, "confidence_note": None, "silence_applied": True, "atteste": False,
    }
    q_a21 = {
        "code": "A.2.1", "sous_theme": "Faisabilité", "status": "NA",
        "penalty": 0, "gain": 0, "mitigation_status": None, "mitigation_label": None,
        "evidence_r": {"passage": "hors périmètre (na_module)", "page": None},
        "confidence_note": None, "silence_applied": False, "atteste": False,
    }
    questions = [q_a11, q_b21, q_b22, q_b31, q_b12, q_a21]

    # --- Top Risk Drivers : OUI triés par impact net (le plus négatif
    # d'abord) — A.1.1 net=-20 avant B.2.1 net=-15, malgré l'ordre grille
    # inverse (A avant B mais -20 "pèse plus" que -15). ---
    drivers = grid_display._top_risk_drivers(questions)
    _test("Top Risk Drivers : seuls les OUI (2/6)", [q["code"] for q in drivers] == ["A.1.1", "B.2.1"],
          f"Obtenu : {[q['code'] for q in drivers]}")
    _test("Top Risk Drivers : triés par impact net décroissant (-20 avant -15)",
          [grid_display._net(q) for q in drivers] == [-20, -15])

    # --- Points à vérifier : INCONNU + OUI avec doute LLM non-silence
    # (B.2.1) — A.1.1 exclu (pas de confidence_note), B.1.2 exclu (NON). ---
    to_check = grid_display._needs_verification(questions)
    _test("À vérifier : B.2.1 (doute LLM) + B.2.2 (INCONNU), rien d'autre",
          {q["code"] for q in to_check} == {"B.2.1", "B.2.2"},
          f"Obtenu : {[q['code'] for q in to_check]}")

    # --- Points favorables : A.1.1 (mitigation prouvée) + B.3.1 (NON
    # attesté) — B.1.2 exclu (NON par silence, pas de preuve explicite). ---
    favorable = grid_display._favorable_points(questions)
    _test("Favorables : A.1.1 (mitigation prouvée) + B.3.1 (NON attesté), rien d'autre",
          {q["code"] for q in favorable} == {"A.1.1", "B.3.1"},
          f"Obtenu : {[q['code'] for q in favorable]}")

    # --- NA (A.2.1) n'apparaît dans AUCUN des 3 groupes ---
    all_grouped_codes = (
        {q["code"] for q in drivers} | {q["code"] for q in to_check} | {q["code"] for q in favorable}
    )
    _test("NA (A.2.1) absent des 3 groupes de l'Executive Summary", "A.2.1" not in all_grouped_codes)

    # --- Helpers d'affichage : libellé métier avant le code ---
    line = grid_display._question_line(q_a11)
    _test("_question_line : sous_theme avant le code (libellé métier en avant)",
          line.index("Communautaire") < line.index("A.1.1"))
    _test("_question_line : impact net affiché", "-20" in line)

    truncated = grid_display._truncated_verbatim(q_a11["evidence_r"]["passage"])
    _test("_truncated_verbatim : tronqué à 200 caractères + ellipse",
          len(truncated) == 201 and truncated.endswith("…"))
    _test("_truncated_verbatim : None si pas de passage", grid_display._truncated_verbatim(None) is None)


def test_analysis_store():
    """Conservation des analyses sur disque (scripts/analysis_store.py,
    2026-08-20, directive "un seul document + conservation") : save/list/
    load, purement déterministe (pas de LLM), isolé dans un dossier
    temporaire — ne touche JAMAIS data/analyses/ du vrai repo (ANALYSES_DIR
    monkey-patché puis restauré, même idiome que config.LLM_BACKEND
    ailleurs dans ce fichier)."""
    print("\n--- 1.17c Conservation des analyses (analysis_store.py) ---\n")

    import tempfile
    from pathlib import Path
    from datetime import datetime as _dt
    import analysis_store

    fake_result = _build_fake_v4_result()
    _orig_dir = analysis_store.ANALYSES_DIR

    with tempfile.TemporaryDirectory() as tmp:
        analysis_store.ANALYSES_DIR = Path(tmp)
        try:
            # --- save_analysis ---
            analyzed_at = _dt(2026, 8, 20, 14, 30, 0)
            path = analysis_store.save_analysis(
                fake_result, document="CBG Expansion.pdf", documents=["CBG Expansion.pdf"],
                analyzed_at=analyzed_at,
            )
            _test("save_analysis : retourne un Path non None", path is not None)
            _test("save_analysis : le fichier existe sur disque", path is not None and path.exists())

            # --- list_analyses ---
            summaries = analysis_store.list_analyses()
            _test("list_analyses : 1 entrée après 1 sauvegarde", len(summaries) == 1,
                  f"Obtenu : {len(summaries)}")
            summary = summaries[0]
            _test("list_analyses : document correct", summary["document"] == "CBG Expansion.pdf")
            _test("list_analyses : score correct", summary["score"] == fake_result["scoring"]["score"])
            _test("list_analyses : color correct", summary["color"] == fake_result["scoring"]["color"])

            # --- load_analysis : round-trip complet (result_v4 intact) ---
            loaded = analysis_store.load_analysis(summary["path"])
            _test("load_analysis : non None", loaded is not None)
            if loaded:
                _test("load_analysis : result_v4 round-trip identique",
                      loaded["result_v4"]["questions"] == fake_result["questions"])

            # --- Plusieurs analyses : la plus récente en tête ---
            analysis_store.save_analysis(
                fake_result, document="Aysha Wind.pdf", documents=["Aysha Wind.pdf"],
                analyzed_at=_dt(2026, 8, 20, 15, 0, 0),
            )
            summaries2 = analysis_store.list_analyses()
            _test("list_analyses : 2 entrées après 2 sauvegardes", len(summaries2) == 2,
                  f"Obtenu : {len(summaries2)}")
            _test("list_analyses : la plus récente en premier",
                  summaries2[0]["document"] == "Aysha Wind.pdf",
                  f"Obtenu : {summaries2[0]['document']!r}")

            # --- Fichier corrompu : ignoré, pas d'exception (fail-open) ---
            (Path(tmp) / "corrupted.json").write_text("{not valid json", encoding="utf-8")
            summaries3 = analysis_store.list_analyses()
            _test("list_analyses : fichier corrompu ignoré, pas d'exception",
                  len(summaries3) == 2, f"Obtenu : {len(summaries3)}")

            # --- load_analysis sur un fichier manquant : None, pas d'exception ---
            missing = analysis_store.load_analysis(Path(tmp) / "does_not_exist.json")
            _test("load_analysis : None sur un fichier manquant (fail-open)", missing is None)
        finally:
            analysis_store.ANALYSES_DIR = _orig_dir


def test_grid_questions_v4():
    """Vérifie la structure de QUESTIONS (grid_questions.py), RECONSTRUITE
    pour CC-V4-11 sur les 12 codes exacts de la Maquette Vierge : 6+6 par
    catégorie, plus de plafond partagé ni de polarité inversée (les deux
    mécanismes restent dormants dans le code, cf. grid_scoring.py/
    grid_prompts.py, mais aucune question ne les active plus)."""
    print("\n--- 1.6 Grille ESG V4 (grid_questions.py, CC-V4-11) ---\n")

    import grid_questions

    questions = grid_questions.QUESTIONS

    _test("GRID_VERSION='V4'", grid_questions.GRID_VERSION == "V4", f"Obtenu : {grid_questions.GRID_VERSION!r}")
    _test("12 questions exactement", len(questions) == 12, f"Obtenu : {len(questions)}")
    _test("ESG_QUESTIONS alias de QUESTIONS (rétro-compat)",
          grid_questions.ESG_QUESTIONS is grid_questions.QUESTIONS)

    # --- Les 12 codes sont EXACTEMENT ceux de la Maquette Vierge, dans l'ordre ---
    expected_codes = [
        "A.1.1", "A.1.2", "A.2.1", "A.2.2", "A.3.1", "A.3.2",
        "B.1.1", "B.1.2", "B.2.1", "B.2.2", "B.3.1", "B.3.2",
    ]
    actual_codes = [q["code"] for q in questions]
    _test("Les 12 codes sont exactement ceux de la Maquette Vierge, dans l'ordre",
          actual_codes == expected_codes, f"Obtenu : {actual_codes}")

    for removed_code in ("A.1.3", "A.4.1", "B.2.3", "B.4.1"):
        _test(f"{removed_code} n'existe plus (absent de la Maquette Vierge, CC-V4-11)",
              grid_questions.get_question(removed_code) is None)

    cat_a = [q for q in questions if q["category"] == "A"]
    cat_b = [q for q in questions if q["category"] == "B"]
    _test("6 questions Cat A (Maquette Vierge, CC-V4-11)", len(cat_a) == 6, f"Obtenu : {len(cat_a)}")
    _test("6 questions Cat B (Maquette Vierge, CC-V4-11)", len(cat_b) == 6, f"Obtenu : {len(cat_b)}")

    _test("Pénalité -25 pour toutes les questions Cat A",
          all(q["penalty"] == -25 for q in cat_a),
          f"Pénalités Cat A : {[q['penalty'] for q in cat_a]}")
    _test("Pénalité -15 pour toutes les questions Cat B",
          all(q["penalty"] == -15 for q in cat_b),
          f"Pénalités Cat B : {[q['penalty'] for q in cat_b]}")
    _test("Gain +5 pour toutes les questions Cat A", all(q["gain"] == 5 for q in cat_a))
    _test("Gain +3 pour toutes les questions Cat B", all(q["gain"] == 3 for q in cat_b))

    # --- Plafond partagé : DORMANT depuis CC-V4-11 (A.1.3 n'existe plus,
    # les deux sujets sont fusionnés dans un seul A.1.1) ---
    _test("Aucune question ne porte shared_cap_group (mécanisme dormant, CC-V4-11)",
          all(q["shared_cap_group"] is None for q in questions),
          f"Obtenu : {[(q['code'], q['shared_cap_group']) for q in questions if q['shared_cap_group']]}")
    _test("get_questions_by_shared_cap('A.1') renvoie [] (dormant)",
          len(grid_questions.get_questions_by_shared_cap("A.1")) == 0)
    _test("get_questions_by_shared_cap('nonexistent') renvoie []",
          len(grid_questions.get_questions_by_shared_cap("nonexistent")) == 0)

    # --- B.3.1 : polarité STANDARD depuis CC-V4-11 (vérifié contre la
    # Maquette Vierge, page 2 — formatée EXACTEMENT comme les 11 autres
    # questions, aucune inversion) ---
    b31 = grid_questions.get_question("B.3.1")
    _test("B.3.1 existe", b31 is not None, "get_question('B.3.1') renvoie None")
    if b31:
        _test("B.3.1 : inverted_polarity=False (standard, CC-V4-11)", b31["inverted_polarity"] is False,
              f"Obtenu : {b31['inverted_polarity']}")
        _test("B.3.1 : has_separate_r=True", b31["has_separate_r"] is True,
              f"Obtenu : {b31['has_separate_r']}")
        _test("B.3.1 : a_condition='r_oui' (standard, CC-V4-11)", b31["a_condition"] == "r_oui",
              f"Obtenu : {b31['a_condition']!r}")
        _test("B.3.1 : gain=3", b31["gain"] == 3, f"Obtenu : {b31['gain']}")
        _test("B.3.1 : question_r porte sur l'absence de baseline",
              "baseline" in b31["question_r"].lower())

    # --- Aucune question n'a plus inverted_polarity=True (CC-V4-11) ---
    _test("Aucune question n'a inverted_polarity=True (CC-V4-11)",
          all(not q["inverted_polarity"] for q in questions),
          f"Obtenu : {[q['code'] for q in questions if q['inverted_polarity']]}")

    pollution_codes = {"B.2.1", "B.2.2"}
    non_pollution = [q for q in questions if q["code"] not in pollution_codes]

    _test("B.2.1/B.2.2 : na_module='B.2' (2 questions pollution, pas 3, CC-V4-11)",
          all(q["na_module"] == "B.2" for q in questions if q["code"] in pollution_codes),
          f"Obtenu : {[(q['code'], q['na_module']) for q in questions if q['code'] in pollution_codes]}")
    _test("Questions hors module B.2 : na_module=None",
          all(q["na_module"] is None for q in non_pollution),
          f"Obtenu : {[(q['code'], q['na_module']) for q in non_pollution]}")

    # --- silence_type et a_condition sur toutes les questions ---
    _test("Toutes les questions ont un silence_type valide",
          all(q["silence_type"] in ("evenement", "etat") for q in questions),
          f"Obtenu : {[(q['code'], q['silence_type']) for q in questions]}")
    _test("Toutes les questions ont a_condition='r_oui' (schéma unique, CC-V4-11)",
          all(q["a_condition"] == "r_oui" for q in questions),
          f"Obtenu : {[(q['code'], q['a_condition']) for q in questions]}")

    _test("get_question('A.1.1') renvoie un dict",
          isinstance(grid_questions.get_question("A.1.1"), dict))
    _test("get_question('Z.9.9') renvoie None",
          grid_questions.get_question("Z.9.9") is None)

    _test("get_active_questions() renvoie 12 questions",
          len(grid_questions.get_active_questions()) == 12,
          f"Obtenu : {len(grid_questions.get_active_questions())}")
    _test("get_active_questions(na_modules=['B.2']) renvoie 10 questions (12-2, CC-V4-11)",
          len(grid_questions.get_active_questions(na_modules=["B.2"])) == 10,
          f"Obtenu : {len(grid_questions.get_active_questions(na_modules=['B.2']))}")

    # --- DOCUMENT_TYPES / RESPONSE_VALUES / QUALIFYING_FLAGS (nouveau V4) ---
    _test("DOCUMENT_TYPES : 4 types", len(grid_questions.DOCUMENT_TYPES) == 4,
          f"Obtenu : {len(grid_questions.DOCUMENT_TYPES)}")
    _test("DOCUMENT_TYPES[1] : reading_mode='instruction'",
          grid_questions.DOCUMENT_TYPES[1]["reading_mode"] == "instruction",
          f"Obtenu : {grid_questions.DOCUMENT_TYPES[1]['reading_mode']!r}")
    _test("DOCUMENT_TYPES[3] : proof_forms=3",
          grid_questions.DOCUMENT_TYPES[3]["proof_forms"] == 3,
          f"Obtenu : {grid_questions.DOCUMENT_TYPES[3]['proof_forms']}")
    _test("RESPONSE_VALUES = {OUI, NON, INCONNU, NA}",
          grid_questions.RESPONSE_VALUES == {"OUI", "NON", "INCONNU", "NA"},
          f"Obtenu : {grid_questions.RESPONSE_VALUES}")
    _test("QUALIFYING_FLAGS contient ALLEGATION_NON_CONFIRMEE et NON_ATTESTE",
          {"ALLEGATION_NON_CONFIRMEE", "NON_ATTESTE"} <= set(grid_questions.QUALIFYING_FLAGS),
          f"Obtenu : {list(grid_questions.QUALIFYING_FLAGS)}")

    # --- CC-07 : MITIGATION_STATUTS / CONCESSIVE_MARKERS non supprimés ---
    _test("MITIGATION_STATUTS toujours présent (4 statuts, CC-07 non régressé)",
          len(grid_questions.MITIGATION_STATUTS) == 4,
          f"Obtenu : {len(grid_questions.MITIGATION_STATUTS)}")
    _test("CONCESSIVE_MARKERS toujours présent (CC-07 non régressé)",
          len(grid_questions.CONCESSIVE_MARKERS) > 0)
    _test("OUI_DEFAILLANTE : description mentionne la mitigation interrompue",
          "interrompue" in grid_questions.MITIGATION_STATUTS["OUI_DEFAILLANTE"]["description"])


# ============================================================================
# 2. TESTS D'INTÉGRATION
# ============================================================================

def test_integration():
    print("\n--- 2. Tests d'intégration ---\n")

    try:
        import search
        from model import compute_grade
        from analyze import analyze

        model, index, metadata = search.load_search_components()

        # 2.1 Flag scores
        test_text = """
        Community opposition around the project. Grievances filed regarding
        involuntary resettlement. Indigenous peoples rights not addressed.
        """
        scores = search.get_flag_scores(test_text, model, index, metadata)
        _test("Flag scores → 3 valeurs",  len(scores) == 3)
        _test("Flag scores 0–100",
              all(0 <= v <= 100 for v in scores.values()),
              f"Scores : {scores}")
        _test("Texte communautaire → flag1 domine",
              scores["flag1_community"] >= scores["flag2_pollution"],
              f"f1={scores['flag1_community']:.1f} vs f2={scores['flag2_pollution']:.1f}",
              warning_only=True)

        # 2.2 Grade ESG (compute_grade, remplace le Cox)
        pred = compute_grade(scores)
        expected_keys = ["risk_score", "risk_label", "risk_grade"]
        _test("Prédiction toutes les clés",
              all(k in pred for k in expected_keys))
        _test("Score ∈ [0, 100]",
              0 <= pred["risk_score"] <= 100,
              f"score = {pred['risk_score']}")
        _test("Grade valide",
              pred["risk_grade"] in ["A", "B", "C", "D"],
              f"grade = {pred['risk_grade']}")

        # 2.4 Pipeline complète
        t0 = time.time()
        result = analyze(test_text)
        dt = time.time() - t0
        expected_result_keys = [
            "flag_scores", "prediction", "similar_passages",
            "detected_signals", "signal_spans", "processing_time_s",
        ]
        _test("analyze() toutes les clés",
              all(k in result for k in expected_result_keys))
        _test(f"analyze() < 45s (actuel: {dt:.1f}s)",
              dt < 45,
              "Trop lent → voir checklist.md, chantier perf LLM/deep_analysis",
              warning_only=True)

    except Exception as e:
        print(f"  ❌ Erreur d'intégration : {e}")
        import traceback
        traceback.print_exc()

    # 2.5 Branchement chunks -> Grille V4 (directive CC-V4-09) — bloc séparé
    # du try ci-dessus : un échec de l'ANCIEN pipeline (corpus/FAISS
    # manquants) ne doit pas masquer ce test, qui ne dépend que de
    # search.chunk_text() (pas du corpus) et d'un backend LLM monkey-patché.
    try:
        test_integration_chunks_to_grid()
        test_integration_section_exclusion_real_chunking()
        test_page_marker_extraction()
    except Exception as e:
        print(f"  ❌ Erreur d'intégration (chunks -> Grille V4) : {e}")
        import traceback
        traceback.print_exc()


def test_integration_chunks_to_grid():
    """Vérifie que le pipeline complet (search.chunk_text -> grid_analyze)
    fonctionne sur un texte réel (directive CC-V4-09). Backend LLM
    monkey-patché (même idiome que 1.5/1.15) — pas d'appel réseau réel,
    et le flag GRID_V4_ENABLED est restauré après le test (isolation)."""
    print("\n--- 2.5a Branchement chunks -> Grille V4 (search.chunk_text réel) ---\n")

    import json
    import config
    import llm_backend
    import search
    import grid_analyze

    # Texte synthétique simulant un ESRS — cf. directive CC-V4-09 (reprend
    # l'exemple fourni telle quelle : assez court pour tenir dans un seul
    # chunk de search.chunk_text(), cf. test_integration_section_exclusion_
    # real_chunking ci-dessous pour un texte assez long pour vérifier la
    # séparation ESAP/plaintes IFC sur plusieurs chunks).
    test_text = """
    Environmental and Social Review Summary
    The project involves the expansion of mining operations.
    There are significant ongoing resettlement impacts.

    Environmental and Social Action Plan
    Item 1: Prepare a Livelihood Restoration Plan
    Item 2: Finalize the compensation report

    Information on E&S Complaints and Grievance Mechanisms
    Affected communities have unrestricted access to the CAO.
    """

    chunks = search.chunk_text(test_text)
    _test("search.chunk_text() sur un texte réel produit au moins 1 chunk",
          len(chunks) >= 1, f"Obtenu : {len(chunks)} chunk(s)")

    chunks_for_grid = [{"text": c, "page": None} for c in chunks]

    def _fake_dispatch(backend, prompt, model, options, timeout, response_format=None):
        # CC-V4-12 : Passe 1 (extraction) uniquement — found=False partout,
        # la Passe 2 n'est donc jamais atteinte pour ce test.
        return json.dumps({
            "code": "X", "found": False, "verbatim": None, "page": None,
            "subject": None, "brief": "rien trouve",
        })

    _orig_enabled = config.GRID_V4_ENABLED
    _orig_dispatch = llm_backend._dispatch
    config.GRID_V4_ENABLED = True
    llm_backend._dispatch = _fake_dispatch
    try:
        result = grid_analyze.analyze_grid(chunks_for_grid, document_type=1)

        _test("Pipeline complet : résultat non None", result is not None)
        if result is not None:
            _test("Pipeline complet : grid_version='V4'", result["grid_version"] == "V4",
                  f"Obtenu : {result.get('grid_version')!r}")
            _test("Pipeline complet : 12 questions", len(result["questions"]) == 12,
                  f"Obtenu : {len(result['questions'])}")
            _test("Pipeline complet : score >= 0", result["scoring"]["score"] >= 0,
                  f"Obtenu : {result['scoring']['score']}")
    finally:
        llm_backend._dispatch = _orig_dispatch
        config.GRID_V4_ENABLED = _orig_enabled


def test_integration_section_exclusion_real_chunking():
    """Vérifie que l'exclusion ESAP / plaintes IFC (grid_sections.py)
    fonctionne quand ESAP et la section plaintes tombent dans des chunks
    SÉPARÉS produits par le vrai search.chunk_text() — cf. directive
    CC-V4-09, "s'assurer que les exclusions fonctionnent sur un vrai
    document". Le texte de la directive (~50 mots) tient dans un seul
    chunk de search.chunk_text() (fenêtre de 175 mots) : ESAP et plaintes
    s'y retrouvent mélangés dans le MÊME chunk, ce qui ne teste pas la
    séparation. Ce test utilise un texte assez long (>500 mots, sections
    espacées de >175 mots) pour que le chunking réel les place dans des
    chunks distincts, condition réaliste sur un rapport de 45-70 pages.
    Purement déterministe (search.chunk_text + grid_sections), pas de LLM.
    """
    print("\n--- 2.5b Exclusion de sections sur chunking réel (search.chunk_text) ---\n")

    import search
    import grid_sections

    def _pad(label, n):
        return " ".join(f"{label}word{i}" for i in range(n))

    test_text = f"""
    Environmental and Social Review Summary. {_pad("context", 220)}

    Environmental and Social Action Plan
    Item 1: Prepare a Livelihood Restoration Plan for the affected households.
    Item 2: Finalize the compensation report for past resettlement.
    Item 3: Complete establishment of a data management system.

    {_pad("monitoring", 220)}

    Information on E&S Complaints and Grievance Mechanisms
    Affected communities have unrestricted access to the Compliance Advisor Ombudsman (CAO).
    The CAO is mandated to address complaints from people affected by IFC projects.

    {_pad("closing", 100)}
    """

    chunks = search.chunk_text(test_text)
    _test("Texte réaliste (~500+ mots) -> plusieurs chunks",
          len(chunks) > 1, f"Obtenu : {len(chunks)} chunk(s)")

    classified = grid_sections.classify_chunks([{"text": c, "page": None} for c in chunks])

    esap_tags = [c for c in classified if c["section_tag"] == "esap"]
    complaints_tags = [c for c in classified if c["section_tag"] == "ifc_complaints"]
    untagged = [c for c in classified if c["section_tag"] is None]

    _test("Au moins un chunk taggé 'esap' sur chunking réel", len(esap_tags) > 0,
          f"Obtenu : {len(esap_tags)}")
    _test("Au moins un chunk taggé 'ifc_complaints' sur chunking réel", len(complaints_tags) > 0,
          f"Obtenu : {len(complaints_tags)}")
    _test("Au moins un chunk normal (non taggé) sur chunking réel", len(untagged) > 0,
          f"Obtenu : {len(untagged)}")
    _test("ESAP et plaintes IFC tombent dans des chunks DISTINCTS (pas mélangés)",
          not any(c in esap_tags for c in complaints_tags))

    # Vérifie que le filtrage en aval (get_chunks_for_question) exclut
    # bien ces chunks une fois issus d'un vrai chunking, pas seulement
    # sur les petits exemples synthétiques de CC-09/CC-V4-04.
    r_chunks = grid_sections.get_chunks_for_question(classified, "B.1.2")
    _test("Question R : chunks plaintes IFC exclus (chunking réel)",
          all(c["section_tag"] != "ifc_complaints" for c in r_chunks))
    _test("Question R : chunks ESAP conservés (chunking réel)",
          any(c["section_tag"] == "esap" for c in r_chunks))

    a_chunks = grid_sections.get_chunks_for_question(classified, "B.1.2", for_mitigation=True)
    _test("Question A (mitigation) : chunks ESAP exclus (chunking réel)",
          all(c["section_tag"] != "esap" for c in a_chunks))
    _test("Question A (mitigation) : chunks plaintes IFC exclus (chunking réel)",
          all(c["section_tag"] != "ifc_complaints" for c in a_chunks))


def test_page_marker_extraction():
    """Diagnostic "page ?" Aysha Wind (2026-08-20) : evidence_r["page"]
    dépendait d'un numéro de page que le LLM devinait depuis des
    marqueurs *ambiants* (en-têtes/pieds de page qu'un PDF donné avait —
    ou non — préservés comme texte extractible), d'où "page ?" sur des
    documents sans ces marqueurs (ESAP tabulaire) même quand des rapports
    institutionnels (Mundra) fonctionnaient par chance. Fix : app.py::
    _extract_uploaded_text() insère désormais un marqueur [PAGE:N] avant
    chaque page PDF, explicite et indépendant de la mise en page source.

    Ce test simule le texte marqué que produirait _extract_uploaded_text()
    (app.py n'est pas importable ici — module Streamlit avec des appels
    st.* au niveau module, cf. absence d'"import app" ailleurs dans ce
    fichier) et vérifie 4 choses : (a) le marqueur survit intact au vrai
    search.chunk_text() (ni avalé par le filtre boilerplate, ni cassé par
    la tokenisation mot-à-mot), (b) grid_prompts.get_extraction_prompt()
    documente la consigne de lecture du marqueur, (c) sur le pipeline
    complet (search.chunk_text -> grid_analyze réel, LLM monkey-patché),
    si le LLM lit correctement le marqueur, la page atteint bien
    evidence_r["page"], (d) evidence_a["page"] (mitigation) hérite de
    evidence_r["page"] plutôt que de rester None — cf. diagnostic
    "mitigation sans page" (2026-08-20) : la Passe 2 (qualification) ne
    voit jamais le marqueur ni les chunks bruts, uniquement le verbatim
    déjà extrait par la Passe 1, donc structurellement la même page.
    Rien de tout ceci ne prouve qu'un LLM réel suivra la consigne de
    lecture du marqueur — seulement que la donnée n'est plus jetée en
    route une fois qu'il la fournit (cf. limite déjà documentée dans
    grid_prompts._PAGE_RULE)."""
    print("\n--- 2.5c Marqueur [PAGE:N] (diagnostic page '?' Aysha Wind) ---\n")

    import json
    import config
    import llm_backend
    import search
    import grid_prompts
    import grid_analyze

    def _pad(label, n):
        return " ".join(f"{label}word{i}" for i in range(n))

    # Simule le texte que _extract_uploaded_text() produit sur un PDF de
    # 3 pages (marqueur inséré avant le texte de chaque page).
    test_text = f"""[PAGE:1]
    {_pad("intro", 150)}

    [PAGE:2]
    A strike occurred on site in March 2025 involving unresolved labor disputes.
    {_pad("context", 150)}

    [PAGE:3]
    {_pad("closing", 150)}
    """

    # --- (a) Le marqueur survit au vrai chunking ---
    chunks = search.chunk_text(test_text)
    _test("Texte marqué [PAGE:N] -> au moins 1 chunk (pas tout filtré comme boilerplate)",
          len(chunks) >= 1, f"Obtenu : {len(chunks)} chunk(s)")
    _test("Marqueur [PAGE:1] survit au chunking (search.chunk_text réel)",
          any("[PAGE:1]" in c for c in chunks))
    _test("Marqueur [PAGE:2] survit au chunking, non cassé par la tokenisation mot-à-mot",
          any("[PAGE:2]" in c for c in chunks))

    # --- (b) Le prompt d'extraction documente la consigne de lecture ---
    prompt = grid_prompts.get_extraction_prompt("A.1.1", ["[PAGE:2]\nA strike occurred on site."])
    _test("Prompt d'extraction : marqueur [PAGE:N] documenté", "[PAGE:N]" in prompt)
    _test("Prompt d'extraction : consigne d'utiliser le marqueur le plus proche",
          "le plus proche" in prompt)
    _test("Prompt d'extraction : consigne d'exclure le marqueur du verbatim",
          "N'inclus JAMAIS le marqueur" in prompt)

    # --- (c) Pipeline complet : la page circule jusqu'à evidence_r si le
    # LLM la renvoie correctement (plomberie de bout en bout). ---
    chunks_for_grid = [{"text": c, "page": None} for c in chunks]

    def _fake_dispatch(backend, prompt, model, options, timeout, response_format=None):
        if "VERBATIM EXTRAIT" in prompt:
            # Passe 2 (qualification) — status=OUI avec une mesure de
            # mitigation, pour vérifier que evidence_a["page"] hérite de
            # evidence_r["page"] (diagnostic "mitigation sans page",
            # 2026-08-20) : la Passe 2 ne voit jamais le marqueur
            # [PAGE:N], seulement le verbatim déjà extrait par la Passe 1.
            return json.dumps({
                "code": "A.1.1", "status": "OUI", "confidence": "HIGH",
                "mitigation_status": "OUI_PROUVEE",
                "verbatim_r": "A strike occurred on site in March 2025.",
                "verbatim_a_mesure": "A mediation agreement was signed with local unions.",
                "verbatim_a_defaillance": None,
                "brief_r": "greve confirmee", "brief_a": "accord signe",
            })
        return json.dumps({
            "code": "A.1.1", "found": True,
            "verbatim": "A strike occurred on site in March 2025 involving unresolved labor disputes.",
            "page": 2, "subject": "SPV", "brief": "greve documentee",
        })

    _orig_enabled = config.GRID_V4_ENABLED
    _orig_dispatch = llm_backend._dispatch
    config.GRID_V4_ENABLED = True
    llm_backend._dispatch = _fake_dispatch
    try:
        result = grid_analyze.analyze_grid(chunks_for_grid, document_type=1)
        _test("Pipeline complet (chunking réel) : résultat non None", result is not None)
        if result is not None:
            q = next(q for q in result["questions"] if q["code"] == "A.1.1")
            _test("Pipeline complet : evidence_r['page'] renseigné (fini le 'page ?' systématique)",
                  q["evidence_r"] is not None and q["evidence_r"]["page"] == 2,
                  f"Obtenu : {q.get('evidence_r')!r}")
            _test("Pipeline complet : evidence_a['page'] hérite de evidence_r['page'] (fini 'mitigation sans page')",
                  q["evidence_a"] is not None and q["evidence_a"]["page"] == 2,
                  f"Obtenu : {q.get('evidence_a')!r}")
    finally:
        llm_backend._dispatch = _orig_dispatch
        config.GRID_V4_ENABLED = _orig_enabled


# ============================================================================
# 3. TESTS DE COHÉRENCE MÉTIER (bloc 3.2 de la roadmap)
# ============================================================================

def test_business():
    print("\n--- 3. Tests de cohérence métier ---\n")

    try:
        from analyze import analyze

        # --- CAS 1 : Community Opposition (haut risque) ---
        text_community = """
        The project has faced significant community opposition. Local populations
        report involuntary displacement without adequate compensation. Multiple
        grievances filed through the CAO mechanism. Indigenous communities claim
        violation of FPIC. Stakeholder engagement described as inadequate.
        """

        # --- CAS 2 : ESAP Delays (risque moyen) ---
        text_esap = """
        The ESAP action plan shows delays in implementation of corrective measures.
        Several environmental monitoring commitments are overdue. Non-compliance
        with PS1 requirements on assessment and management. Pollution discharge
        thresholds exceeded on two occasions. Remediation plan not yet submitted.
        """

        # --- CAS 3 : Biodiversity Risk (haut risque flag3) ---
        text_biodiversity = """
        The project area overlaps with a critical habitat for endangered species.
        PS6 requirements for biodiversity management have not been fully met.
        No adequate biodiversity offset program established. Environmental impact
        assessment did not properly identify all species at risk. Habitat
        fragmentation concerns raised by independent environmental review.
        """

        # --- CAS 4 : Projet propre (bas risque) ---
        text_clean = """
        The project's environmental and social management system is fully operational.
        Regular monitoring reports demonstrate compliance with all Performance Standards.
        Community feedback mechanisms are active with positive stakeholder sentiment.
        Biodiversity offset program is on track. ESAP actions completed ahead of schedule.
        """

        # Analyser chaque cas
        print("  Analyse des 4 cas de test...")
        r1 = analyze(text_community)
        r2 = analyze(text_esap)
        r3 = analyze(text_biodiversity)
        r4 = analyze(text_clean)

        # Résumé visuel (CHANTIER SIMPLIFICATION PIPELINE, 2026-08-08 : plus de
        # probability_12m/Cox — risk_score = max(flag_scores), 0-100)
        print(f"\n  Cas 1 (Community Opposition) : {r1['prediction']['risk_grade']} "
              f"({r1['prediction']['risk_label']}) — score={r1['prediction']['risk_score']}/100")
        print(f"  Cas 2 (ESAP Delays)          : {r2['prediction']['risk_grade']} "
              f"({r2['prediction']['risk_label']}) — score={r2['prediction']['risk_score']}/100")
        print(f"  Cas 3 (Biodiversity Risk)    : {r3['prediction']['risk_grade']} "
              f"({r3['prediction']['risk_label']}) — score={r3['prediction']['risk_score']}/100")
        print(f"  Cas 4 (Projet propre)        : {r4['prediction']['risk_grade']} "
              f"({r4['prediction']['risk_label']}) — score={r4['prediction']['risk_score']}/100")

        # Vérifications
        _test("Cas 1 (risque) > Cas 4 (propre)",
              r1["prediction"]["risk_score"] > r4["prediction"]["risk_score"],
              f"{r1['prediction']['risk_score']} vs {r4['prediction']['risk_score']}")

        # Convention conservée (2026-08-08) : A = pire (Escalade), D = meilleur
        # (Vigilance) — même sens qu'avec le Cox, cf. model.py.
        _test("Cas 1 = grade A ou B",
              r1["prediction"]["risk_grade"] in ["A", "B"],
              f"Grade = {r1['prediction']['risk_grade']}",
              warning_only=True)

        _test("Cas 4 = grade C ou D",
              r4["prediction"]["risk_grade"] in ["C", "D"],
              f"Grade = {r4['prediction']['risk_grade']}",
              warning_only=True)

        _test("Cas 1 : signaux détectés",
              len(r1["detected_signals"]) > 0,
              "Aucun signal — vérifier les mots-clés dans analyze._extract_signals")

        # Ordonnancement global
        scores = [r1["prediction"]["risk_score"],
                  r2["prediction"]["risk_score"],
                  r3["prediction"]["risk_score"],
                  r4["prediction"]["risk_score"]]
        _test("Cas propre a le score le plus bas",
              scores[3] == min(scores),
              f"Scores : community={scores[0]} esap={scores[1]} "
              f"bio={scores[2]} propre={scores[3]}",
              warning_only=True)

    except Exception as e:
        print(f"  ❌ Erreur tests métier : {e}")
        import traceback
        traceback.print_exc()


# ============================================================================
# RÉSUMÉ + MAIN
# ============================================================================

def print_summary():
    total = _passed + _failed + _warnings
    print("\n" + "=" * 60)
    print(f"RÉSULTATS : {_passed}/{total} passés, {_failed} échecs, {_warnings} warnings")
    if _failed == 0:
        print("🎉 Tous les tests critiques passent !")
    else:
        print("🔴 Des tests critiques ont échoué — à corriger avant la démo")
    print("=" * 60)


if __name__ == "__main__":
    args = sys.argv[1:]
    any_tier_flag = "--unit" in args or "--integ" in args or "--business" in args

    # BUG CORRIGÉ (2026-08-20) : c'était un elif, pas 3 `if` indépendants —
    # `python scripts/test.py --unit --integ --business` (commande
    # documentée dans CLAUDE.md, "avant chaque commit") n'exécutait donc
    # QUE --unit, --integ et --business n'étaient jamais lancés malgré la
    # ligne de commande. Seul `python scripts/test.py` sans flag (branche
    # `else` ci-dessous) lançait vraiment les 3 tiers.
    if "--unit" in args:
        test_unit()
    if "--integ" in args:
        test_integration()
    if "--business" in args:
        test_business()
    if not any_tier_flag:
        # Tout lancer (aucun flag reconnu passé)
        test_unit()
        test_integration()
        test_business()

    print_summary()